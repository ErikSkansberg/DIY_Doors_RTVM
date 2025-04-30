import sys
import subprocess

def install_and_import(package_name, import_name=None):
    """
    Attempts to import the module with name `import_name` (or package_name if not provided).
    If the import fails, the package is installed via pip and then imported.
    """
    import_name = import_name or package_name
    try:
        __import__(import_name)
    except ImportError:
        print(f"Package '{import_name}' not found. Installing '{package_name}'...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        try:
            __import__(import_name)
        except ImportError:
            print(f"Failed to install and import '{package_name}'. Please install it manually.")
            sys.exit(1)

# List of required packages along with any custom import names.
# For example, "pyspellchecker" is imported as "spellchecker" in your code.
required_packages = [
    ("pandas", None),
    ("numpy", None),
    ("matplotlib", None),
    ("openpyxl", None),
    ("pyspellchecker", "spellchecker"),
    ("reportlab", None),
    ("pdfrw", None),
]

# Install and import each package.
for pkg, imp_name in required_packages:
    install_and_import(pkg, import_name=imp_name)

# Now you can proceed with your other imports.
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import re
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('TkAgg')
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import Counter
from openpyxl import load_workbook
from spellchecker import SpellChecker

# Further down in your code, you'll also be using:
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import acroform
from pdfrw import PdfReader


class PatternDialog(tk.Toplevel):
    def __init__(self, master, app, obj_identifier, di_number, current_row):
        super().__init__(master)
        self.title("Pattern Generator")
        self.app = app  # Reference to the RTVMApp instance
        self.current_row = current_row  # Store the current row index

        # Initialize variables
        self.obj_identifier = obj_identifier
        self.di_number = di_number
        self.deletions = []
        
        # Set window size for better usability
        self.geometry("650x600")
        
        # Add a help frame with key guidelines from the RTVM Desk Guide
        self.help_frame = tk.LabelFrame(self, text="RTVM Formatting Guidelines")
        self.help_frame.grid(row=0, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        
        help_text = (
            "• NEVER USE SEMICOLONS (;) except as delimiters between elements\n"
            "• Use commas (,) between location elements\n"
            "• Capitalize The First Letter Of Each Word in location\n" 
            "• Avoid abbreviations (use 'Page' not 'PG', 'Sheet' not 'SHT')\n"
            "• For unknown Sheet/Page/Plan elements, use [] as placeholders\n"
            "• Format: ADD;DI-Number;CDRL Name, Page X, Plan View Y;Status\n"
            "• Valid statuses: SAT, UNSAT (don't use TBD)\n"
            "• Use DEL;WCC-VERI-DOC-XXXX to request deletion"
        )
        
        self.help_label = tk.Label(self.help_frame, text=help_text, justify="left", anchor="w")
        self.help_label.pack(fill="x", padx=5, pady=5)
        
        # Toggle button to show/hide guidelines
        self.help_visible = tk.BooleanVar(value=True)
        self.toggle_help_button = tk.Button(self.help_frame, text="Hide Guidelines", command=self.toggle_help)
        self.toggle_help_button.pack(side="right", padx=5, pady=2)

        # Object Identifier
        self.obj_identifier_label = tk.Label(self, text="Object Identifier")
        self.obj_identifier_label.grid(row=1, column=0, sticky="e", padx=(10, 5), pady=5)
        self.obj_identifier_entry = tk.Entry(self)
        self.obj_identifier_entry.grid(row=1, column=1, columnspan=2, sticky="w", padx=5, pady=5)
        self.obj_identifier_entry.insert(0, self.obj_identifier)

        # DI Number
        self.di_number_label = tk.Label(self, text="DI Number")
        self.di_number_label.grid(row=2, column=0, sticky="e", padx=(10, 5), pady=5)
        self.di_number_entry = tk.Entry(self)
        self.di_number_entry.grid(row=2, column=1, columnspan=2, sticky="w", padx=5, pady=5)
        self.di_number_entry.insert(0, self.di_number)

        # CDRL Name
        self.cdrl_name_label = tk.Label(self, text="(4) CDRL File Name")
        self.cdrl_name_label.grid(row=3, column=0, sticky="e", padx=(10, 5), pady=5)
        self.cdrl_name_entry = tk.Entry(self, width=30)
        self.cdrl_name_entry.grid(row=3, column=1, columnspan=2, sticky="w", padx=5, pady=5)
        
        # Add pattern template suggestion
        variant_frame = tk.Frame(self)
        variant_frame.grid(row=4, column=0, columnspan=3, sticky="w", padx=10, pady=2)
        
        variant_label = tk.Label(variant_frame, text="Variant Template:")
        variant_label.pack(side="left", padx=(0, 5))
        
        self.variant_160_button = tk.Button(variant_frame, text="160-WLIC", command=lambda: self.insert_variant_template("160-WLIC"))
        self.variant_160_button.pack(side="left", padx=5)
        
        self.variant_180_button = tk.Button(variant_frame, text="180-WLR", command=lambda: self.insert_variant_template("180-WLR"))
        self.variant_180_button.pack(side="left", padx=5)

        # Detailed Location - Page/Sheet
        self.page_sheet_label = tk.Label(self, text="(4) Page/Sheet")
        self.page_sheet_label.grid(row=5, column=0, sticky="e", padx=(10, 5), pady=5)
        self.page_sheet_option_var = tk.StringVar(self)
        self.page_sheet_option_var.set("Page")  # default value
        self.page_sheet_option_menu = ttk.Combobox(
            self, textvariable=self.page_sheet_option_var,
            values=["Page", "Sheet"], width=8, state="readonly")
        self.page_sheet_option_menu.grid(row=5, column=1, sticky="w", padx=5, pady=5)
        self.page_sheet_entry = tk.Entry(self, width=10)
        self.page_sheet_entry.grid(row=5, column=2, sticky="w", padx=5, pady=5)

        # Detailed Location - Plan View/Section
        self.plan_view_label = tk.Label(self, text="(4) Plan View/Section")
        self.plan_view_label.grid(row=6, column=0, sticky="e", padx=(10, 5), pady=5)
        self.plan_view_option_var = tk.StringVar(self)
        self.plan_view_option_var.set("Plan View")  # default value
        self.plan_view_option_menu = ttk.Combobox(
            self, textvariable=self.plan_view_option_var,
            values=["Plan View", "Section"], width=8, state="readonly")
        self.plan_view_option_menu.grid(row=6, column=1, sticky="w", padx=5, pady=5)
        self.plan_view_entry = tk.Entry(self, width=10)
        self.plan_view_entry.grid(row=6, column=2, sticky="w", padx=5, pady=5)

        # Contractor Assessed Status
        self.status_label = tk.Label(self, text="(5) Contractor Assessed Status")
        self.status_label.grid(row=7, column=0, sticky="e", padx=(10, 5), pady=5)
        self.status_var = tk.StringVar(self)
        self.status_var.set("")  # Default to blank
        self.status_dropdown = ttk.Combobox(
            self, textvariable=self.status_var, values=["SAT", "UNSAT"], 
            state="readonly", width=8)
        self.status_dropdown.grid(row=7, column=1, columnspan=2, sticky="w", padx=5, pady=5)
        
        # Add warning label about TBD status
        self.status_warning = tk.Label(
            self, 
            text="Note: According to the RTVM Desk Guide, don't use TBD status in submissions.", 
            fg="red", 
            font=("Helvetica", 9, "italic"))
        self.status_warning.grid(row=8, column=0, columnspan=3, sticky="w", padx=10, pady=2)

        # Pattern Type Selection Frame
        pattern_type_frame = tk.LabelFrame(self, text="Pattern Type")
        pattern_type_frame.grid(row=9, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        
        self.pattern_type_var = tk.StringVar(value="update_existing")
        
        # Update existing VeriDoc
        self.update_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Update Existing VeriDoc", 
            variable=self.pattern_type_var, 
            value="update_existing",
            command=self.toggle_pattern_type)
        self.update_radio.grid(row=0, column=0, sticky="w", padx=5, pady=2)
        
        # Add new CDRL
        self.add_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Add New CDRL", 
            variable=self.pattern_type_var, 
            value="add_new",
            command=self.toggle_pattern_type)
        self.add_radio.grid(row=0, column=1, sticky="w", padx=5, pady=2)
        
        # Delete existing VeriDoc
        self.delete_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Delete Existing VeriDoc", 
            variable=self.pattern_type_var, 
            value="delete",
            command=self.toggle_pattern_type)
        self.delete_radio.grid(row=0, column=2, sticky="w", padx=5, pady=2)
        
        # Buttons frame
        self.button_frame = tk.Frame(self)
        self.button_frame.grid(row=10, column=0, columnspan=3, pady=5)

        # Generate Button
        self.generate_button = tk.Button(
            self.button_frame, text="(6) Generate Pattern", command=self.generate_pattern)
        self.generate_button.grid(row=0, column=0, padx=5, pady=5)

        # Copy Button
        self.copy_button = tk.Button(
            self.button_frame, text="(7) Copy to Clipboard", command=self.copy_to_clipboard)
        self.copy_button.grid(row=0, column=1, padx=5, pady=5)

        # Save to Excel Button
        self.save_button = tk.Button(
            self.button_frame, text="Save Generated Pattern to Excel", command=self.save_to_excel)
        self.save_button.grid(row=0, column=2, padx=5, pady=5)

        # Reset Button
        self.reset_button = tk.Button(
            self.button_frame, text="Reset", command=self.reset_fields)
        self.reset_button.grid(row=0, column=3, padx=5, pady=5)

        # New Button to create the 180-Vessel Version
        self.create_180_button = tk.Button(
            self.button_frame, text="Also create a 180-Vessel Version", command=self.create_180_version)
        self.create_180_button.grid(row=1, column=0, columnspan=4, padx=5, pady=5)

        # Preview Label
        self.preview_label = tk.Label(self, text="Pattern Preview:")
        self.preview_label.grid(row=11, column=0, sticky="w", padx=10, pady=(10, 2))
        
        # Generated Pattern
        self.output_label = tk.Label(self, text="Generated Pattern: For column G")
        self.output_label.grid(row=12, column=0, columnspan=3, sticky="w", padx=10, pady=(2, 2))
        self.output_text = tk.Text(self, height=6, width=70)
        self.output_text.grid(row=13, column=0, columnspan=3, padx=10, pady=(2, 10), sticky="ew")
        
        # Add example patterns from the RTVM Desk Guide
        example_frame = tk.LabelFrame(self, text="Example Patterns from RTVM Desk Guide")
        example_frame.grid(row=14, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        
        examples_text = (
            "Update Existing: WCC-VERI-DOC-169;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;SAT\n"
            "Add New CDRL: ADD;070-001;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT\n"
            "Delete VeriDoc: DEL;WCC-VERI-DOC-169"
        )
        
        examples_label = tk.Label(example_frame, text=examples_text, justify="left", font=("Courier", 9))
        examples_label.pack(fill="x", padx=5, pady=5)

        # Configure grid weights for proper resizing
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        
        # Initialize by calling toggle_pattern_type once
        self.toggle_pattern_type()


    def toggle_help(self):
        """Toggle the visibility of the help guidelines."""
        if self.help_visible.get():
            self.help_label.pack_forget()
            self.help_visible.set(False)
            self.toggle_help_button.config(text="Show Guidelines")
        else:
            self.help_label.pack(fill="x", padx=5, pady=5)
            self.help_visible.set(True)
            self.toggle_help_button.config(text="Hide Guidelines")
            
    def insert_variant_template(self, variant):
        """Insert variant template into CDRL Name field."""
        variant_format = f"{variant}_WCC_{self.di_number.replace('-', '_')}"
        self.cdrl_name_entry.delete(0, tk.END)
        self.cdrl_name_entry.insert(0, variant_format)

    def toggle_history_tables(self):
        if self.show_history_var.get() == 1:
            # Show the history tables and labels
            self.comment_history_label.grid()
            self.comment_history_table.grid()
            self.gov_comment_history_label.grid()
            self.gov_comment_history_table.grid()
            # Adjust column weights
            self.root.grid_columnconfigure(3, weight=1)
            self.root.grid_columnconfigure(4, weight=1)
        else:
            # Hide the history tables and labels
            self.comment_history_label.grid_remove()
            self.comment_history_table.grid_remove()
            self.gov_comment_history_label.grid_remove()
            self.gov_comment_history_table.grid_remove()
            # Adjust column weights
            self.root.grid_columnconfigure(3, weight=0)
            self.root.grid_columnconfigure(4, weight=0)
            
    def toggle_pattern_type(self):
        """Enable/disable appropriate fields based on pattern type selection."""
        pattern_type = self.pattern_type_var.get()
        
        # Enable all fields initially
        self.obj_identifier_entry.config(state="normal")
        self.di_number_entry.config(state="normal")
        self.cdrl_name_entry.config(state="normal")
        self.page_sheet_option_menu.config(state="readonly")
        self.page_sheet_entry.config(state="normal")
        self.plan_view_option_menu.config(state="readonly")
        self.plan_view_entry.config(state="normal")
        self.status_dropdown.config(state="readonly")
        self.create_180_button.config(state="normal")
        
        if pattern_type == "update_existing":
            # For updating, we need the VeriDoc number, location details, and status
            self.di_number_entry.config(state="disabled")  # DI number is part of the existing VeriDoc
            
        elif pattern_type == "add_new":
            # For adding, we need DI number, location details, and status
            # VeriDoc number is not needed as it will be assigned by the system
            self.obj_identifier_entry.config(state="disabled")
            
        elif pattern_type == "delete":
            # For deleting, we only need the VeriDoc number
            self.di_number_entry.config(state="disabled")
            self.cdrl_name_entry.config(state="disabled")
            self.page_sheet_option_menu.config(state="disabled")
            self.page_sheet_entry.config(state="disabled")
            self.plan_view_option_menu.config(state="disabled")
            self.plan_view_entry.config(state="disabled")
            self.status_dropdown.config(state="disabled")
            self.create_180_button.config(state="disabled")




    def generate_pattern(self):
        # Clear output_text
        self.output_text.delete("1.0", tk.END)
        patterns = []

        # Check if we can generate the pattern
        can_generate_pattern = True
        error_messages = []

        pattern_type = self.pattern_type_var.get()
        
        if pattern_type != "delete" and not self.status_var.get():
            error_messages.append("Contractor Assessed Status dropdown is blank.")
            can_generate_pattern = False

        if pattern_type != "delete" and not self.page_sheet_entry.get().strip():
            error_messages.append("Page/Sheet input box is blank.")
            can_generate_pattern = False

        if pattern_type != "delete" and not self.plan_view_entry.get().strip():
            error_messages.append("Plan View/Section input box is blank.")
            can_generate_pattern = False
            
        if pattern_type == "update_existing" and not self.obj_identifier_entry.get().strip():
            error_messages.append("Object Identifier is required for updating existing VeriDoc.")
            can_generate_pattern = False
            
        if pattern_type == "add_new" and not self.di_number_entry.get().strip():
            error_messages.append("DI Number is required for adding a new CDRL.")
            can_generate_pattern = False
            
        if pattern_type == "delete" and not self.obj_identifier_entry.get().strip():
            error_messages.append("Object Identifier is required for deletion.")
            can_generate_pattern = False

        if can_generate_pattern:
            obj_identifier = self.obj_identifier_entry.get().strip().upper()
            di_number = self.di_number_entry.get().strip()
            cdrl_name = self.cdrl_name_entry.get().strip().upper()
            page_sheet = self.page_sheet_entry.get().strip()
            page_sheet_type = self.page_sheet_option_var.get()
            plan_view = self.plan_view_entry.get().strip()
            plan_view_type = self.plan_view_option_var.get()
            status = self.status_var.get()

            # Generate the pattern based on the chosen type
            if pattern_type == "update_existing":
                # Format for updating an existing assignment
                detailed_location = f"{cdrl_name}, {page_sheet_type} {page_sheet}, {plan_view_type} {plan_view}"
                self.pattern1 = f"{obj_identifier};{detailed_location};{status}"
                
            elif pattern_type == "add_new":
                # Format for adding a new assignment
                detailed_location = f"{cdrl_name}, {page_sheet_type} {page_sheet}, {plan_view_type} {plan_view}"
                self.pattern1 = f"ADD;{di_number};{detailed_location};{status}"
                
            elif pattern_type == "delete":
                # Format for deleting an existing assignment
                self.pattern1 = f"DEL;{obj_identifier}"

            patterns.append(self.pattern1)
        else:
            if error_messages:
                messagebox.showerror("Error", "\n".join(error_messages))
                return  # Exit the method to avoid adding empty patterns

        # Output patterns
        self.output_text.insert(tk.END, "\n".join(patterns))



    def create_180_version(self):
        # Check if the first pattern has been generated
        if not hasattr(self, 'pattern1'):
            messagebox.showerror("Error", "Please generate the initial pattern first.")
            return
            
        # Check if pattern type is "delete" (180 version doesn't make sense for delete)
        if self.pattern_type_var.get() == "delete":
            messagebox.showerror("Error", "Cannot create 180-Vessel version for a deletion pattern.")
            return

        # Generate the second pattern for 180-WLR variant
        pattern_type = self.pattern_type_var.get()
        obj_identifier = self.obj_identifier_entry.get().upper()
        di_number = self.di_number_entry.get()
        cdrl_name = self.cdrl_name_entry.get().upper()
        page_sheet = self.page_sheet_entry.get()
        page_sheet_type = self.page_sheet_option_var.get()
        plan_view = self.plan_view_entry.get()
        plan_view_type = self.plan_view_option_var.get()
        status = self.status_var.get()

        # Replace 160-WLIC with 180-WLR in the detailed location
        cdrl_name_wlr = cdrl_name.replace("160-WLIC", "180-WLR")
        detailed_location_wlr = f"{cdrl_name_wlr}, {page_sheet_type} {page_sheet}, {plan_view_type} {plan_view}"
        
        if pattern_type == "update_existing":
            messagebox.showinfo("Information", 
                "According to the RTVM Desk Guide, for variant-specific updates, use the ADD format for the second variant.")
            self.pattern2 = f"ADD;{di_number};{detailed_location_wlr};{status}"
        else:  # add_new
            self.pattern2 = f"ADD;{di_number};{detailed_location_wlr};{status}"

        # Append the second pattern to the output
        self.output_text.insert(tk.END, "\n" + self.pattern2)


    def copy_to_clipboard(self):
        # Clear the clipboard
        self.clipboard_clear()

        # Get the content from the output text box
        patterns = self.output_text.get("1.0", tk.END).strip()

        # Copy the content to the clipboard
        self.clipboard_append(patterns)

        # Update the root window to ensure the clipboard retains the copied content
        self.update()

        messagebox.showinfo("Copied", "Pattern copied to clipboard.")


    def reset_fields(self):
        self.cdrl_name_entry.delete(0, tk.END)
        self.page_sheet_entry.delete(0, tk.END)
        self.plan_view_entry.delete(0, tk.END)
        self.status_var.set("")  # Set to blank on reset
        self.output_text.delete("1.0", tk.END)
        self.deletions.clear()
        # Reset to default pattern type
        self.pattern_type_var.set("update_existing")
        self.toggle_pattern_type()

    def save_to_excel(self):
        # Get the generated pattern
        patterns = self.output_text.get("1.0", tk.END).strip()
        if not patterns:
            messagebox.showerror("Error", "No pattern generated to save.")
            return

        # Get existing content from cell G (column index 6) of the current row
        existing_content = self.app.df.iloc[self.current_row, 6]
        if pd.isna(existing_content):
            existing_content = ""
        elif not isinstance(existing_content, str):
            existing_content = str(existing_content)

        # Append the new patterns to the existing content
        if existing_content.strip():
            new_content = existing_content.strip() + "\n" + patterns
        else:
            new_content = patterns

        # Update the Excel file directly using openpyxl
        from openpyxl import load_workbook

        try:
            # Load the workbook
            wb = load_workbook(self.app.excel_file_path)
            ws = wb.active  # You may need to select the correct sheet if there are multiple

            # Calculate the Excel row number (considering headers)
            excel_row = self.current_row + 2  # Assuming header is on the first row

            # Update the cell in column G (which is column index 7 in openpyxl)
            ws.cell(row=excel_row, column=7, value=new_content)

            # Save the workbook
            wb.save(self.app.excel_file_path)

            # Update the DataFrame in memory
            self.app.df.iloc[self.current_row, 6] = new_content

            messagebox.showinfo("Success", "Pattern saved to Excel file successfully.")
            # Update the Contractor Proposed Change Request Input table
            self.app.update_proposed_changes_table()
            
            # Close the pattern dialog window
            self.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel file: {e}")

import os
import queue
import threading
import shutil
from openpyxl import load_workbook
from copy import copy
import pandas as pd
import time
import re
from tkinter import ttk, messagebox, filedialog, simpledialog

class RTVMApp:
    def __init__(self, root):
        self.root = root
        root.title("RTVM Pattern Generator")

        # Initialize variables
        self.current_row = 0
        self.df = None
        self.deletions = []
        self.excel_file_path = ""  # Store the path to the Excel file
        self.pattern_dialog = None  # Reference to PatternDialog instance
        self.selected_base_path = None  # Will store user-selected base directory path
        # Add this line to initialize self.current_comments
        self.current_comments = {}  # To store comments for the current row

        # Initialize unique statuses
        self.unique_object_statuses = set()
        self.unique_contractor_statuses = set()
        self.unique_government_statuses = set()

        # Initialize filter variables
        self.filtered_row_indices = []
        self.current_filtered_index = 0

        # Adjust the window size if needed
        root.geometry("1600x600")  # Width x Height

        # Up and Down Buttons at the top along with Upload button
        self.button_frame_top = tk.Frame(root)
        self.button_frame_top.grid(row=0, column=0, columnspan=10, pady=10, sticky="ew")

        # Up and Down Buttons
        self.up_button = tk.Button(
            self.button_frame_top, text="Up", command=lambda: self.navigate_cells('up'))
        self.up_button.grid(row=0, column=0, padx=5, pady=5)

        self.down_button = tk.Button(
            self.button_frame_top, text="Down", command=lambda: self.navigate_cells('down'))
        self.down_button.grid(row=0, column=1, padx=5, pady=5)


        # Management Button
        self.management_button = tk.Button(
            self.button_frame_top, text="Management", command=self.open_management_window)
        self.management_button.grid(row=0, column=10, padx=5, pady=5)
        # Add the Tools Button
        self.tools_button = tk.Button(
             self.button_frame_top, text="Tools", command=self.open_tools_menu)
        self.tools_button.grid(row=0, column=11, padx=5, pady=5)

        # Jump to Cell Entry
        self.jump_to_label = tk.Label(self.button_frame_top, text="Jump to Row:")
        self.jump_to_label.grid(row=0, column=2, padx=5, pady=5)

        self.jump_to_var = tk.StringVar()
        self.jump_to_entry = tk.Entry(
            self.button_frame_top, textvariable=self.jump_to_var, width=5)
        self.jump_to_entry.grid(row=0, column=3, padx=5, pady=5)

        self.go_button = tk.Button(
            self.button_frame_top, text="Go", command=self.jump_to_cell)
        self.go_button.grid(row=0, column=4, padx=5, pady=5)

        # Upload Excel File Button on the same row
        self.upload_button = tk.Button(
            self.button_frame_top, text="Upload Excel File", command=self.upload_excel_file)
        self.upload_button.grid(row=0, column=5, padx=10, pady=5)

        # Apply Filter Button
        self.apply_filter_button = tk.Button(
            self.button_frame_top, text="Apply Filter", command=self.open_filter_dialog)
        self.apply_filter_button.grid(row=0, column=6, padx=5, pady=5)

        # Clear Filter Button
        self.clear_filter_button = tk.Button(
            self.button_frame_top, text="Clear Filter", command=self.clear_filters)
        self.clear_filter_button.grid(row=0, column=7, padx=5, pady=5)

        # Create Pie Charts Button
        self.create_pie_charts_button = tk.Button(
            self.button_frame_top, text="Create Pie Charts", command=self.open_pie_chart_window)
        self.create_pie_charts_button.grid(row=0, column=8, padx=5, pady=5)

        # Show/Hide History Tables Checkbox
        self.show_history_var = tk.IntVar(value=0)  # 0 means unchecked (hide tables by default)
        self.show_history_checkbox = tk.Checkbutton(
            self.button_frame_top,
            text="Show History Tables",
            variable=self.show_history_var,
            command=self.toggle_history_tables
        )
        self.show_history_checkbox.grid(row=0, column=9, padx=5, pady=5)

        # Show/Hide Progress Bar Checkbox
        self.show_progress_var = tk.IntVar(value=0)  # 0 means unchecked (hide progress bar by default)
        self.show_progress_checkbox = tk.Checkbutton(
            self.button_frame_top,
            text="Show Progress Bar",
            variable=self.show_progress_var,
            command=self.toggle_progress_bar
        )
        self.show_progress_checkbox.grid(row=1, column=9, padx=5, pady=5)

        # Row Indicator
        self.row_indicator_var = tk.StringVar(value="Row: 0")
        self.row_indicator_label = tk.Label(
            root, textvariable=self.row_indicator_var)
        self.row_indicator_label.grid(row=1, column=1, sticky="w")

        # Specification Text Label and Text Box
        self.spec_text_label = tk.Label(
            root, text="Specification Text")
        self.spec_text_label.grid(row=2, column=1, columnspan=4, sticky="w")

        self.spec_text_box = tk.Text(
            root, height=4, width=70, bg="lightblue")
        self.spec_text_box.grid(row=3, column=1, columnspan=9, sticky="nsew")
        self.spec_text_box.config(state=tk.DISABLED)

        # Version and POC Note
        self.version_note = tk.Label(root, text=(
            "Program Version 4\n"
            "POC: Eriks@birdon.us"
        ), justify="left", anchor="w")
        self.version_note.grid(row=1, column=6, rowspan=6, sticky="nw")

        # New Label Above Progress Bar
        self.progress_label = tk.Label(
            root, text="Progress Bar")
        self.progress_label.grid(row=4, column=0, sticky="w")

        # Progress Bar Table
        self.progress_table = ttk.Treeview(
            root, columns=("Row Number",), show="headings", height=20)
        self.progress_table.grid(row=5, column=0, sticky="nsew")
        self.progress_table.heading("Row Number", text="Row Number")
        self.progress_table.column("Row Number", width=50, anchor="center")

        # Bind click event
        self.progress_table.bind("<ButtonRelease-1>", self.on_progress_bar_click)


        # Initially hide the progress bar table
        self.progress_label.grid_remove()
        self.progress_table.grid_remove()

        # DI Number Breakdown Label
        self.table_label = tk.Label(
            root, text="DI Number Breakdown")
        self.table_label.grid(row=4, column=1, sticky="w")

        # DI Number Breakdown Table
        self.table = ttk.Treeview(root, columns=(
            "VeriDoc Number", "DI Number", "CDRL Subtitle", "Object Status",
            "Contractor Assessed Status", "Government Assessed Status"),
            show="headings")
        self.table.grid(row=5, column=1, sticky="nsew")

        # Configure headings
        self.table.heading("VeriDoc Number", text="VeriDoc Number")
        self.table.heading("DI Number", text="DI Number")
        self.table.heading("CDRL Subtitle", text="CDRL Subtitle")
        self.table.heading("Object Status", text="Object Status")
        self.table.heading("Contractor Assessed Status", text="Contractor Assessed Status")
        self.table.heading("Government Assessed Status", text="Government Assessed Status")

        # Comment Section Label
        self.comment_section_label = tk.Label(
            root, text="Comment from Birdon to USCG")
        self.comment_section_label.grid(row=4, column=2, sticky="w")

        # Comment Table
        self.comment_table = ttk.Treeview(
            root, columns=("Comments",), show="headings")
        self.comment_table.grid(row=5, column=2, sticky="nsew")
        self.comment_table.heading("Comments", text="Comments")

        # Bind the right-click event to the comment table
        self.comment_table.bind("<Button-3>", self.show_comment_table_context_menu)

        # Remove double-click editing if desired
        self.comment_table.unbind('<Double-1>')

        # Proposed Changes Label
        self.proposed_changes_label = tk.Label(
            root, text="Contractor Proposed Change Request Input")
        self.proposed_changes_label.grid(row=4, column=3, sticky="w")

        # Proposed Changes Table
        self.proposed_changes_table = ttk.Treeview(
            root, columns=("Pattern",), show="headings")
        self.proposed_changes_table.grid(
            row=5, column=3, sticky="nsew")
        self.proposed_changes_table.heading("Pattern", text="Pattern")

        # Comment History Label
        self.comment_history_label = tk.Label(
            root, text="Contractor Proposed Change Comment History")
        self.comment_history_label.grid(row=4, column=4, sticky="w")

        # Comment History Table
        self.comment_history_table = ttk.Treeview(
            root, columns=("History",), show="headings")
        self.comment_history_table.grid(row=5, column=4, sticky="nsew")
        self.comment_history_table.heading("History", text="History")

        # Government Comment History Label
        self.gov_comment_history_label = tk.Label(
            root, text="Government Adjudication Comment History")
        self.gov_comment_history_label.grid(row=4, column=5, sticky="w")

        # Government Comment History Table
        self.gov_comment_history_table = ttk.Treeview(
            root, columns=("History",), show="headings")
        self.gov_comment_history_table.grid(row=5, column=5, sticky="nsew")
        self.gov_comment_history_table.heading("History", text="History")

        # Configure column widths (Optional)
        self.table.column("VeriDoc Number", width=120)
        self.table.column("DI Number", width=100)
        self.table.column("CDRL Subtitle", width=150)
        self.table.column("Object Status", width=120)
        self.table.column("Contractor Assessed Status", width=180)
        self.table.column("Government Assessed Status", width=180)

        self.comment_table.column("Comments", width=200)
        self.proposed_changes_table.column("Pattern", width=400)
        self.comment_history_table.column("History", width=400)
        self.gov_comment_history_table.column("History", width=400)
        self.progress_table.column("Row Number", width=50)

        # Configure styles for highlighting
        self.style = ttk.Style()
        self.style.configure("Treeview", rowheight=25)

        # Create custom styles for the cells
        self.style.configure("GovAgree.Cell", background="green", foreground="white")
        self.style.configure("GovDisagree.Cell", background="red", foreground="white")
        self.style.configure("sat_row", background="green", foreground="white")
        self.style.configure("unsat_row", background="red", foreground="white")
        self.style.configure("tbd_row", background="yellow", foreground="black")
        self.style.configure("highlight", background="lightblue")

        # Keep track of highlighted items
        self.highlighted_items = []

        # Bind events
        self.table.bind("<<TreeviewSelect>>", self.on_table_row_select)
        self.table.bind("<Button-3>", self.show_table_context_menu)
        self.proposed_changes_table.bind("<Button-3>", self.show_proposed_changes_context_menu)

        # Configure grid weights for proper resizing
        # Initially set weight=0 for progress bar column
        self.root.grid_columnconfigure(0, weight=0)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_columnconfigure(2, weight=1)
        self.root.grid_columnconfigure(3, weight=1)
        self.root.grid_columnconfigure(4, weight=1)
        self.root.grid_columnconfigure(5, weight=1)
        self.root.grid_rowconfigure(5, weight=1)

        # Initialize the visibility of history tables
        self.toggle_history_tables()

        # Initialize the visibility of progress bar
        self.toggle_progress_bar()

        #this prevents save functions from running multiple times
        self.save_lock = threading.Lock()  # Prevent simultaneous saves


    #this will allow the saving process to happen in its own thread in the background. 
    def save_comments_to_excel_background(self):
        def save_worker():
            try:
                with self.save_lock:
                    self.save_comments_to_excel()
            except Exception as e:
                # Optionally, if you need to show an error, schedule it on the main thread:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Background save error: {e}"))
        threading.Thread(target=save_worker, daemon=True).start()



    def open_tools_menu(self):
        # Create a new top-level window for tool selection
        self.tools_menu_window = tk.Toplevel(self.root)
        self.tools_menu_window.title("Select a Tool")
        self.tools_menu_window.attributes("-topmost", True)
        self.tools_menu_window.transient(self.root)
        self.tools_menu_window.lift()

        # Label
        tk.Label(self.tools_menu_window, text="Select a Tool:").pack(pady=10)

        # For future scalability, let's store tools in a dictionary
        # Key: Tool name (string), Value: function to open that tool
        self.available_tools = {
            "Compaire Tool": self.open_comair_tool_window,
            "Remove Previously Submitted Requests": self.open_remove_requests_tool_window,
            "RTVM Subset Management": self.open_rvtm_subset_management_window,
            "Disagreement Manager": self.open_disagreement_manager
        }
            # Later on, you can add more tools here:
            # "Another Tool": self.open_another_tool_window

        # Create a StringVar and Combobox for tool selection
        self.selected_tool = tk.StringVar(value="Compaire Tool")
        tool_list = list(self.available_tools.keys())
        self.tool_dropdown = ttk.Combobox(
            self.tools_menu_window, textvariable=self.selected_tool, values=tool_list, state="readonly")
        self.tool_dropdown.pack(pady=5)

        # Button to open the selected tool
        open_button = tk.Button(self.tools_menu_window, text="Open Selected Tool", command=self.open_selected_tool)
        open_button.pack(pady=10)

    def open_selected_tool(self):
        # Get the currently selected tool name
        tool_name = self.selected_tool.get()
        if tool_name in self.available_tools:
            # Call the corresponding function to open the tool
            self.tools_menu_window.destroy()  # Close the menu window
            self.available_tools[tool_name]()
        else:
            messagebox.showerror("Error", "Selected tool is not available.")





### START OF DIAGREEMENT TOOL 
    def open_disagreement_manager(self):
        if self.df is None:
            messagebox.showerror("Error", "No main Excel file is loaded. Please upload a main file first.")
            return

        # Filter rows
        # We need Government Assessed Status = "Disagree" and Object Status = "Accepted"
        # We'll use self.status_data which should have these fields
        # self.status_data is a list of dicts with keys like 'row_index', 'object_status', 'government_status', etc.
        filtered = [item for item in self.status_data 
                    if item['government_status'].strip().lower() == 'disagree' and 
                       item['object_status'].strip().lower() == 'accepted']

        if not filtered:
            messagebox.showinfo("No Disagreements", "No rows found with Government Assessed Status = Disagree and Object Status = Accepted.")
            return

        # Create the disagreement manager window
        self.disagreement_window = tk.Toplevel(self.root)
        self.disagreement_window.title("Disagreement Manager")

        # Create an instance of DisagreementManager class
        self.disagreement_manager = DisagreementManager(self.disagreement_window, self, filtered)


### END OF DIAGREEMENT TOOL             






     #### START OF REMOVAL TOOL
    def open_remove_requests_tool_window(self):
        # Create a new window
        self.remove_requests_window = tk.Toplevel(self.root)
        self.remove_requests_window.title("Remove Previously Submitted Contractor Proposed Change Request")

        # Explanation Label at the top
        explanation = ("This tool compares the currently loaded file with an older submission.\n"
                       "Any previously submitted contractor proposed change requests found in the old file\n"
                       "will be removed from the new file upon clicking 'Remove Previously Submitted Requests'.")
        tk.Label(self.remove_requests_window, text=explanation, justify="left").pack(pady=10, padx=10)

        # Frame for displaying current file and old file info
        files_frame = tk.Frame(self.remove_requests_window)
        files_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        # Label for the currently loaded new file
        tk.Label(files_frame, text="Current (New) File:").grid(row=0, column=0, sticky="w")
        self.new_file_label_var = tk.StringVar(value=self.excel_file_path if self.excel_file_path else "No File Loaded")
        tk.Label(files_frame, textvariable=self.new_file_label_var).grid(row=0, column=1, sticky="w")

        # Label and button for uploading old file
        tk.Label(files_frame, text="Old File:").grid(row=1, column=0, sticky="w")
        self.old_file_label_var = tk.StringVar(value="No Old File Selected")
        tk.Label(files_frame, textvariable=self.old_file_label_var).grid(row=1, column=1, sticky="w")
        self.old_file_button = tk.Button(files_frame, text="Browse Old File", command=self.browse_old_file)
        self.old_file_button.grid(row=1, column=2, padx=5, sticky="w")

        # Button to remove previously submitted requests
        self.remove_requests_button = tk.Button(self.remove_requests_window, text="Remove Previously Submitted Requests",
                                                command=self.remove_previously_submitted_requests)
        self.remove_requests_button.pack(pady=20)

        # Create a frame for the progress bar
        self.progress_frame = tk.Frame(self.remove_requests_window)
        self.progress_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        self.progress_frame.grid_remove()  # Hide initially

        tk.Label(self.progress_frame, text="Processing...").pack(side=tk.LEFT, padx=5)
        self.removal_progress = ttk.Progressbar(self.progress_frame, orient='horizontal', length=300, mode='determinate')
        self.removal_progress.pack(side=tk.LEFT, padx=5)



    def browse_old_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            self.old_file_path = file_path
            self.old_file_label_var.set(file_path)

    def remove_previously_submitted_requests(self):
        # Check if main new file and old file are uploaded
        if not hasattr(self, 'excel_file_path') or not self.excel_file_path:
            messagebox.showerror("Error", "No main (new) file is loaded. Please upload a main file first.")
            return
        if not hasattr(self, 'old_file_path') or not self.old_file_path:
            messagebox.showerror("Error", "No old file is selected.")
            return

        try:
            old_df = pd.read_excel(self.old_file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read old Excel file: {e}")
            return

        # Identify columns in the current DF
        # Proposed Changes Column (G) is known to be at index 6
        proposed_changes_col = 6

        # Identify the "Contractor Proposed Change Comment Input" column by name
        comment_col_name = "Contractor Proposed Change Comment Input"
        if comment_col_name not in self.df.columns:
            messagebox.showerror("Error", f"Column '{comment_col_name}' not found in new file.")
            return
        comment_col = self.df.columns.get_loc(comment_col_name)

        # Validate column indexes in the old file
        if proposed_changes_col >= len(self.df.columns):
            messagebox.showerror("Error", "Proposed Changes column not found in new file.")
            return
        if proposed_changes_col >= len(old_df.columns):
            messagebox.showerror("Error", "Proposed Changes column not found in old file.")
            return

        if comment_col >= len(self.df.columns):
            messagebox.showerror("Error", f"'{comment_col_name}' column not found in new file.")
            return
        if comment_col >= len(old_df.columns):
            messagebox.showerror("Error", f"'{comment_col_name}' column not found in old file.")
            return

        # Convert object ID columns to string (Object ID is assumed to be column 0)
        object_id_col = 0
        self.df.iloc[:, object_id_col] = self.df.iloc[:, object_id_col].astype(str)
        old_df.iloc[:, object_id_col] = old_df.iloc[:, object_id_col].astype(str)

        # Create a dictionary from old_df for old patterns
        old_patterns_map = {}
        for idx, row in old_df.iterrows():
            obj_id = str(row.iloc[object_id_col])
            old_patterns_str = row.iloc[proposed_changes_col]
            if pd.isna(old_patterns_str):
                old_patterns_str = ""
            old_lines = [line.strip() for line in old_patterns_str.split('\n') if line.strip()]
            old_patterns_set = set(old_lines)
            old_patterns_map[obj_id] = {
                'patterns': old_patterns_set
            }

        # Also create a dictionary for old comments (from the old file)
        old_comments_map = {}
        for idx, row in old_df.iterrows():
            obj_id = str(row.iloc[object_id_col])
            old_comments_str = row.iloc[comment_col]
            if pd.isna(old_comments_str):
                old_comments_str = ""
            old_comment_lines = [line.strip() for line in old_comments_str.split('\n') if line.strip()]
            old_comments_set = set(old_comment_lines)
            if obj_id not in old_patterns_map:
                old_patterns_map[obj_id] = {'patterns': set()}
            # Add comments to the same map for convenience
            old_patterns_map[obj_id]['comments'] = old_comments_set

        changes_made = False

        # Remove previously submitted requests from Proposed Changes (column G)
        for idx, row in self.df.iterrows():
            obj_id = str(row.iloc[object_id_col])
            new_patterns_str = row.iloc[proposed_changes_col]
            if pd.isna(new_patterns_str):
                new_patterns_str = ""
            new_lines = [line.strip() for line in new_patterns_str.split('\n') if line.strip()]

            if obj_id in old_patterns_map:
                old_patterns_set = old_patterns_map[obj_id]['patterns']
                # Filter out any patterns that exist in old file
                filtered_lines = [line for line in new_lines if line not in old_patterns_set]
                if len(filtered_lines) != len(new_lines):
                    changes_made = True
                    updated_str = '\n'.join(filtered_lines)
                    self.df.iat[idx, proposed_changes_col] = updated_str

        # Remove previously submitted comments from "Contractor Proposed Change Comment Input" column
        for idx, row in self.df.iterrows():
            obj_id = str(row.iloc[object_id_col])
            new_comments_str = row.iloc[comment_col]
            if pd.isna(new_comments_str):
                new_comments_str = ""
            new_comment_lines = [line.strip() for line in new_comments_str.split('\n') if line.strip()]

            if obj_id in old_patterns_map and 'comments' in old_patterns_map[obj_id]:
                old_comments_set = old_patterns_map[obj_id]['comments']
                filtered_comment_lines = [line for line in new_comment_lines if line not in old_comments_set]
                if len(filtered_comment_lines) != len(new_comment_lines):
                    changes_made = True
                    updated_comments_str = '\n'.join(filtered_comment_lines)
                    self.df.iat[idx, comment_col] = updated_comments_str

        if not changes_made:
            messagebox.showinfo("No Changes", "No previously submitted requests or comments were found to remove.")
            return

        # If changes were made, update the Excel file preserving formatting
        from openpyxl import load_workbook

        try:
            wb = load_workbook(self.excel_file_path)
            ws = wb.active  # Adjust if you need a specific sheet

            # Find the column indexes in the Excel sheet
            # Column G (Proposed Changes) = 7 in 1-based indexing
            # For 'Contractor Proposed Change Comment Input', find the correct column by header
            header_row = 1
            comment_col_letter = None
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(row=header_row, column=col).value
                if header_val == comment_col_name:
                    comment_col_letter = col
                    break

            # Update only the changed cells
            for idx, row in self.df.iterrows():
                new_content = row.iloc[proposed_changes_col]
                if pd.isna(new_content):
                    new_content = ""
                ws.cell(row=idx+2, column=7, value=new_content)  # Proposed Changes column (G)

                new_comment_content = row.iloc[comment_col]
                if pd.isna(new_comment_content):
                    new_comment_content = ""
                if comment_col_letter is not None:
                    ws.cell(row=idx+2, column=comment_col_letter, value=new_comment_content)

            wb.save(self.excel_file_path)

            # DataFrame in memory is already updated
            messagebox.showinfo("Success", "Previously submitted requests and comments have been removed and the file has been updated.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save updated file: {e}")


### END of removal Tool

### START OF COMPAIRE TOOL 
    def open_comair_tool_window(self):
        # This is your original `open_tools_window()` code
        # Renamed to `open_comair_tool_window()` for clarity.

        self.comair_window = tk.Toplevel(self.root)
        self.comair_window.title("Comair Tool")

        # Set window size (optional)
        self.comair_window.geometry("800x600")  # Adjust as needed

        # The rest of the original open_tools_window code goes here...
        # For example, your file comparison UI and logic:
        file_frame = tk.Frame(self.comair_window)
        file_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        self.file1_label = tk.Label(file_frame, text="Select First Excel File:")
        self.file1_label.grid(row=0, column=0, sticky="w")
        self.file1_entry = tk.Entry(file_frame, width=50)
        self.file1_entry.grid(row=0, column=1, padx=5)
        self.file1_button = tk.Button(file_frame, text="Browse", command=self.browse_file1)
        self.file1_button.grid(row=0, column=2, padx=5)

        self.file2_label = tk.Label(file_frame, text="Select Second Excel File:")
        self.file2_label.grid(row=1, column=0, sticky="w")
        self.file2_entry = tk.Entry(file_frame, width=50)
        self.file2_entry.grid(row=1, column=1, padx=5)
        self.file2_button = tk.Button(file_frame, text="Browse", command=self.browse_file2)
        self.file2_button.grid(row=1, column=2, padx=5)

        self.compare_button = tk.Button(self.comair_window, text="Compare Files", command=self.compare_files)
        self.compare_button.pack(pady=10)

        self.save_results_button = tk.Button(self.comair_window, text="Save Results", command=self.save_comparison_results)
        self.save_results_button.pack(pady=5)

        self.results_frame = tk.Frame(self.comair_window)
        self.results_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("Row", "Column", "Value in File 1", "Value in File 2")
        self.results_table = ttk.Treeview(self.results_frame, columns=columns, show="headings")
        self.results_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for col in columns:
            self.results_table.heading(col, text=col)
            self.results_table.column(col, anchor='center', width=150)

        scrollbar = ttk.Scrollbar(self.results_frame, orient=tk.VERTICAL, command=self.results_table.yview)
        self.results_table.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)



    def save_comparison_results(self):
        # Get the data from the results table
        rows = self.results_table.get_children()
        if not rows:
            messagebox.showinfo("No Data", "No differences to save.")
            return

        # Prepare data for DataFrame
        data = []
        for row_id in rows:
            row = self.results_table.item(row_id)['values']
            data.append(row)

        df = pd.DataFrame(data, columns=['Row', 'Column', 'Value in File 1', 'Value in File 2'])

        # Ask user for a file location
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 title="Save Comparison Results")
        if file_path:
            try:
                df.to_excel(file_path, index=False)
                messagebox.showinfo("Success", f"Results saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save results: {e}")

    def get_differences(self, df1, df2):
        try:
            # Align the DataFrames
            df1, df2 = df1.align(df2, join='outer', axis=1)
            df1.fillna('', inplace=True)
            df2.fillna('', inplace=True)

            # Use pandas built-in comparison
            diff = df1.compare(df2, keep_equal=False)
            diff.reset_index(inplace=True)

            # Prepare a list to collect differences
            differences = []

            for index, row in diff.iterrows():
                row_num = row['index'] + 2  # Adjust for human-readable indexing (assuming header row is row 1)
                for col in df1.columns:
                    if col in diff.columns.get_level_values(0):
                        val1 = diff.at[index, (col, 'self')]
                        val2 = diff.at[index, (col, 'other')]
                        differences.append({
                            'Row': row_num,
                            'Column': col,
                            'Value in File 1': val1,
                            'Value in File 2': val2
                        })

            # Create the DataFrame from the list of differences
            diff_df = pd.DataFrame(differences, columns=['Row', 'Column', 'Value in File 1', 'Value in File 2'])
            return diff_df
        except Exception as e:
            messagebox.showerror("Error", f"Failed to compare Excel files: {e}")
            return pd.DataFrame()


    def browse_file1(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            self.file1_entry.delete(0, tk.END)
            self.file1_entry.insert(0, file_path)

    def browse_file2(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            self.file2_entry.delete(0, tk.END)
            self.file2_entry.insert(0, file_path)


    def compare_files(self):
        file1_path = self.file1_entry.get()
        file2_path = self.file2_entry.get()

        if not file1_path or not file2_path:
            messagebox.showerror("Error", "Please select both Excel files to compare.")
            return

        try:
            df1 = pd.read_excel(file1_path)
            df2 = pd.read_excel(file2_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel files: {e}")
            return

        # Clear previous results
        self.results_table.delete(*self.results_table.get_children())

        # Compare the dataframes
        try:
            diff_df = self.get_differences(df1, df2)
            if diff_df.empty:
                messagebox.showinfo("No Differences", "The two files are identical.")
            else:
                # Display differences in the results table
                for index, row in diff_df.iterrows():
                    self.results_table.insert("", "end", values=(row['Row'], row['Column'], row['Value in File 1'], row['Value in File 2']))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to compare Excel files: {e}")




### End of compair tool ################################################################################################################################################################################################################################

### Start of RTVM Subsets tool pack ################################################################################################################################################################################################################################






    import os
    import re
    import queue
    import json
    import threading
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from datetime import datetime
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog

    def enhance_rvtm_subset_management_window(self):
        """
        Enhanced version of the RTVM Subset Management window with status reporting
        """
        self.rvtm_subset_window = tk.Toplevel(self.root)
        self.rvtm_subset_window.title("RTVM Subset Management")
        self.rvtm_subset_window.geometry("900x700")  # Increased window size

        # Split the window into two main frames - top commands and bottom results
        top_frame = tk.Frame(self.rvtm_subset_window)
        top_frame.pack(fill='x', padx=10, pady=10)
    
        notebook = ttk.Notebook(self.rvtm_subset_window)
        notebook.pack(fill='both', expand=True, padx=10, pady=5)
    
        # First tab: Management Operations
        operations_frame = ttk.Frame(notebook)
        notebook.add(operations_frame, text='Management Operations')
    
        # Second tab: Subset Status
        status_frame = ttk.Frame(notebook)
        notebook.add(status_frame, text='Subset Status')
    
        # Third tab: Subset Analytics
        analytics_frame = ttk.Frame(notebook)
        notebook.add(analytics_frame, text='Analytics')

        explanation = (
            "This RTVM Subset Management tool allows you to:\n"
            "- Create a summary report of verification data\n"
            "- Export photos of the generated report\n"
            "- Create subsets of the RTVM based on predefined SWBS groups\n"
            "- Recombine these subsets back into a single file\n"
            "- Merge a single subset into the main RTVM file\n"
            "- Combine and update subsets while preserving changes\n\n"
            "The data is taken from the currently loaded main RTVM file.\n"
            "Please select a base location first."
        )
    
        tk.Label(top_frame, text=explanation, justify="left").pack(pady=10, padx=10)

        # Button to select the base location
        select_location_button = tk.Button(
            top_frame, text="Select Base Location", command=self.select_base_location
        )
        select_location_button.pack(side="left", padx=5, pady=5)

        # Display the currently loaded file name
        tk.Label(top_frame, text="Current File:", anchor='w').pack(side="left")
        self.file_name_var = tk.StringVar(value=self.excel_file_path if self.excel_file_path else "No File Loaded")
        tk.Label(top_frame, textvariable=self.file_name_var, width=50, anchor='w').pack(side="left", padx=5, pady=5)

        # === OPERATIONS TAB ===
        operations_buttons_frame = tk.Frame(operations_frame)
        operations_buttons_frame.pack(fill='x', padx=10, pady=10)

        # Create Summary Report Button
        self.create_report_button = tk.Button(
            operations_buttons_frame, text="Create Summary Report", command=self.create_summary_report
        )
        self.create_report_button.grid(row=0, column=0, padx=5, pady=5)

        # Export Photos Button
        self.export_photos_button = tk.Button(
            operations_buttons_frame, text="Export Photos of Report", command=self.export_photos_of_report
        )
        self.export_photos_button.grid(row=0, column=1, padx=5, pady=5)

        # Create Subsets Button
        self.create_subsets_button = tk.Button(
            operations_buttons_frame, text="Create Subsets", command=self.create_subsets
        )
        self.create_subsets_button.grid(row=0, column=2, padx=5, pady=5)

        # Recombine Subsets Button
        self.recombine_subsets_button = tk.Button(
            operations_buttons_frame, text="Recombine Subsets", command=self.recombine_subsets
        )
        self.recombine_subsets_button.grid(row=0, column=3, padx=5, pady=5)

        # Merge Single Subset Button
        self.merge_single_subset_button = tk.Button(
            operations_buttons_frame, text="Merge Single Subset", command=self.merge_single_subset
        )
        self.merge_single_subset_button.grid(row=0, column=4, padx=5, pady=5)

        # NEW: Add Combine & Update Subsets Button
        self.combine_update_button = tk.Button(
            operations_buttons_frame, text="Combine & Update Subsets", command=self.combine_and_update_subsets
        )
        self.combine_update_button.grid(row=0, column=5, padx=5, pady=5)

        # === STATUS TAB ===
        status_top_frame = tk.Frame(status_frame)
        status_top_frame.pack(fill='x', padx=10, pady=10)
    
        # Scan Subsets Button
        self.scan_subsets_button = tk.Button(
            status_top_frame, text="Scan Subset Files", command=self.scan_subset_files,
            bg="#4CAF50", fg="white", font=("Helvetica", 10, "bold")
        )
        self.scan_subsets_button.pack(side="left", padx=5, pady=5)
    
        # PMR Selection for status tab
        tk.Label(status_top_frame, text="PMR Number:").pack(side="left", padx=5, pady=5)
        self.status_pmr_var = tk.StringVar()
        pmr_entry = tk.Entry(status_top_frame, textvariable=self.status_pmr_var, width=8)
        pmr_entry.pack(side="left", padx=5, pady=5)
    
        # Auto-detect PMRs Button
        self.detect_pmrs_button = tk.Button(
            status_top_frame, text="Detect PMRs", command=self.detect_pmrs
        )
        self.detect_pmrs_button.pack(side="left", padx=5, pady=5)
    
        # Export Status Report Button
        self.export_status_button = tk.Button(
            status_top_frame, text="Export Status Report", command=self.export_subset_status_report
        )
        self.export_status_button.pack(side="left", padx=5, pady=5)
    
        # Create a frame for the subset status table
        status_table_frame = tk.Frame(status_frame)
        status_table_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
        # Create the subset status table
        columns = (
            "Subset", "SWBS", "Modified", "Total Patterns", 
            "ADD Requests", "DEL Requests", "Detail Updates", 
            "SAT Status", "UNSAT Status", "Completion %"
        )
    
        self.subset_status_table = ttk.Treeview(
            status_table_frame, 
            columns=columns, 
            show="headings",
            selectmode="browse"
        )
    
        # Configure column headings and widths
        for col in columns:
            self.subset_status_table.heading(col, text=col, command=lambda c=col: self.sort_subset_table(c))
    
        self.subset_status_table.column("Subset", width=150)
        self.subset_status_table.column("SWBS", width=80)
        self.subset_status_table.column("Modified", width=120)
        self.subset_status_table.column("Total Patterns", width=90, anchor='e')
        self.subset_status_table.column("ADD Requests", width=90, anchor='e')
        self.subset_status_table.column("DEL Requests", width=90, anchor='e')
        self.subset_status_table.column("Detail Updates", width=90, anchor='e')
        self.subset_status_table.column("SAT Status", width=80, anchor='e')
        self.subset_status_table.column("UNSAT Status", width=80, anchor='e')
        self.subset_status_table.column("Completion %", width=90, anchor='e')
    
        # Add scrollbars
        status_y_scroll = ttk.Scrollbar(status_table_frame, orient="vertical", command=self.subset_status_table.yview)
        status_x_scroll = ttk.Scrollbar(status_table_frame, orient="horizontal", command=self.subset_status_table.xview)
        self.subset_status_table.configure(yscrollcommand=status_y_scroll.set, xscrollcommand=status_x_scroll.set)
    
        # Pack the table and scrollbars
        status_y_scroll.pack(side="right", fill="y")
        status_x_scroll.pack(side="bottom", fill="x")
        self.subset_status_table.pack(side="left", fill="both", expand=True)
    
        # Double-click to open subset file
        self.subset_status_table.bind("<Double-1>", self.open_selected_subset)
        # Right-click context menu
        self.subset_status_table.bind("<Button-3>", self.show_subset_context_menu)
    
        # === ANALYTICS TAB ===
        analytics_top_frame = tk.Frame(analytics_frame)
        analytics_top_frame.pack(fill='x', padx=10, pady=10)
    
        # Generate Analytics Button
        self.generate_analytics_button = tk.Button(
            analytics_top_frame, text="Generate Analytics", command=self.generate_subset_analytics
        )
        self.generate_analytics_button.pack(side="left", padx=5, pady=5)
    
        # Create frame for charts
        self.analytics_charts_frame = tk.Frame(analytics_frame)
        self.analytics_charts_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
        # Initialize charts area with empty placeholder
        self.init_analytics_placeholder()
    
        # Call this function to set the assigned_verification_cells attribute
        self.set_assigned_verification_cells()

    def show_subset_context_menu(self, event):
        """Show context menu for subset status table"""
        # Identify the row under the cursor
        row_id = self.subset_status_table.identify_row(event.y)
        if not row_id:
            return
    
        # Select the row
        self.subset_status_table.selection_set(row_id)
    
        # Create context menu
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(label="Open Subset", command=lambda: self.open_selected_subset(None))
        context_menu.add_command(label="Show Folder Location", command=self.show_subset_folder)
        context_menu.add_separator()
        context_menu.add_command(label="View Pattern Details", command=self.view_subset_patterns)
        context_menu.add_command(label="Check for Issues", command=self.check_subset_issues)
        context_menu.add_separator()
        context_menu.add_command(label="Mark as Complete", command=lambda: self.mark_subset_status("complete"))
        context_menu.add_command(label="Mark as In Progress", command=lambda: self.mark_subset_status("in_progress"))
    
        # Display the menu
        context_menu.post(event.x_root, event.y_root)

    def open_selected_subset(self, event):
        """Open the selected subset file"""
        selected_items = self.subset_status_table.selection()
        if not selected_items:
            return
    
        item = selected_items[0]
        filename = self.subset_status_table.item(item, "values")[0]  # Get the filename from the UI
    
        # We now need to find the full path from subset_stats since we're only displaying filenames
        file_path = None
        for stat in self.subset_stats:
            if stat["filename"] == filename:
                file_path = stat["file_path"]
                break
    
        if not file_path:
            messagebox.showerror("Error", f"Could not find full path for: {filename}")
            return
    
        try:
            # Try to open the file with the default application
            import os
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")

    def export_subset_status_report(self):
        """Export the subset status data to Excel"""
        if not hasattr(self, 'subset_stats') or not self.subset_stats:
            messagebox.showinfo("No Data", "Please scan subset files first")
            return
    
        # Ask user for save location
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="RTVM_Subset_Status_Report.xlsx"
        )
    
        if not save_path:
            return
    
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Subset Status"
        
            # Add headers
            headers = ["Subset File", "SWBS", "Last Modified", "Total Patterns", 
                       "ADD Requests", "DEL Requests", "Detail Updates", 
                       "SAT Status", "UNSAT Status", "Completion %"]
        
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Add data
            for row, subset in enumerate(self.subset_stats, start=2):
                ws.cell(row=row, column=1, value=os.path.basename(subset["file_path"]))
                ws.cell(row=row, column=2, value=subset["swbs"])
                ws.cell(row=row, column=3, value=subset["modified"])
                ws.cell(row=row, column=4, value=subset["total_patterns"])
                ws.cell(row=row, column=5, value=subset["add_requests"])
                ws.cell(row=row, column=6, value=subset["del_requests"])
                ws.cell(row=row, column=7, value=subset["detail_updates"])
                ws.cell(row=row, column=8, value=subset["sat_status"])
                ws.cell(row=row, column=9, value=subset["unsat_status"])
            
                # Format completion percentage
                completion = subset["completion_pct"]
                ws.cell(row=row, column=10, value=completion)
            
                # Conditional formatting based on completion
                if completion != "N/A" and completion != "Error":
                    try:
                        completion_value = float(completion)
                        if completion_value >= 90:
                            ws.cell(row=row, column=10).fill = PatternFill(start_color="a5d6a7", end_color="a5d6a7", fill_type="solid")
                        elif completion_value >= 50:
                            ws.cell(row=row, column=10).fill = PatternFill(start_color="fff59d", end_color="fff59d", fill_type="solid")
                        elif completion_value > 0:
                            ws.cell(row=row, column=10).fill = PatternFill(start_color="ffcc80", end_color="ffcc80", fill_type="solid")
                        else:
                            ws.cell(row=row, column=10).fill = PatternFill(start_color="ef9a9a", end_color="ef9a9a", fill_type="solid")
                    except ValueError:
                        pass
        
            # Add summary sheet
            ws_summary = wb.create_sheet(title="Summary")
        
            # Add headers
            summary_headers = ["Metric", "Value"]
            for col, header in enumerate(summary_headers, start=1):
                ws_summary.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Calculate summary statistics
            total_files = len(self.subset_stats)
            total_patterns = sum(subset["total_patterns"] for subset in self.subset_stats if isinstance(subset["total_patterns"], int))
            total_add = sum(subset["add_requests"] for subset in self.subset_stats if isinstance(subset["add_requests"], int))
            total_del = sum(subset["del_requests"] for subset in self.subset_stats if isinstance(subset["del_requests"], int))
            total_updates = sum(subset["detail_updates"] for subset in self.subset_stats if isinstance(subset["detail_updates"], int))
            total_sat = sum(subset["sat_status"] for subset in self.subset_stats if isinstance(subset["sat_status"], int))
            total_unsat = sum(subset["unsat_status"] for subset in self.subset_stats if isinstance(subset["unsat_status"], int))
        
            if total_patterns > 0:
                overall_completion = (total_sat / total_patterns) * 100
            else:
                overall_completion = 0
        
            # Add summary data
            row = 2
            ws_summary.cell(row=row, column=1, value="Total Subset Files")
            ws_summary.cell(row=row, column=2, value=total_files)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total Patterns")
            ws_summary.cell(row=row, column=2, value=total_patterns)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total ADD Requests")
            ws_summary.cell(row=row, column=2, value=total_add)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total DEL Requests")
            ws_summary.cell(row=row, column=2, value=total_del)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total Detail Updates")
            ws_summary.cell(row=row, column=2, value=total_updates)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total SAT Status")
            ws_summary.cell(row=row, column=2, value=total_sat)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Total UNSAT Status")
            ws_summary.cell(row=row, column=2, value=total_unsat)
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Overall Completion %")
            ws_summary.cell(row=row, column=2, value=f"{overall_completion:.1f}%")
        
            # Calculate completion by SWBS
            swbs_data = {}
            for subset in self.subset_stats:
                swbs = subset["swbs"]
                if swbs not in swbs_data:
                    swbs_data[swbs] = {
                        "total_patterns": 0,
                        "sat_status": 0
                    }
            
                # Add pattern counts if they're numeric
                if isinstance(subset["total_patterns"], int):
                    swbs_data[swbs]["total_patterns"] += subset["total_patterns"]
                if isinstance(subset["sat_status"], int):
                    swbs_data[swbs]["sat_status"] += subset["sat_status"]
        
            # Add SWBS breakdown section
            row += 2
            ws_summary.cell(row=row, column=1, value="Completion by SWBS Group").font = Font(bold=True)
        
            for swbs, data in swbs_data.items():
                row += 1
                if data["total_patterns"] > 0:
                    swbs_completion = (data["sat_status"] / data["total_patterns"]) * 100
                else:
                    swbs_completion = 0
            
                ws_summary.cell(row=row, column=1, value=f"{swbs} Completion %")
                ws_summary.cell(row=row, column=2, value=f"{swbs_completion:.1f}%")
            
                # Conditional formatting
                if swbs_completion >= 90:
                    ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="a5d6a7", end_color="a5d6a7", fill_type="solid")
                elif swbs_completion >= 50:
                    ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="fff59d", end_color="fff59d", fill_type="solid")
                elif swbs_completion > 0:
                    ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="ffcc80", end_color="ffcc80", fill_type="solid")
                else:
                    ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="ef9a9a", end_color="ef9a9a", fill_type="solid")
        
            # Auto-adjust column widths
            for sheet in [ws, ws_summary]:
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column].width = adjusted_width
        
            # Save the workbook
            wb.save(save_path)
            messagebox.showinfo("Export Complete", f"Status report exported to {save_path}")
    
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export status report: {str(e)}")

    def generate_subset_analytics(self):
        """Generate analytics charts for subset data"""
        if not hasattr(self, 'subset_stats') or not self.subset_stats:
            messagebox.showinfo("No Data", "Please scan subset files first")
            return
    
        # Clear previous charts
        for widget in self.analytics_charts_frame.winfo_children():
            widget.destroy()
    
        # Create a grid layout for charts
        analytics_top = tk.Frame(self.analytics_charts_frame)
        analytics_top.pack(fill='both', expand=True)
    
        analytics_bottom = tk.Frame(self.analytics_charts_frame)
        analytics_bottom.pack(fill='both', expand=True)
    
        # Calculate analytics data
        total_patterns = sum(subset["total_patterns"] for subset in self.subset_stats if isinstance(subset["total_patterns"], int))
        total_add = sum(subset["add_requests"] for subset in self.subset_stats if isinstance(subset["add_requests"], int))
        total_del = sum(subset["del_requests"] for subset in self.subset_stats if isinstance(subset["del_requests"], int))
        total_updates = sum(subset["detail_updates"] for subset in self.subset_stats if isinstance(subset["detail_updates"], int))
        total_sat = sum(subset["sat_status"] for subset in self.subset_stats if isinstance(subset["sat_status"], int))
        total_unsat = sum(subset["unsat_status"] for subset in self.subset_stats if isinstance(subset["unsat_status"], int))
    
        # Group by SWBS
        swbs_data = {}
        for subset in self.subset_stats:
            swbs = subset["swbs"]
            if swbs not in swbs_data:
                swbs_data[swbs] = {
                    "total_patterns": 0,
                    "sat_status": 0,
                    "unsat_status": 0,
                    "add_requests": 0,
                    "del_requests": 0,
                    "detail_updates": 0
                }
        
            # Add numeric values only
            if isinstance(subset["total_patterns"], int):
                swbs_data[swbs]["total_patterns"] += subset["total_patterns"]
            if isinstance(subset["sat_status"], int):
                swbs_data[swbs]["sat_status"] += subset["sat_status"]
            if isinstance(subset["unsat_status"], int):
                swbs_data[swbs]["unsat_status"] += subset["unsat_status"]
            if isinstance(subset["add_requests"], int):
                swbs_data[swbs]["add_requests"] += subset["add_requests"]
            if isinstance(subset["del_requests"], int):
                swbs_data[swbs]["del_requests"] += subset["del_requests"]
            if isinstance(subset["detail_updates"], int):
                swbs_data[swbs]["detail_updates"] += subset["detail_updates"]
    
        # Chart 1: Pattern Types Distribution (Pie Chart)
        fig1 = plt.Figure(figsize=(5, 4), dpi=100)
        ax1 = fig1.add_subplot(111)
    
        # Data for pie chart
        labels = ['ADD Requests', 'DEL Requests', 'Detail Updates']
        sizes = [total_add, total_del, total_updates]
        colors = ['#ff9999','#66b3ff','#99ff99']
        explode = (0.1, 0, 0)  # Explode ADD wedge
    
        # Plot pie chart
        ax1.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=90)
        ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
        ax1.set_title('Pattern Types Distribution')
    
        # Embed in canvas
        canvas1 = FigureCanvasTkAgg(fig1, master=analytics_top)
        canvas1.draw()
        canvas1.get_tk_widget().pack(side='left', fill='both', expand=True)
    
        # Chart 2: Status Distribution (Pie Chart)
        fig2 = plt.Figure(figsize=(5, 4), dpi=100)
        ax2 = fig2.add_subplot(111)
    
        # Data for pie chart
        labels2 = ['SAT', 'UNSAT']
        sizes2 = [total_sat, total_unsat]
        colors2 = ['#99ff99', '#ff9999']
    
        # Plot pie chart
        ax2.pie(sizes2, labels=labels2, colors=colors2, autopct='%1.1f%%', shadow=True, startangle=90)
        ax2.axis('equal')
        ax2.set_title('Status Distribution')
    
        # Embed in canvas
        canvas2 = FigureCanvasTkAgg(fig2, master=analytics_top)
        canvas2.draw()
        canvas2.get_tk_widget().pack(side='left', fill='both', expand=True)
    
        # Chart 3: Completion by SWBS (Bar Chart)
        fig3 = plt.Figure(figsize=(5, 4), dpi=100)
        ax3 = fig3.add_subplot(111)
    
        # Prepare data
        swbs_names = []
        completion_rates = []
    
        for swbs, data in swbs_data.items():
            swbs_names.append(swbs)
            if data["total_patterns"] > 0:
                completion = (data["sat_status"] / data["total_patterns"]) * 100
            else:
                completion = 0
            completion_rates.append(completion)
    
        # Sort data for better visualization
        sorted_data = sorted(zip(swbs_names, completion_rates), key=lambda x: x[1], reverse=True)
        swbs_names, completion_rates = zip(*sorted_data) if sorted_data else ([], [])
    
        # Create bars
        bars = ax3.bar(swbs_names, completion_rates, color='skyblue')
    
        # Add labels and title
        ax3.set_xlabel('SWBS Group')
        ax3.set_ylabel('Completion %')
        ax3.set_title('Completion by SWBS Group')
    
        # Adjust layout
        fig3.tight_layout()
    
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax3.text(
                bar.get_x() + bar.get_width()/2., 
                height + 1,
                f'{height:.1f}%',
                ha='center', 
                va='bottom',
                rotation=0
            )
    
        # Embed in canvas
        canvas3 = FigureCanvasTkAgg(fig3, master=analytics_bottom)
        canvas3.draw()
        canvas3.get_tk_widget().pack(side='left', fill='both', expand=True)
    
        # Chart 4: Pattern Types by SWBS (Stacked Bar Chart)
        fig4 = plt.Figure(figsize=(5, 4), dpi=100)
        ax4 = fig4.add_subplot(111)
    
        # Prepare data
        swbs_list = list(swbs_data.keys())
        add_counts = [swbs_data[swbs]["add_requests"] for swbs in swbs_list]
        del_counts = [swbs_data[swbs]["del_requests"] for swbs in swbs_list]
        update_counts = [swbs_data[swbs]["detail_updates"] for swbs in swbs_list]
    
        # Create stacked bars
        bar_width = 0.35
        ax4.bar(swbs_list, add_counts, bar_width, label='ADD', color='#ff9999')
        ax4.bar(swbs_list, del_counts, bar_width, bottom=add_counts, label='DEL', color='#66b3ff')
    
        # Calculate position for updates bar
        bottom_vals = [a + d for a, d in zip(add_counts, del_counts)]
        ax4.bar(swbs_list, update_counts, bar_width, bottom=bottom_vals, label='Updates', color='#99ff99')
    
        # Add labels and title
        ax4.set_xlabel('SWBS Group')
        ax4.set_ylabel('Number of Patterns')
        ax4.set_title('Pattern Types by SWBS')
        ax4.legend()
    
        # Adjust layout
        fig4.tight_layout()
    
        # Embed in canvas
        canvas4 = FigureCanvasTkAgg(fig4, master=analytics_bottom)
        canvas4.draw()
        canvas4.get_tk_widget().pack(side='left', fill='both', expand=True)
    
        # Add export button
        export_frame = tk.Frame(self.analytics_charts_frame)
        export_frame.pack(fill='x', padx=10, pady=10)
    
        export_button = tk.Button(
            export_frame, 
            text="Export Analytics to PDF", 
            command=lambda: self.export_analytics_to_pdf([fig1, fig2, fig3, fig4])
        )
        export_button.pack(side='right', padx=5)
    
        # Add analytics summary text
        summary_frame = tk.Frame(self.analytics_charts_frame)
        summary_frame.pack(fill='x', padx=10, pady=5)
    
        # Calculate overall completion
        if total_patterns > 0:
            overall_completion = (total_sat / total_patterns) * 100
        else:
            overall_completion = 0
    
        summary_text = f"""
        Analytics Summary:
    
        • Total Subset Files: {len(self.subset_stats)}
        • Total Patterns: {total_patterns}
        • Pattern Types: {total_add} ADD, {total_del} DEL, {total_updates} Updates
        • Status Distribution: {total_sat} SAT, {total_unsat} UNSAT
        • Overall Completion: {overall_completion:.1f}%
        • SWBS Groups: {len(swbs_data)}
        """
    
        summary_label = tk.Label(summary_frame, text=summary_text, justify='left', font=("Helvetica", 10))
        summary_label.pack(side='left', padx=10, pady=5, anchor='nw')

    def export_analytics_to_pdf(self, figures):
        """Export the analytics charts to a PDF file"""
        # Ask user for save location
        save_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile="RTVM_Subset_Analytics.pdf"
        )
    
        if not save_path:
            return
    
        try:
            # Create PDF with matplotlib
            from matplotlib.backends.backend_pdf import PdfPages
        
            with PdfPages(save_path) as pdf:
                # Add each figure to the PDF
                for fig in figures:
                    pdf.savefig(fig)
            
                # Set PDF metadata
                d = pdf.infodict()
                d['Title'] = 'RTVM Subset Analytics'
                d['Author'] = os.getenv('USERNAME', 'RTVM Tool User')
                d['Subject'] = 'RTVM Subset Management Analytics'
                d['Keywords'] = 'RTVM, analytics, subset, statistics'
                d['CreationDate'] = datetime.now()
        
            messagebox.showinfo("Export Complete", f"Analytics exported to {save_path}")
    
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export analytics: {str(e)}")

    # Integration method for the main RTVMApp class
    def open_rvtm_subset_management_window(self):
        """
        Open the enhanced RTVM Subset Management window
        """
        # Call the new enhanced version
        self.enhance_rvtm_subset_management_window()
      # Assuming full path is stored
    
        try:
            # Try to open the file with the default application
            import os
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")

    def show_subset_folder(self):
        """Open the folder containing the selected subset"""
        selected_items = self.subset_status_table.selection()
        if not selected_items:
            return
    
        item = selected_items[0]
        filename = self.subset_status_table.item(item, "values")[0]  # Get the filename from the UI
    
        # We now need to find the full path from subset_stats since we're only displaying filenames
        file_path = None
        for stat in self.subset_stats:
            if stat["filename"] == filename:
                file_path = stat["file_path"]
                break
    
        if not file_path:
            messagebox.showerror("Error", f"Could not find full path for: {filename}")
            return
    
        folder_path = os.path.dirname(file_path)
    
        try:
            # Open the folder in File Explorer
            import os
            os.startfile(folder_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder: {e}")

    def sort_subset_table(self, column):
        """Sort the subset status table by the clicked column"""
        # Get current sort direction
        sort_order = getattr(self, 'subset_sort_order', True)  # Default ascending
        self.subset_sort_order = not sort_order  # Toggle for next click
    
        # Get all items along with the specified column value
        data = [(self.subset_status_table.set(item, column), item) for item in self.subset_status_table.get_children('')]
    
        # Determine sort key based on column type
        numeric_columns = [
            "Total Patterns", "ADD Requests", "DEL Requests", "Detail Updates", 
            "SAT Status", "UNSAT Status", "Completion %"
        ]
    
        # Sort the data
        if column in numeric_columns:
            data.sort(key=lambda x: float(x[0]) if x[0] and x[0] != 'N/A' else -1, reverse=sort_order)
        else:
            data.sort(key=lambda x: x[0], reverse=sort_order)
    
        # Rearrange items according to sort
        for index, (_, item) in enumerate(data):
            self.subset_status_table.move(item, '', index)
    
        # Update column heading to show sort indicator
        for col in self.subset_status_table["columns"]:
            if col == column:
                direction = "▼" if sort_order else "▲"
                self.subset_status_table.heading(col, text=f"{col} {direction}")
            else:
                self.subset_status_table.heading(col, text=col)

    def detect_pmrs(self):
        """Auto-detect available PMR numbers in the base folder"""
        if not self.selected_base_path:
            messagebox.showerror("Error", "Please select a base location first")
            return
    
        pmr_folders = []
    
        # Look for folders with pattern "PMR X"
        for item in os.listdir(self.selected_base_path):
            item_path = os.path.join(self.selected_base_path, item)
            if os.path.isdir(item_path) and item.startswith("PMR "):
                try:
                    pmr_num = item.split("PMR ")[1].strip()
                    pmr_folders.append(pmr_num)
                except:
                    continue
    
        if not pmr_folders:
            messagebox.showinfo("No PMRs Found", "No PMR folders detected in the base location")
            return
    
        # If PMRs are found, show dialog for user to select one
        pmr_window = tk.Toplevel(self.root)
        pmr_window.title("Select PMR")
        pmr_window.geometry("300x300")
        pmr_window.transient(self.root)
        pmr_window.grab_set()
    
        tk.Label(pmr_window, text="Select a PMR Number:").pack(pady=10)
    
        # Create listbox for PMRs
        pmr_listbox = tk.Listbox(pmr_window, height=10, width=20)
        pmr_listbox.pack(pady=5, fill='both', expand=True, padx=10)
    
        # Add PMRs to listbox
        for pmr in sorted(pmr_folders, key=lambda x: int(x) if x.isdigit() else 0):
            pmr_listbox.insert(tk.END, pmr)
    
        # Function to handle selection
        def select_pmr():
            selection = pmr_listbox.curselection()
            if selection:
                selected_pmr = pmr_listbox.get(selection[0])
                self.status_pmr_var.set(selected_pmr)
                # Trigger a scan with the selected PMR
                pmr_window.destroy()
                self.scan_subset_files()
    
        # Add select button
        select_button = tk.Button(pmr_window, text="Select", command=select_pmr)
        select_button.pack(pady=10)

    def init_analytics_placeholder(self):
        """Initialize empty analytics area with instructions"""
        # Clear any existing widgets
        for widget in self.analytics_charts_frame.winfo_children():
            widget.destroy()
    
        # Add placeholder message
        placeholder = tk.Label(self.analytics_charts_frame, 
                               text="Click 'Generate Analytics' to view charts and statistics about your RTVM subsets",
                               font=("Helvetica", 12))
        placeholder.pack(expand=True, fill='both')

    def scan_subset_files(self):
        """
        Scan the selected base location for subset files and populate the status table
        """
        if not self.selected_base_path:
            messagebox.showerror("Error", "Please select a base location first")
            return
    
        # Get PMR number
        pmr_number = self.status_pmr_var.get().strip()
        if not pmr_number:
            messagebox.showerror("Error", "Please enter a PMR number")
            return
    
        pmr_folder = os.path.join(self.selected_base_path, f"PMR {pmr_number}")
        if not os.path.exists(pmr_folder):
            messagebox.showerror("Error", f"PMR folder not found: {pmr_folder}")
            return
    
        # Show progress dialog
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Scanning Subset Files")
        progress_window.geometry("400x150")
        progress_window.transient(self.root)
        progress_window.grab_set()
    
        tk.Label(progress_window, text="Scanning subset files...").pack(pady=(20, 10))
    
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100, length=350)
        progress_bar.pack(pady=(10, 10), padx=20)
    
        status_var = tk.StringVar(value="Starting scan...")
        status_label = tk.Label(progress_window, textvariable=status_var)
        status_label.pack(pady=10)
    
        # Clear existing items in the table
        for item in self.subset_status_table.get_children():
            self.subset_status_table.delete(item)
    
        # Use queue for thread-safe communication
        result_queue = queue.Queue()
    
        # Define the scanning function to run in a separate thread
        def scan_worker():
            try:
                subset_files = []
                # Find all subset files
                for root, dirs, files in os.walk(pmr_folder):
                    for file in files:
                        if file.lower().endswith(('.xlsx', '.xls')) and "RTVM Subset" in file:
                            subset_files.append(os.path.join(root, file))
            
                if not subset_files:
                    result_queue.put(("error", "No subset files found"))
                    return
            
                # Update progress
                result_queue.put(("status", f"Found {len(subset_files)} subset files. Processing..."))
                result_queue.put(("max", len(subset_files)))
            
                # Process each file
                subset_stats = []
                for idx, file_path in enumerate(subset_files):
                    # Update progress
                    result_queue.put(("progress", idx + 1))
                    result_queue.put(("status", f"Processing file {idx + 1}/{len(subset_files)}: {os.path.basename(file_path)}"))
                
                    try:
                        # Extract SWBS from path (clean format for display)
                        swbs = "Unknown"
                        path_parts = file_path.split(os.sep)
                        for part in path_parts:
                            # Look for SWBS folder - should contain "SWBS" in name
                            if "SWBS" in part:
                                swbs = part
                                break
                    
                        # Get file modification time
                        mod_time = os.path.getmtime(file_path)
                        mod_time_str = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d %H:%M")
                    
                        # Open the file and analyze
                        if os.path.getsize(file_path) > 0:  # Skip empty files
                            try:
                                wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                            
                                if 'RTVM' in wb.sheetnames:
                                    ws = wb['RTVM']
                                else:
                                    ws = wb.active
                            
                                # Initialize counters with default values of 0
                                total_patterns = 0
                                add_requests = 0
                                del_requests = 0
                                detail_updates = 0
                                sat_status = 0
                                unsat_status = 0
                            
                                # Safely process each row
                                try:
                                    # Process cells in column G (index 6) - Contractor Proposed Change Request Input
                                    for row in ws.iter_rows(min_row=2):
                                        # Skip rows that don't have enough cells
                                        if len(row) <= 6:
                                            continue
                                        
                                        # Get the cell from column G (index 6)
                                        cell = row[6]
                                    
                                        # Skip empty cells
                                        if cell is None or cell.value is None:
                                            continue
                                    
                                        # Convert cell value to string and handle different types
                                        try:
                                            cell_content = str(cell.value)
                                        except Exception:
                                            # If conversion fails, skip this cell
                                            continue
                                    
                                        # Skip empty strings
                                        if not cell_content.strip():
                                            continue
                                    
                                        # Split into lines and filter out empty ones
                                        lines = [line.strip() for line in cell_content.split('\n') if line.strip()]
                                    
                                        # Count total patterns
                                        total_patterns += len(lines)
                                    
                                        # Process each pattern line
                                        for line in lines:
                                            try:
                                                # Count by pattern type
                                                if line.startswith('ADD;'):
                                                    add_requests += 1
                                                elif line.startswith('DEL;'):
                                                    del_requests += 1
                                                # Count non-ADD, non-DEL patterns that contain semicolons as detail updates
                                                elif ';' in line and not line.startswith('ADD;') and not line.startswith('DEL;'):
                                                    detail_updates += 1
                                            
                                                # Count by status (case-insensitive)
                                                line_upper = line.upper()
                                                if ';SAT' in line_upper:
                                                    sat_status += 1
                                                elif ';UNSAT' in line_upper:
                                                    unsat_status += 1
                                            except Exception:
                                                # If any error occurs processing a line, continue to the next
                                                continue
                                except Exception as row_error:
                                    # If there's an error processing rows, log it but continue with counts we have
                                    print(f"Error processing rows in {file_path}: {row_error}")
                                    # We'll still use the counts we managed to get
                            
                                # Calculate completion percentage (handle division by zero)
                                if total_patterns > 0:
                                    completion_pct = (sat_status / total_patterns) * 100
                                else:
                                    completion_pct = 0
                            
                                # Create subset info dictionary with the counts we have
                                subset_info = {
                                    "file_path": file_path,
                                    "filename": os.path.basename(file_path),  # Store just the filename separately
                                    "swbs": swbs,
                                    "modified": mod_time_str,
                                    "total_patterns": total_patterns,
                                    "add_requests": add_requests,
                                    "del_requests": del_requests,
                                    "detail_updates": detail_updates,
                                    "sat_status": sat_status,
                                    "unsat_status": unsat_status,
                                    "completion_pct": f"{completion_pct:.1f}"
                                }
                            
                                subset_stats.append(subset_info)
                            
                            except Exception as wb_error:
                                # If there's an error with the workbook, create an entry with zeros
                                print(f"Error opening workbook {file_path}: {wb_error}")
                                subset_info = {
                                    "file_path": file_path,
                                    "filename": os.path.basename(file_path),  # Store just the filename separately
                                    "swbs": swbs,
                                    "modified": mod_time_str,
                                    "total_patterns": 0,
                                    "add_requests": 0,
                                    "del_requests": 0,
                                    "detail_updates": 0,
                                    "sat_status": 0,
                                    "unsat_status": 0,
                                    "completion_pct": "0.0"
                                }
                                subset_stats.append(subset_info)
                        else:
                            # Empty file
                            subset_info = {
                                "file_path": file_path,
                                "filename": os.path.basename(file_path),  # Store just the filename separately
                                "swbs": swbs,
                                "modified": mod_time_str,
                                "total_patterns": 0,
                                "add_requests": 0,
                                "del_requests": 0,
                                "detail_updates": 0,
                                "sat_status": 0,
                                "unsat_status": 0,
                                "completion_pct": "0.0"
                            }
                            subset_stats.append(subset_info)
                
                    except Exception as e:
                        print(f"Error processing file {file_path}: {e}")
                        # Add entry with zeros for error case
                        subset_info = {
                            "file_path": file_path,
                            "filename": os.path.basename(file_path),  # Store just the filename separately
                            "swbs": swbs if 'swbs' in locals() else "Unknown",
                            "modified": mod_time_str if 'mod_time_str' in locals() else "Unknown",
                            "total_patterns": 0,
                            "add_requests": 0,
                            "del_requests": 0,
                            "detail_updates": 0,
                            "sat_status": 0,
                            "unsat_status": 0,
                            "completion_pct": "0.0"
                        }
                        subset_stats.append(subset_info)
            
                # Store subset statistics for later use
                self.subset_stats = subset_stats
            
                # Signal completion
                result_queue.put(("complete", subset_stats))
        
            except Exception as e:
                result_queue.put(("error", str(e)))
    
        # Start scanning thread
        scan_thread = threading.Thread(target=scan_worker)
        scan_thread.daemon = True
        scan_thread.start()
    
        # Function to update progress and handle completion
        def check_progress():
            try:
                # Check for messages from the worker thread
                while True:
                    message_type, data = result_queue.get_nowait()
                
                    if message_type == "status":
                        status_var.set(data)
                
                    elif message_type == "progress":
                        progress_var.set((data / progress_bar["maximum"]) * 100)
                
                    elif message_type == "max":
                        progress_bar["maximum"] = data
                
                    elif message_type == "error":
                        messagebox.showerror("Error", data)
                        progress_window.destroy()
                        return
                
                    elif message_type == "complete":
                        # Process is complete, update table with results
                        subset_stats = data
                    
                        # Update table in the main thread
                        for subset in subset_stats:
                            values = (
                                subset["filename"],  # Show just the filename instead of full path
                                subset["swbs"],      # Keep the SWBS display cleaner
                                subset["modified"],
                                subset["total_patterns"],
                                subset["add_requests"],
                                subset["del_requests"],
                                subset["detail_updates"],
                                subset["sat_status"],
                                subset["unsat_status"],
                                subset["completion_pct"]
                            )
                        
                            # Insert with color coding based on completion percentage
                            try:
                                completion = float(subset["completion_pct"])
                                if completion >= 90:
                                    tag = "complete"
                                elif completion >= 50:
                                    tag = "progress"
                                elif completion > 0:
                                    tag = "started"
                                else:
                                    tag = "not_started"
                            except:
                                tag = "error"
                        
                            self.subset_status_table.insert("", "end", values=values, tags=(tag,))
                    
                        # Configure tags for color coding
                        self.subset_status_table.tag_configure("complete", background="#a5d6a7")  # Light green
                        self.subset_status_table.tag_configure("progress", background="#fff59d")  # Light yellow
                        self.subset_status_table.tag_configure("started", background="#ffcc80")   # Light orange
                        self.subset_status_table.tag_configure("not_started", background="#ef9a9a")  # Light red
                        self.subset_status_table.tag_configure("error", background="#e0e0e0")  # Grey
                    
                        # Close progress window after a short delay
                        progress_window.after(500, progress_window.destroy)
                        return
        
            except queue.Empty:
                # Queue is empty, check again later
                progress_window.after(100, check_progress)
    
        # Start checking for progress updates
        progress_window.after(100, check_progress)


    def view_subset_patterns(self):
        """View detailed pattern breakdown for selected subset"""
        selected_items = self.subset_status_table.selection()
        if not selected_items:
            return
    
        item = selected_items[0]
        file_path = self.subset_status_table.item(item, "values")[0]
    
        try:
            # Open workbook
            wb = load_workbook(filename=file_path, read_only=True)
        
            if 'RTVM' in wb.sheetnames:
                ws = wb['RTVM']
            else:
                ws = wb.active
        
            # Collect patterns
            patterns = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > 6 and row[6]:
                    cell_content = str(row[6])
                    lines = [line.strip() for line in cell_content.split('\n') if line.strip()]
                
                    for line in lines:
                        pattern_type = "Unknown"
                        if line.startswith('ADD;'):
                            pattern_type = "ADD Request"
                        elif line.startswith('DEL;'):
                            pattern_type = "DEL Request"
                        elif ';' in line and not line.startswith('ADD;') and not line.startswith('DEL;'):
                            pattern_type = "Detail Update"
                    
                        status = "Unknown"
                        if ';SAT' in line:
                            status = "SAT"
                        elif ';UNSAT' in line:
                            status = "UNSAT"
                    
                        patterns.append({
                            "row": row_idx,
                            "type": pattern_type,
                            "status": status,
                            "pattern": line
                        })
        
            # Create detail window
            detail_window = tk.Toplevel(self.root)
            detail_window.title(f"Pattern Details: {os.path.basename(file_path)}")
            detail_window.geometry("900x600")
            detail_window.transient(self.root)
        
            # Create table frame
            table_frame = tk.Frame(detail_window)
            table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
            # Create table
            columns = ("Row", "Pattern Type", "Status", "Pattern")
            pattern_table = ttk.Treeview(table_frame, columns=columns, show="headings")
        
            # Configure columns
            pattern_table.heading("Row", text="Excel Row")
            pattern_table.heading("Pattern Type", text="Pattern Type")
            pattern_table.heading("Status", text="Status")
            pattern_table.heading("Pattern", text="Pattern")
        
            # Configure column widths
            pattern_table.column("Row", width=80, anchor="center")
            pattern_table.column("Pattern Type", width=120)
            pattern_table.column("Status", width=80)
            pattern_table.column("Pattern", width=600)
        
            # Add scrollbars
            y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=pattern_table.yview)
            x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=pattern_table.xview)
            pattern_table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
            # Pack table and scrollbars
            y_scroll.pack(side="right", fill="y")
            x_scroll.pack(side="bottom", fill="x")
            pattern_table.pack(side="left", fill="both", expand=True)
        
            # Add control buttons
            button_frame = tk.Frame(detail_window)
            button_frame.pack(fill="x", padx=10, pady=10)
        
            # Add filter options
            filter_frame = tk.Frame(button_frame)
            filter_frame.pack(side="left")
        
            tk.Label(filter_frame, text="Filter:").pack(side="left", padx=5)
        
            # Filter by type
            type_var = tk.StringVar(value="All")
            type_dropdown = ttk.Combobox(filter_frame, textvariable=type_var, 
                                          values=["All", "ADD Request", "DEL Request", "Detail Update"],
                                          width=15, state="readonly")
            type_dropdown.pack(side="left", padx=5)
        
            # Filter by status
            status_var = tk.StringVar(value="All")
            status_dropdown = ttk.Combobox(filter_frame, textvariable=status_var,
                                            values=["All", "SAT", "UNSAT"],
                                            width=10, state="readonly")
            status_dropdown.pack(side="left", padx=5)
        
            # Apply filter button
            def apply_filter():
                # Clear table
                for item in pattern_table.get_children():
                    pattern_table.delete(item)
            
                # Get filter values
                pattern_type = type_var.get()
                status = status_var.get()
            
                # Apply filters
                filtered_patterns = patterns
                if pattern_type != "All":
                    filtered_patterns = [p for p in filtered_patterns if p["type"] == pattern_type]
                if status != "All":
                    filtered_patterns = [p for p in filtered_patterns if p["status"] == status]
            
                # Populate table with filtered data
                for pattern in filtered_patterns:
                    values = (pattern["row"], pattern["type"], pattern["status"], pattern["pattern"])
                    pattern_table.insert("", "end", values=values, tags=(pattern["status"].lower(),))
            
                # Update count
                count_label.config(text=f"Showing {len(filtered_patterns)} of {len(patterns)} patterns")
        
            filter_button = tk.Button(filter_frame, text="Apply Filter", command=apply_filter)
            filter_button.pack(side="left", padx=5)
        
            # Export button
            export_button = tk.Button(button_frame, text="Export to Excel", 
                                      command=lambda: self.export_patterns_to_excel(patterns, file_path))
            export_button.pack(side="right", padx=5)
        
            # Count label
            count_label = tk.Label(button_frame, text=f"Total patterns: {len(patterns)}")
            count_label.pack(side="right", padx=10)
        
            # Populate table
            for pattern in patterns:
                values = (pattern["row"], pattern["type"], pattern["status"], pattern["pattern"])
                pattern_table.insert("", "end", values=values, tags=(pattern["status"].lower(),))
        
            # Configure tags for color coding
            pattern_table.tag_configure("sat", background="#a5d6a7")  # Light green
            pattern_table.tag_configure("unsat", background="#ffcc80")  # Light orange
            pattern_table.tag_configure("unknown", background="#e0e0e0")  # Grey
    
        except Exception as e:
            messagebox.showerror("Error", f"Could not read pattern data: {str(e)}")

    def export_patterns_to_excel(self, patterns, file_path):
        """Export the pattern details to an Excel file"""
        if not patterns:
            messagebox.showinfo("No Data", "No patterns to export")
            return
    
        # Ask user for save location
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Patterns_{os.path.basename(file_path)}"
        )
    
        if not save_path:
            return
    
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pattern Analysis"
        
            # Add headers
            headers = ["Excel Row", "Pattern Type", "Status", "Pattern"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Add data
            for row, pattern in enumerate(patterns, start=2):
                ws.cell(row=row, column=1, value=pattern["row"])
                ws.cell(row=row, column=2, value=pattern["type"])
                ws.cell(row=row, column=3, value=pattern["status"])
                ws.cell(row=row, column=4, value=pattern["pattern"])
            
                # Apply conditional formatting
                if pattern["status"] == "SAT":
                    ws.cell(row=row, column=3).fill = PatternFill(start_color="a5d6a7", end_color="a5d6a7", fill_type="solid")
                elif pattern["status"] == "UNSAT":
                    ws.cell(row=row, column=3).fill = PatternFill(start_color="ffcc80", end_color="ffcc80", fill_type="solid")
        
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 100)
        
            # Add summary on a second sheet
            ws_summary = wb.create_sheet(title="Summary")
        
            # Add summary headers
            summary_headers = ["Metric", "Count", "Percentage"]
            for col, header in enumerate(summary_headers, start=1):
                ws_summary.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Total patterns
            total = len(patterns)
            ws_summary.cell(row=2, column=1, value="Total Patterns")
            ws_summary.cell(row=2, column=2, value=total)
        
            # Pattern types
            type_counts = {"ADD Request": 0, "DEL Request": 0, "Detail Update": 0, "Unknown": 0}
            for pattern in patterns:
                type_counts[pattern["type"]] += 1
        
            row = 3
            for pattern_type, count in type_counts.items():
                if count > 0:
                    ws_summary.cell(row=row, column=1, value=f"{pattern_type} Count")
                    ws_summary.cell(row=row, column=2, value=count)
                    ws_summary.cell(row=row, column=3, value=f"{(count/total)*100:.1f}%")
                    row += 1
        
            # Status counts
            status_counts = {"SAT": 0, "UNSAT": 0, "Unknown": 0}
            for pattern in patterns:
                status_counts[pattern["status"]] += 1
        
            row += 1
            ws_summary.cell(row=row, column=1, value="Status Breakdown").font = Font(bold=True)
            row += 1
        
            for status, count in status_counts.items():
                if count > 0:
                    ws_summary.cell(row=row, column=1, value=f"{status} Count")
                    ws_summary.cell(row=row, column=2, value=count)
                    ws_summary.cell(row=row, column=3, value=f"{(count/total)*100:.1f}%")
                    row += 1
        
            # Auto-adjust column widths for summary
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
                adjusted_width = (max_length + 2) * 1.2
                ws_summary.column_dimensions[column].width = adjusted_width
        
            # Save the workbook
            wb.save(save_path)
            messagebox.showinfo("Export Complete", f"Pattern details exported to {save_path}")
    
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export patterns: {str(e)}")

    def check_subset_issues(self):
        """Check the selected subset for potential issues"""
        selected_items = self.subset_status_table.selection()
        if not selected_items:
            return
    
        item = selected_items[0]
        file_path = self.subset_status_table.item(item, "values")[0]
    
        try:
            # Open workbook
            wb = load_workbook(filename=file_path, read_only=True)
        
            if 'RTVM' in wb.sheetnames:
                ws = wb['RTVM']
            else:
                ws = wb.active
        
            # Define potential issues
            issues = []
        
            # Rows with patterns but no comments
            # Rows with TBD status
            # Rows with missing data in patterns
            # Patterns with invalid format
        
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) <= 7:
                    continue
                
                # Get patterns from column G
                cell_content = str(row[6]) if row[6] is not None else ""
                pattern_lines = [line.strip() for line in cell_content.split('\n') if line.strip()]
            
                # Get comments from column H
                comment_content = str(row[7]) if row[7] is not None else ""
            
                # Check for patterns without comments
                if pattern_lines and not comment_content.strip():
                    issues.append({
                        "row": row_idx,
                        "type": "Missing Comment",
                        "description": f"Row has {len(pattern_lines)} patterns but no comments"
                    })
            
                # Check each pattern for issues
                for line in pattern_lines:
                    # Check for TBD status
                    if ';TBD' in line:
                        issues.append({
                            "row": row_idx,
                            "type": "TBD Status",
                            "description": f"Pattern contains TBD status: {line}"
                        })
                
                    # Check for patterns with spaces around semicolons
                    if re.search(r"\s+;|;\s+", line):
                        issues.append({
                            "row": row_idx,
                            "type": "Format Error",
                            "description": f"Pattern has spaces around semicolons: {line}"
                        })
                
                    # Validate ADD patterns
                    if line.startswith('ADD;'):
                        parts = line.split(';')
                        if len(parts) < 2:
                            issues.append({
                                "row": row_idx,
                                "type": "Incomplete Pattern",
                                "description": f"ADD pattern missing DI Number: {line}"
                            })
                        elif len(parts) >= 4 and parts[3].strip() not in ["SAT", "UNSAT"]:
                            issues.append({
                                "row": row_idx,
                                "type": "Invalid Status",
                                "description": f"ADD pattern has invalid status: {line}"
                            })
                
                    # Validate DEL patterns
                    elif line.startswith('DEL;'):
                        parts = line.split(';')
                        if len(parts) != 2 or not parts[1].strip():
                            issues.append({
                                "row": row_idx,
                                "type": "Invalid DEL Pattern",
                                "description": f"DEL pattern should have format DEL;WCC-VERI-DOC-XXXX: {line}"
                            })
                
                    # Validate update patterns
                    elif ';' in line and not line.startswith('ADD;') and not line.startswith('DEL;'):
                        parts = line.split(';')
                        if not parts[0].startswith('WCC-VERI-DOC-'):
                            issues.append({
                                "row": row_idx,
                                "type": "Invalid VeriDoc",
                                "description": f"VeriDoc number should start with WCC-VERI-DOC-: {line}"
                            })
                    
                        if len(parts) >= 3 and parts[2].strip() not in ["SAT", "UNSAT"]:
                            issues.append({
                                "row": row_idx,
                                "type": "Invalid Status",
                                "description": f"Update pattern has invalid status: {line}"
                            })
        
            # Show issues in a new window
            issues_window = tk.Toplevel(self.root)
            issues_window.title(f"Issues in {os.path.basename(file_path)}")
            issues_window.geometry("900x600")
            issues_window.transient(self.root)
        
            # Create table frame
            table_frame = tk.Frame(issues_window)
            table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
            # Create table
            columns = ("Row", "Issue Type", "Description")
            issues_table = ttk.Treeview(table_frame, columns=columns, show="headings")
        
            # Configure columns
            issues_table.heading("Row", text="Excel Row")
            issues_table.heading("Issue Type", text="Issue Type")
            issues_table.heading("Description", text="Description")
        
            issues_table.column("Row", width=80, anchor="center")
            issues_table.column("Issue Type", width=150)
            issues_table.column("Description", width=650)
        
            # Add scrollbars
            y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=issues_table.yview)
            issues_table.configure(yscrollcommand=y_scroll.set)
        
            # Pack table and scrollbar
            y_scroll.pack(side="right", fill="y")
            issues_table.pack(side="left", fill="both", expand=True)
        
            # Add issues to table
            if issues:
                for issue in issues:
                    values = (issue["row"], issue["type"], issue["description"])
                    tag = issue["type"].lower().replace(" ", "_")
                    issues_table.insert("", "end", values=values, tags=(tag,))
            
                # Configure tags for color coding
                issues_table.tag_configure("tbd_status", background="#fff59d")  # Light yellow
                issues_table.tag_configure("missing_comment", background="#ffcc80")  # Light orange
                issues_table.tag_configure("format_error", background="#ef9a9a")  # Light red
                issues_table.tag_configure("incomplete_pattern", background="#ef9a9a")  # Light red
                issues_table.tag_configure("invalid_status", background="#ef9a9a")  # Light red
                issues_table.tag_configure("invalid_del_pattern", background="#ef9a9a")  # Light red
                issues_table.tag_configure("invalid_veridoc", background="#ef9a9a")  # Light red
            else:
                # No issues found
                no_issues_label = tk.Label(issues_window, text="No issues found in this subset", font=("Helvetica", 14))
                no_issues_label.pack(expand=True)
            
            # Add export button
            button_frame = tk.Frame(issues_window)
            button_frame.pack(fill="x", padx=10, pady=10)
        
            # Summary label
            summary_label = tk.Label(button_frame, text=f"Found {len(issues)} issues")
            summary_label.pack(side="left", padx=10)
        
            # Export button
            if issues:
                export_button = tk.Button(button_frame, text="Export Issues", 
                                        command=lambda: self.export_issues_to_excel(issues, file_path))
                export_button.pack(side="right", padx=5)
    
        except Exception as e:
            messagebox.showerror("Error", f"Could not check subset for issues: {str(e)}")

    def export_issues_to_excel(self, issues, file_path):
        """Export the issues list to an Excel file"""
        if not issues:
            messagebox.showinfo("No Data", "No issues to export")
            return
    
        # Ask user for save location
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Issues_{os.path.basename(file_path)}"
        )
    
        if not save_path:
            return
    
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Identified Issues"
        
            # Add headers
            headers = ["Excel Row", "Issue Type", "Description"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Add data
            for row, issue in enumerate(issues, start=2):
                ws.cell(row=row, column=1, value=issue["row"])
                ws.cell(row=row, column=2, value=issue["type"])
                ws.cell(row=row, column=3, value=issue["description"])
            
                # Apply conditional formatting
                if issue["type"] == "TBD Status":
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="fff59d", end_color="fff59d", fill_type="solid")
                elif issue["type"] == "Missing Comment":
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="ffcc80", end_color="ffcc80", fill_type="solid")
                else:
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="ef9a9a", end_color="ef9a9a", fill_type="solid")
        
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 100)
            
            # Add a summary on a second sheet
            ws_summary = wb.create_sheet(title="Summary")
        
            # Add headers
            summary_headers = ["Issue Type", "Count", "Percentage"]
            for col, header in enumerate(summary_headers, start=1):
                ws_summary.cell(row=1, column=col, value=header).font = Font(bold=True)
        
            # Count issues by type
            issue_types = {}
            for issue in issues:
                issue_type = issue["type"]
                if issue_type not in issue_types:
                    issue_types[issue_type] = 0
                issue_types[issue_type] += 1
        
            # Add data
            total_issues = len(issues)
            for row, (issue_type, count) in enumerate(issue_types.items(), start=2):
                ws_summary.cell(row=row, column=1, value=issue_type)
                ws_summary.cell(row=row, column=2, value=count)
                percentage = (count / total_issues) * 100
                ws_summary.cell(row=row, column=3, value=f"{percentage:.1f}%")
            
                # Conditional formatting
                if issue_type == "TBD Status":
                    ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="fff59d", end_color="fff59d", fill_type="solid")
                elif issue_type == "Missing Comment":
                    ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="ffcc80", end_color="ffcc80", fill_type="solid")
                else:
                    ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="ef9a9a", end_color="ef9a9a", fill_type="solid")
        
            # Add total row
            row = len(issue_types) + 2
            ws_summary.cell(row=row, column=1, value="Total Issues").font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=total_issues).font = Font(bold=True)
            ws_summary.cell(row=row, column=3, value="100%").font = Font(bold=True)
        
            # Auto-adjust column widths for summary
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
                adjusted_width = (max_length + 2) * 1.2
                ws_summary.column_dimensions[column].width = adjusted_width
        
            # Save the workbook
            wb.save(save_path)
            messagebox.showinfo("Export Complete", f"Issues exported to {save_path}")
    
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export issues: {str(e)}")

    def mark_subset_status(self, status):
        """Mark the selected subset with a status for tracking purposes"""
        selected_items = self.subset_status_table.selection()
        if not selected_items:
            return
    
        item = selected_items[0]
        filename = self.subset_status_table.item(item, "values")[0]  # Get the filename from the UI
    
        # We now need to find the full path from subset_stats since we're only displaying filenames
        file_path = None
        for stat in self.subset_stats:
            if stat["filename"] == filename:
                file_path = stat["file_path"]
                break
    
        if not file_path:
            messagebox.showerror("Error", f"Could not find full path for: {filename}")
            return
    
        # Create a status marker file or update existing
        try:
            marker_file = os.path.join(os.path.dirname(file_path), ".subset_status.json")
        
            # Load existing status data if available
            status_data = {}
            if os.path.exists(marker_file):
                with open(marker_file, 'r') as f:
                    status_data = json.load(f)
        
            # Update status for this file
            status_data[os.path.basename(file_path)] = {
                "status": status,
                "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "by": os.getenv('USERNAME', 'unknown')
            }
        
            # Save back to file
            with open(marker_file, 'w') as f:
                json.dump(status_data, f, indent=2)
        
            # Update the UI
            if status == "complete":
                self.subset_status_table.item(item, tags=("complete",))
            elif status == "in_progress":
                self.subset_status_table.item(item, tags=("progress",))
        
            messagebox.showinfo("Status Updated", f"Marked subset as {status}")
    
        except Exception as e:
            messagebox.showerror("Error", f"Could not update status: {str(e)}")





##################################################################################################################################################


    def open_rvtm_subset_management_window(self):
        self.rvtm_subset_window = tk.Toplevel(self.root)
        """
        Open the enhanced RTVM Subset Management window
        """
        # Call the new enhanced version
        self.enhance_rvtm_subset_management_window()




    def combine_and_update_subsets(self):
        """
        This function:
        1. Finds all subset files in the base location
        2. Combines their content into the main file
        3. Handles conflicts when the same requirement appears in multiple subsets
        4. Creates updated subsets from the newly combined main file
        """
        if not self.selected_base_path:
            messagebox.showerror("Error", "No base location selected. Please select a base location first.")
            return

        if not self.excel_file_path:
            messagebox.showerror("Error", "No main Excel file loaded. Please upload a main file first.")
            return

        # Step 1: Select folder containing subsets
        subset_folder = filedialog.askdirectory(
            title="Select Folder Containing RTVM Subsets to Combine", 
            initialdir=self.selected_base_path
        )
        if not subset_folder:
            return

        # Step 2: Find all the subset Excel files
        subset_files = []
        for root, dirs, files in os.walk(subset_folder):
            for file in files:
                if file.endswith(('.xlsx', '.xls')) and 'RTVM Subset' in file:
                    subset_files.append(os.path.join(root, file))

        if not subset_files:
            messagebox.showinfo("No Subsets Found", "No RTVM subset files were found in the selected folder.")
            return

        # Step 3: Show progress dialog and start the combination process
        self.show_combine_update_progress(subset_files)

    def show_combine_update_progress(self, subset_files):
        """Creates a progress dialog and starts the combination process in a background thread."""
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Combining and Updating Subsets")
        progress_window.geometry("500x250")
        progress_window.transient(self.root)
        progress_window.grab_set()  # Make the window modal

        # Status label
        status_var = tk.StringVar(value="Preparing to combine subset files...")
        status_label = tk.Label(progress_window, textvariable=status_var)
        status_label.pack(pady=(20, 10))

        # Progress bar
        progress_var = tk.DoubleVar(value=0)
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100, length=450)
        progress_bar.pack(pady=10, padx=20)

        # Detail label for current operation
        detail_var = tk.StringVar(value="")
        detail_label = tk.Label(progress_window, textvariable=detail_var, font=("Helvetica", 9))
        detail_label.pack(pady=10)

        # Create a queue for thread-safe communication
        result_queue = queue.Queue()

        # Start the combination process in a background thread
        threading.Thread(
            target=self.process_combine_and_update,
            args=(subset_files, result_queue),
            daemon=True
        ).start()

        # Update function to handle messages from the worker thread
        def update_progress():
            try:
                while True:
                    message_type, data = result_queue.get_nowait()
                
                    if message_type == 'status':
                        status_var.set(data)
                
                    elif message_type == 'detail':
                        detail_var.set(data)
                
                    elif message_type == 'progress':
                        current, total = data
                        progress_percent = (current / total) * 100 if total > 0 else 0
                        progress_var.set(progress_percent)
                
                    elif message_type == 'conflict':
                        # Handle conflict in the main thread
                        self.handle_requirement_conflict(data, result_queue)
                
                    elif message_type == 'error':
                        messagebox.showerror("Error", data)
                        progress_window.destroy()
                        return
                
                    elif message_type == 'complete':
                        status_var.set("Complete! Creating updated subsets...")
                        progress_var.set(100)
                        # Start creating updated subsets
                        threading.Thread(
                            target=self.create_updated_subsets,
                            args=(result_queue, data),  # data contains the path to the combined file
                            daemon=True
                        ).start()
                
                    elif message_type == 'all_complete':
                        status_var.set("All operations complete!")
                        progress_var.set(100)
                    
                        # Close the progress window after a short delay
                        progress_window.after(2000, progress_window.destroy)
                    
                        # Show completion message
                        messagebox.showinfo(
                            "Operation Complete", 
                            "Successfully combined subsets and created updated subset files."
                        )
                        return
                    
            except queue.Empty:
                # Schedule another check
                progress_window.after(100, update_progress)
        
        # Start the progress update loop
        progress_window.after(100, update_progress)

    def process_combine_and_update(self, subset_files, result_queue):
        """
        Worker function to process the subset files and combine them.
        This runs in a background thread.
        """
        try:
            # Load the main file to get the structure
            result_queue.put(('status', "Loading main Excel file..."))
            main_wb = load_workbook(filename=self.excel_file_path)
            main_ws = None
            if 'RTVM' in main_wb.sheetnames:
                main_ws = main_wb['RTVM']
            else:
                main_ws = main_wb.active
        
            # Create a map of DOORS SPEC ID to row in the main file
            result_queue.put(('status', "Building index of main file..."))
            id_to_row = {}
            for row_idx, row in enumerate(main_ws.iter_rows(min_row=2, values_only=False), start=2):
                doors_spec_id = row[0].value
                if doors_spec_id:
                    id_to_row[str(doors_spec_id)] = row_idx
        
            # Create a temporary file to build the combined result
            result_queue.put(('status', "Creating temporary file for combined data..."))
            combined_file = os.path.join(self.selected_base_path, "temp_combined_rtvm.xlsx")
            shutil.copy2(self.excel_file_path, combined_file)
        
            # Load the combined workbook
            result_queue.put(('status', "Loading temporary combined file..."))
            combined_wb = load_workbook(filename=combined_file)
            if 'RTVM' in combined_wb.sheetnames:
                combined_ws = combined_wb['RTVM']
            else:
                combined_ws = combined_wb.active
            
            # Track conflicts that need user resolution
            conflicts = []
            doors_id_to_subset_data = {}
            
            # Process each subset file
            total_files = len(subset_files)
            for idx, subset_file in enumerate(subset_files, start=1):
                result_queue.put(('progress', (idx, total_files)))
                result_queue.put(('detail', f"Processing file {idx}/{total_files}: {os.path.basename(subset_file)}"))
            
                try:
                    # Load the subset workbook
                    subset_wb = load_workbook(filename=subset_file)
                    if 'RTVM' in subset_wb.sheetnames:
                        subset_ws = subset_wb['RTVM']
                    else:
                        subset_ws = subset_wb.active
                
                    # Process each row in the subset
                    for subset_row in subset_ws.iter_rows(min_row=2, values_only=False):
                        doors_spec_id = subset_row[0].value
                        if not doors_spec_id:
                            continue
                        
                        doors_spec_id = str(doors_spec_id)
                    
                        # Check if this DOORS SPEC ID exists in the main file
                        if doors_spec_id in id_to_row:
                            main_row_idx = id_to_row[doors_spec_id]
                        
                            # Get the change request input (column G, index 6)
                            subset_change_request = subset_row[6].value
                            if subset_change_request:
                                if doors_spec_id not in doors_id_to_subset_data:
                                    doors_id_to_subset_data[doors_spec_id] = []
                                
                                doors_id_to_subset_data[doors_spec_id].append({
                                    'file': os.path.basename(subset_file),
                                    'change_request': subset_change_request,
                                    'comment_input': subset_row[7].value if len(subset_row) > 7 else None
                                })
                
                    subset_wb.close()
                
                except Exception as e:
                    result_queue.put(('detail', f"Error processing {os.path.basename(subset_file)}: {str(e)}"))
                    continue
        
            # Now check for conflicts and resolve them
            result_queue.put(('status', "Checking for conflicts..."))
        
            # Resolve conflicts for each DOORS SPEC ID
            for doors_spec_id, subset_entries in doors_id_to_subset_data.items():
                if len(subset_entries) > 1:
                    # There are multiple entries for this DOORS SPEC ID
                    result_queue.put(('detail', f"Found conflict for SPEC ID: {doors_spec_id}"))
                
                    # See if entries actually conflict
                    has_conflict = False
                    base_change_request = subset_entries[0]['change_request']
                    for entry in subset_entries[1:]:
                        if entry['change_request'] != base_change_request:
                            has_conflict = True
                            break
                        
                    if has_conflict:
                        # Send conflict to main thread for resolution
                        result_queue.put(('conflict', {
                            'doors_spec_id': doors_spec_id,
                            'entries': subset_entries
                        }))
                    
                        # Wait for response with the resolved value
                        while True:
                            message_type, data = result_queue.get()
                            if message_type == 'conflict_resolved':
                                resolved_entry = data
                                break
                            
                        # Use the resolved value
                        main_row_idx = id_to_row[doors_spec_id]
                        combined_ws.cell(row=main_row_idx, column=7, value=resolved_entry['change_request'])
                        if resolved_entry.get('comment_input'):
                            combined_ws.cell(row=main_row_idx, column=8, value=resolved_entry['comment_input'])
                    else:
                        # No actual conflict, values are the same - use the first one
                        main_row_idx = id_to_row[doors_spec_id]
                        combined_ws.cell(row=main_row_idx, column=7, value=base_change_request)
                        if subset_entries[0].get('comment_input'):
                            combined_ws.cell(row=main_row_idx, column=8, value=subset_entries[0]['comment_input'])
                else:
                    # Only one entry - no conflict
                    entry = subset_entries[0]
                    main_row_idx = id_to_row[doors_spec_id]
                    combined_ws.cell(row=main_row_idx, column=7, value=entry['change_request'])
                    if entry.get('comment_input'):
                        combined_ws.cell(row=main_row_idx, column=8, value=entry['comment_input'])
        
            # Save the combined workbook
            result_queue.put(('status', "Saving combined file..."))
            combined_wb.save(combined_file)
            combined_wb.close()
        
            # Replace the main file with the combined file
            result_queue.put(('status', "Updating main file..."))
            os.replace(combined_file, self.excel_file_path)
        
            # Reload the main file in the app
            result_queue.put(('detail', "Reloading main file..."))
            self.df = pd.read_excel(self.excel_file_path)
        
            # Signal completion and provide the path to the updated main file
            result_queue.put(('complete', self.excel_file_path))
        
        except Exception as e:
            result_queue.put(('error', f"Error during combination process: {str(e)}"))

    def handle_requirement_conflict(self, conflict_data, result_queue):
        """
        Shows a dialog to resolve a conflict between different versions of the same requirement.
        This runs in the main thread.
    
        The enhanced version includes:
        - Detailed conflict analysis
        - Color-coded line comparison
        - Conflict categorization
        - Line-by-line comparison for better decision-making
        """
        doors_spec_id = conflict_data['doors_spec_id']
        entries = conflict_data['entries']
    
        # Create conflict resolution dialog
        conflict_window = tk.Toplevel(self.root)
        conflict_window.title(f"Resolve Conflict for SPEC ID: {doors_spec_id}")
        conflict_window.geometry("900x700")  # Larger window for enhanced UI
        conflict_window.transient(self.root)
        conflict_window.grab_set()  # Make the window modal
    
        # Add explanation
        header_frame = tk.Frame(conflict_window)
        header_frame.pack(fill="x", padx=10, pady=10)
    
        tk.Label(
            header_frame, 
            text=f"Conflicting Requirements for SPEC ID: {doors_spec_id}",
            font=("Helvetica", 12, "bold")
        ).pack(side="top", anchor="w")
    
        tk.Label(
            header_frame,
            text="This requirement has different change requests in multiple subset files.\nPlease review the conflicts and select how to resolve them.",
            justify="left"
        ).pack(side="top", anchor="w", pady=5)
    
        # Add a notebook for tabs
        notebook = ttk.Notebook(conflict_window)
        notebook.pack(fill="both", expand=True, padx=10, pady=5)
    
        # Tab 1: Overview and Recommended Action
        overview_frame = tk.Frame(notebook)
        notebook.add(overview_frame, text="Overview & Recommended Action")
    
        # First, analyze the conflicts
        if len(entries) >= 2:
            # Find the first two entries with different change requests to analyze
            entry1 = entries[0]
            entry2 = None
            for e in entries[1:]:
                if e['change_request'] != entry1['change_request']:
                    entry2 = e
                    break
                
            if entry2 is None:
                # Fallback if no actual conflict is found (shouldn't happen)
                entry2 = entries[1]
            
            # Run conflict analysis
            conflict_analysis = self.analyze_conflicts(
                entry1['change_request'] or "", 
                entry2['change_request'] or ""
            )
        else:
            # Not enough entries for meaningful analysis
            conflict_analysis = {
                'common_lines_count': 0,
                'unique_to_1_count': 0, 
                'unique_to_2_count': 0,
                'operations': {'add': {'1': [], '2': [], 'common': []},
                              'del': {'1': [], '2': [], 'common': []},
                              'update': {'1': [], '2': [], 'common': []}},
                'veridoc_conflicts': [],
                'merged_lines': []
            }
    
        # Extract all unique lines from all change requests for merging
        all_lines = set()
        for entry in entries:
            if entry['change_request']:
                lines = entry['change_request'].split('\n')
                for line in lines:
                    if line.strip():
                        all_lines.add(line)
    
        merged_text = "\n".join(sorted(all_lines))
    
        # Create overview summary
        overview_summary_frame = tk.LabelFrame(overview_frame, text="Conflict Summary")
        overview_summary_frame.pack(fill="x", padx=5, pady=5)
    
        # Build summary text
        summary_text = (
            f"Files with different values: {len(entries)}\n"
            f"Total unique change request lines: {len(all_lines)}\n"
            f"Common lines to all files: {conflict_analysis['common_lines_count']}\n"
            f"Lines unique to first file: {conflict_analysis['unique_to_1_count']}\n"
            f"Lines unique to second file: {conflict_analysis['unique_to_2_count']}\n\n"
        )
    
        # Add VeriDoc conflict details if any
        if conflict_analysis['veridoc_conflicts']:
            summary_text += "VeriDoc Conflicts Detected:\n"
            for i, conflict in enumerate(conflict_analysis['veridoc_conflicts']):
                summary_text += f"- Conflict {i+1}: {conflict['veridoc']} - {conflict['conflict_type']}\n"
            summary_text += "\n"
    
        summary_text += "Operations breakdown:\n"
        summary_text += f"- ADD operations: {len(conflict_analysis['operations']['add']['common'])} common, "
        summary_text += f"{len(conflict_analysis['operations']['add']['1'])} unique to first file, "
        summary_text += f"{len(conflict_analysis['operations']['add']['2'])} unique to second file\n"
    
        summary_text += f"- DEL operations: {len(conflict_analysis['operations']['del']['common'])} common, "
        summary_text += f"{len(conflict_analysis['operations']['del']['1'])} unique to first file, "
        summary_text += f"{len(conflict_analysis['operations']['del']['2'])} unique to second file\n"
    
        summary_text += f"- UPDATE operations: {len(conflict_analysis['operations']['update']['common'])} common, "
        summary_text += f"{len(conflict_analysis['operations']['update']['1'])} unique to first file, "
        summary_text += f"{len(conflict_analysis['operations']['update']['2'])} unique to second file\n"
    
        summary_label = tk.Label(overview_summary_frame, text=summary_text, justify="left")
        summary_label.pack(fill="x", padx=10, pady=5, anchor="w")
    
        # Recommendation section
        recommendation_frame = tk.LabelFrame(overview_frame, text="Recommended Action")
        recommendation_frame.pack(fill="x", padx=5, pady=5)
    
        # Get recommendation based on conflict analysis
        recommendation = ""
    
        if conflict_analysis['veridoc_conflicts']:
            recommendation = (
                "CONFLICTS DETECTED: There are conflicting operations for the same VeriDoc numbers.\n"
                "Recommendation: Review the individual conflicts in the 'Detailed Comparison' tab and consider manual resolution.\n"
                "WARNING: Automatic merging might include conflicting operations."
            )
        elif conflict_analysis['unique_to_1_count'] > 0 and conflict_analysis['unique_to_2_count'] > 0:
            recommendation = (
                "PARTIAL CONFLICTS: Each file contains unique change requests not found in the other.\n"
                "Recommendation: Merge all values to include all unique operations from each file."
            )
        elif conflict_analysis['unique_to_1_count'] == 0 and conflict_analysis['unique_to_2_count'] > 0:
            recommendation = (
                "FILE 2 CONTAINS ADDITIONAL DATA: The second file includes all operations from the first file plus additional ones.\n"
                "Recommendation: Use the second file as it contains more information."
            )
        elif conflict_analysis['unique_to_1_count'] > 0 and conflict_analysis['unique_to_2_count'] == 0:
            recommendation = (
                "FILE 1 CONTAINS ADDITIONAL DATA: The first file includes all operations from the second file plus additional ones.\n"
                "Recommendation: Use the first file as it contains more information."
            )
        else:
            recommendation = (
                "NO SIGNIFICANT CONFLICT: The files appear to have matching operations.\n"
                "Recommendation: Use either file as they contain equivalent information."
            )
    
        recommendation_label = tk.Label(
            recommendation_frame, 
            text=recommendation, 
            justify="left",
            wraplength=800
        )
        recommendation_label.pack(fill="x", padx=10, pady=5, anchor="w")
    
        # Create radio buttons for selection
        option_var = tk.StringVar(value="merge")  # Default to merge
    
        # Options frame
        options_frame = tk.LabelFrame(overview_frame, text="Resolution Options")
        options_frame.pack(fill="both", expand=True, padx=5, pady=5)
    
        # Merge option
        merge_frame = tk.Frame(options_frame)
        merge_frame.pack(fill="x", padx=5, pady=5)
    
        merge_radio = tk.Radiobutton(
            merge_frame, 
            text="Merge all values (combines all unique change requests from all files)", 
            variable=option_var, 
            value="merge",
            font=("Helvetica", 10, "bold")
        )
        merge_radio.pack(side="left", anchor="w", padx=5)
    
        # Preview and choose from individual files
        files_frame = tk.LabelFrame(options_frame, text="Choose Individual File")
        files_frame.pack(fill="both", expand=True, padx=5, pady=5)
    
        # Create a frame with scrollable content for file options
        file_options_canvas = tk.Canvas(files_frame)
        file_options_scrollbar = ttk.Scrollbar(files_frame, orient="vertical", command=file_options_canvas.yview)
        file_options_scrollable_frame = ttk.Frame(file_options_canvas)
    
        file_options_scrollable_frame.bind(
            "<Configure>",
            lambda e: file_options_canvas.configure(scrollregion=file_options_canvas.bbox("all"))
        )
    
        file_options_canvas.create_window((0, 0), window=file_options_scrollable_frame, anchor="nw")
        file_options_canvas.configure(yscrollcommand=file_options_scrollbar.set)
    
        file_options_canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        file_options_scrollbar.pack(side="right", fill="y")
    
        # Add radio buttons and previews for each file
        for i, entry in enumerate(entries):
            file_frame = tk.Frame(file_options_scrollable_frame)
            file_frame.pack(fill="x", padx=2, pady=5)
        
            radio = tk.Radiobutton(
                file_frame, 
                text=f"Use only values from: {entry['file']}", 
                variable=option_var, 
                value=str(i)
            )
            radio.pack(side="top", anchor="w", padx=5)
        
            preview_frame = tk.Frame(file_frame)
            preview_frame.pack(fill="x", padx=20, pady=2)
        
            preview = tk.Text(preview_frame, height=3, width=80)
            preview.pack(fill="x", side="left", expand=True)
            preview.insert("1.0", entry['change_request'] if entry['change_request'] else "")
            preview.config(state="disabled")
        
            # Add scrollbar to preview
            preview_scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=preview.yview)
            preview.configure(yscrollcommand=preview_scrollbar.set)
            preview_scrollbar.pack(side="right", fill="y")
    
        # Tab 2: Line by Line Comparison
        comparison_frame = ttk.Frame(notebook)
        notebook.add(comparison_frame, text="Detailed Comparison")
    
        # Create a tree view for line-by-line comparison
        columns = ["Line", "Origin", "Type", "Content"]
        tree = ttk.Treeview(comparison_frame, columns=columns, show="headings", selectmode="browse")
    
        # Configure column headings and widths
        tree.heading("Line", text="#")
        tree.heading("Origin", text="Origin")
        tree.heading("Type", text="Type")
        tree.heading("Content", text="Content")
    
        tree.column("Line", width=40, anchor="center")
        tree.column("Origin", width=120, anchor="center")
        tree.column("Type", width=80, anchor="center")
        tree.column("Content", width=600)
    
        tree.pack(fill="both", expand=True, side="left")
    
        # Add scrollbar to tree
        tree_scrollbar = ttk.Scrollbar(comparison_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=tree_scrollbar.set)
        tree_scrollbar.pack(side="right", fill="y")
    
        # Populate the tree with all unique lines from all files
        line_num = 0
    
        # Add headers for Common Lines (lines in both/all files)
        if conflict_analysis['common_lines_count'] > 0:
            tree.insert("", "end", values=("", "COMMON", "HEADER", "Lines that appear in all files"), tags=('header',))
            for op_type in ['add', 'del', 'update']:
                for line in conflict_analysis['operations'][op_type]['common']:
                    line_num += 1
                    tree.insert("", "end", values=(line_num, "All Files", op_type.upper(), line), tags=(op_type,))
    
        # Add headers for lines unique to file 1
        if conflict_analysis['unique_to_1_count'] > 0:
            tree.insert("", "end", values=("", entry1['file'], "HEADER", f"Lines unique to {entry1['file']}"), tags=('header',))
            for op_type in ['add', 'del', 'update']:
                for line in conflict_analysis['operations'][op_type]['1']:
                    line_num += 1
                    tree.insert("", "end", values=(line_num, entry1['file'], op_type.upper(), line), tags=(op_type,))
    
        # Add headers for lines unique to file 2
        if conflict_analysis['unique_to_2_count'] > 0 and entry2 is not None:
            tree.insert("", "end", values=("", entry2['file'], "HEADER", f"Lines unique to {entry2['file']}"), tags=('header',))
            for op_type in ['add', 'del', 'update']:
                for line in conflict_analysis['operations'][op_type]['2']:
                    line_num += 1
                    tree.insert("", "end", values=(line_num, entry2['file'], op_type.upper(), line), tags=(op_type,))
    
        # Add styles for the tree rows
        tree.tag_configure('header', background='lightgrey', font=('Helvetica', 10, 'bold'))
        tree.tag_configure('add', background='#e6ffe6')  # Light green
        tree.tag_configure('del', background='#ffe6e6')  # Light red
        tree.tag_configure('update', background='#e6e6ff')  # Light blue
    
        # Tab 3: Manual Edit
        manual_frame = ttk.Frame(notebook)
        notebook.add(manual_frame, text="Manual Edit")
    
        # Instructions
        tk.Label(
            manual_frame,
            text="You can manually edit the combined content here:",
            anchor="w"
        ).pack(fill="x", padx=10, pady=5)
    
        # Manual edit text area
        manual_text = tk.Text(manual_frame, height=20, width=80)
        manual_text.pack(fill="both", expand=True, padx=10, pady=5)
        manual_text.insert("1.0", merged_text)
    
        # Add scrollbar to manual edit
        manual_scrollbar = ttk.Scrollbar(manual_frame, orient="vertical", command=manual_text.yview)
        manual_text.configure(yscrollcommand=manual_scrollbar.set)
        manual_scrollbar.pack(side="right", fill="y")
    
        # Button frame at the bottom of the window
        button_frame = tk.Frame(conflict_window)
        button_frame.pack(fill="x", padx=10, pady=10)
    
        def resolve_conflict():
            choice = option_var.get()
        
            resolved_entry = {}
        
            if choice == "merge":
                # Use merged content
                resolved_entry = {
                    'change_request': merged_text,
                    'comment_input': None  # No comments merged for now
                }
            elif choice == "manual":
                # Use manual entry from manual edit tab
                manual_content = manual_text.get("1.0", "end-1c")
                resolved_entry = {
                    'change_request': manual_content,
                    'comment_input': None
                }
            else:
                # Use the selected entry
                idx = int(choice)
                resolved_entry = entries[idx]
        
            # Send the resolved entry back to the worker thread
            result_queue.put(('conflict_resolved', resolved_entry))
        
            # Close the dialog
            conflict_window.destroy()
    
        # "Apply Selected Resolution" button
        resolve_button = tk.Button(
            button_frame, 
            text="Apply Selected Resolution", 
            command=resolve_conflict,
            width=20,
            bg="#e6ffe6",  # Light green
            font=("Helvetica", 10, "bold")
        )
        resolve_button.pack(side="right", padx=5)
    
        # Cancel button
        cancel_button = tk.Button(
            button_frame, 
            text="Cancel", 
            command=conflict_window.destroy,
            width=10
        )
        cancel_button.pack(side="right", padx=5)




    def analyze_conflicts(self, change_request1, change_request2):
        """
        Analyzes two different change request inputs to identify specific conflicts.
        Returns a dictionary with analysis that can help users understand the conflict.
    
        Args:
            change_request1: First change request string
            change_request2: Second change request string
        
        Returns:
            Dictionary with conflict analysis
        """
        # Split into lines and create sets of lines for each change request
        lines1 = [line.strip() for line in change_request1.split('\n') if line.strip()]
        lines2 = [line.strip() for line in change_request2.split('\n') if line.strip()]
    
        set1 = set(lines1)
        set2 = set(lines2)
    
        # Find common lines
        common_lines = set1.intersection(set2)
    
        # Find lines unique to each set
        unique_to_1 = set1 - set2
        unique_to_2 = set2 - set1
    
        # Categorize by operation (ADD, DEL, WCC-VERI-DOC)
        operations = {
            'add': {'1': [], '2': [], 'common': []},
            'del': {'1': [], '2': [], 'common': []},
            'update': {'1': [], '2': [], 'common': []}
        }
    
        # Process common lines
        for line in common_lines:
            if line.startswith('ADD;'):
                operations['add']['common'].append(line)
            elif line.startswith('DEL;'):
                operations['del']['common'].append(line)
            elif ';' in line and line.startswith('WCC-VERI-DOC-'):
                operations['update']['common'].append(line)
    
        # Process unique lines in set 1
        for line in unique_to_1:
            if line.startswith('ADD;'):
                operations['add']['1'].append(line)
            elif line.startswith('DEL;'):
                operations['del']['1'].append(line)
            elif ';' in line and line.startswith('WCC-VERI-DOC-'):
                operations['update']['1'].append(line)
    
        # Process unique lines in set 2
        for line in unique_to_2:
            if line.startswith('ADD;'):
                operations['add']['2'].append(line)
            elif line.startswith('DEL;'):
                operations['del']['2'].append(line)
            elif ';' in line and line.startswith('WCC-VERI-DOC-'):
                operations['update']['2'].append(line)
    
        # Look for VERIDOC conflicts (same VERIDOC but different operations)
        veridoc_conflicts = []
    
        # Extract all VeriDocs from operations
        veridocs = {}
    
        # Add unique VeriDocs from file 1 updates
        for line in operations['update']['1']:
            parts = line.split(';')
            veridoc = parts[0]
            if veridoc not in veridocs:
                veridocs[veridoc] = {'1': [], '2': []}
            veridocs[veridoc]['1'].append(line)
    
        # Add unique VeriDocs from file 2 updates
        for line in operations['update']['2']:
            parts = line.split(';')
            veridoc = parts[0]
            if veridoc not in veridocs:
                veridocs[veridoc] = {'1': [], '2': []}
            veridocs[veridoc]['2'].append(line)
    
        # Check for DEL operations on the same VeriDoc
        for line in operations['del']['1'] + operations['del']['2']:
            parts = line.split(';')
            if len(parts) > 1:
                veridoc = parts[1]
                if veridoc in veridocs:
                    # Conflict: trying to update and delete the same VeriDoc
                    veridoc_conflicts.append({
                        'veridoc': veridoc,
                        'conflict_type': 'update_vs_delete',
                        'lines_1': veridocs[veridoc]['1'],
                        'lines_2': veridocs[veridoc]['2'],
                        'del_line': line
                    })
    
        # Check for conflicts in VeriDoc updates (same VeriDoc, different details)
        for veridoc, data in veridocs.items():
            if data['1'] and data['2']:
                # Same VeriDoc is being updated in both files
                # Compare the updates to see if they're different
                for line1 in data['1']:
                    for line2 in data['2']:
                        if line1 != line2:
                            veridoc_conflicts.append({
                                'veridoc': veridoc,
                                'conflict_type': 'different_updates',
                                'line_1': line1,
                                'line_2': line2
                            })
    
        return {
            'common_lines_count': len(common_lines),
            'unique_to_1_count': len(unique_to_1),
            'unique_to_2_count': len(unique_to_2),
            'operations': operations,
            'veridoc_conflicts': veridoc_conflicts,
            # If merged, include all lines from both files (without duplicates)
            'merged_lines': sorted(list(set1.union(set2)))
        }



    def create_updated_subsets(self, result_queue, main_file_path):
        """
        Worker function to create updated subsets from the newly combined main file.
        This is essentially a modified version of create_subsets().
        """
        try:
            import shutil
            from copy import copy
            import os
        
            # Prompt the user for the PMR number (or use a previously selected one)
            main_thread_queue = queue.Queue()
        
            # Ask for PMR number in the main thread
            self.root.after(0, lambda: self.prompt_for_pmr_number(main_thread_queue))
        
            # Wait for response
            pmr_number = main_thread_queue.get()
            if pmr_number is None:
                result_queue.put(('error', "Operation cancelled. No PMR number provided."))
                return
            
            pmr_folder = os.path.join(self.selected_base_path, f"PMR {pmr_number}")
        
            # Create the PMR folder if it doesn't exist
            if not os.path.exists(pmr_folder):
                os.makedirs(pmr_folder)
            
            # Check if template_file_path is defined
            if not hasattr(self, 'template_file_path') or not self.template_file_path:
                # Ask for template in the main thread
                self.root.after(0, lambda: self.prompt_for_template(main_thread_queue))
            
                template_file_path = main_thread_queue.get()
                if not template_file_path:
                    result_queue.put(('error', "Operation cancelled. No template file selected."))
                    return
                
                self.template_file_path = template_file_path
        
            result_queue.put(('status', "Creating updated subset files..."))
        
            # Get the swbs_groups
            swbs_groups = {
                'SWBS 000': [
                    '040-001', '040-002', '040-003', '040-004', '040-005',
                    '041-001', '041-002', '041-003', '041-004', '041-005', '041-006', 
                    '042-001', '042-003', '042-004', '042-005', '042-006', '042-007', '042-008', '042-009', '042-010',
                    '045-001', '045-002', '045-003', 
                    '068-001', '068-002', '068-003',
                    '070-001', '070-002', '070-003', '070-004', '070-005', '070-006',
                    '073-001', '073-002', '073-003', '073-004', '073-005', '073-006', '073-007', '073-008', '073-009', '073-010',
                    '074-001', 
                    '076-001', '076-002', '076-003', '076-004',
                    '077-001', '077-002', '077-003',
                    '079-001', '079-002', '079-003',
                    '080-001', '080-002', '080-003', 
                    '081-001',
                    '083-001', '083-002', '083-003', '083-004', '083-005', '083-006',
                    '084-001', '084-002', '084-003',
                    '085-001', '085-002', '085-003', '085-004', '085-005', '085-006',
                    '086-001', '086-002', '086-003', '086-004', '086-005', '086-006', '086-007',
                    '088-001', '088-002', '088-003', '088-004', '088-005', '088-006', '088-007', '088-008',
                    '092-001', '092-002', '092-003', '092-004',
                    '094-001', '094-002',
                    '096-001', '096-002', '096-003', '096-004',
                    '097-001', '097-002',
                    '098-001'
                ],
                'SWBS 100': [
                    '100-001', '100-002', '100-003', '100-004', '100-005','100-006','100-007','100-008','100-009','100-010','100-011','100-012', '100-013',
                    '167-001'
                ],
                'SWBS 200': [
                    '200-001', '200-002', '200-003',
                    '233-001', '245-001', '245-002', '245-003',
                    '249-001', '249-002', '249-003', '249-004',
                    '252-001',
                    '259-001', '259-002'
                ],
                'SWBS 202': [
                    '202-001','202-002','202-003','202-004','202-005','202-006','202-007','202-008','202-009','202-010','202-011','202-012','202-013','202-014','202-015','202-016'
                ],
                'SWBS 300': [
                    '300-001', '300-002', '300-003', '300-006', '300-007','300-008', '300-009', '300-010', '300-011', 
                    '302-001',
                    '303-001',
                    '310-001',
                    '313-001',
                    '314-001', '314-002',
                    '320-001', '320-002', '320-003', '320-004',
                    '324-001', '324-002',
                    '330-001', '330-002'
                ],
                'SWBS 400': [
                    '400-001', '400-002', '400-003', '400-004', '400-005', '400-006', '400-007', '400-008', '400-009', '400-010', '400-011','400-012',
                    '402-001', '402-002',
                    '405-001',
                    '407-001',
                    '420-001',
                    '428-001',
                    '432-001', '432-002'
                    '435-001', 
                    '440-001' 
                ],
                'SWBS 500': [
                    '504-001',
                    '506-001', '506-002',
                    '507-001',
                    '508-001',
                    '512-001', '512-002', '512-003',
                    '521-001', '521-002', '521-002', '521-003', '521-004',
                    '528-001', '528-002',
                    '529-001', '529-002', '529-003', '529-004',
                    '532-001', '532-002', '532-003', '532-004',
                    '541-001', '541-002',
                    '549-001', 
                    '551-001', '551-002',
                    '555-001', '555-002', '555-003', '555-004',
                    '580-001', '580-002', '580-003', '580-004', 
                    '581-001', '581-002', '581-003', '581-004', 
                    '582-001', '582-001',
                    '583-001', '583-002', '583-003', '583-004', 
                    '589-001', '589-002',
                    '593-001', '593-002', '593-003','593-004', '593-005'
                ],
                'SWBS 600': [
                    '602-001', 
                    '604-001',
                    '621-001',
                    '625-001',
                    '630-001', 
                    '631-001',
                    '633-001',
                    '634-001',
                    '635-001',
                    '640-001', '640-002'
                ]
            }
        
            # Create a set of all DI Numbers for quick lookup
            all_di_numbers = set()
            di_number_to_swbs = {}
            for swbs, di_numbers in swbs_groups.items():
                for di_number in di_numbers:
                    all_di_numbers.add(di_number)
                    di_number_to_swbs[di_number] = swbs
        
            # Copy the template file to create subset files
            di_number_to_file_path = {}
            for swbs, di_numbers in swbs_groups.items():
                swbs_number = swbs.replace('SWBS ', '')
                swbs_folder = os.path.join(pmr_folder, f"{swbs_number} SWBS")
                if not os.path.exists(swbs_folder):
                    os.makedirs(swbs_folder)
            
                for di_number in di_numbers:
                    filename = f"{di_number} - Revision X - RTVM Subset.xlsx"
                    dest_file_path = os.path.join(swbs_folder, filename)
                
                    try:
                        shutil.copyfile(self.template_file_path, dest_file_path)
                        di_number_to_file_path[di_number] = dest_file_path
                    except Exception as e:
                        result_queue.put(('detail', f"Error copying template for {di_number}: {str(e)}"))
        
            # Now, process the main Excel file
            try:
                wb_main = load_workbook(filename=main_file_path)
                if 'RTVM' not in wb_main.sheetnames:
                    result_queue.put(('error', "The sheet 'RTVM' was not found in the main Excel file."))
                    return
                
                ws_main = wb_main['RTVM']
            
                total_rows = ws_main.max_row - 1
                result_queue.put(('detail', f"Processing {total_rows} rows from main file..."))
            
                di_number_to_wb = {}
                di_number_to_ws = {}
                di_number_to_next_row = {}
            
                for idx, row in enumerate(ws_main.iter_rows(min_row=2, values_only=False), start=1):
                    if idx % 50 == 0 or idx == total_rows:
                        result_queue.put(('detail', f"Processed row {idx}/{total_rows}"))
                        result_queue.put(('progress', (idx, total_rows)))
                
                    cell_f = row[5]  # Column F - Assigned Verification Documents
                    cell_value = cell_f.value
                
                    if cell_value:
                        entries = cell_value.split('______________________')
                        di_numbers_in_cell = set()
                        for entry in entries:
                            entry = entry.strip()
                            if not entry:
                                continue
                            lines = entry.split('\n')
                            for line in lines:
                                line = line.strip()
                                if line.startswith('DI Number:'):
                                    di_number = line[len('DI Number:'):].strip()
                                    if di_number:
                                        di_numbers_in_cell.add(di_number)
                    
                        for di_number in di_numbers_in_cell:
                            if di_number in all_di_numbers:
                                if di_number not in di_number_to_wb:
                                    subset_file_path = di_number_to_file_path[di_number]
                                    wb_subset = load_workbook(filename=subset_file_path)
                                    if 'RTVM' in wb_subset.sheetnames:
                                        ws_subset = wb_subset['RTVM']
                                    else:
                                        ws_subset = wb_subset.active
                                    di_number_to_wb[di_number] = wb_subset
                                    di_number_to_ws[di_number] = ws_subset
                                    di_number_to_next_row[di_number] = 2
                                else:
                                    wb_subset = di_number_to_wb[di_number]
                                    ws_subset = di_number_to_ws[di_number]
                            
                                next_row = di_number_to_next_row[di_number]
                            
                                for cell in row:
                                    new_cell = ws_subset.cell(row=next_row, column=cell.column, value=cell.value)
                                    if cell.has_style:
                                        new_cell.font = copy(cell.font)
                                        new_cell.border = copy(cell.border)
                                        new_cell.fill = copy(cell.fill)
                                        new_cell.number_format = copy(cell.number_format)
                                        new_cell.protection = copy(cell.protection)
                                        new_cell.alignment = copy(cell.alignment)
                                    if cell.hyperlink:
                                        new_cell.hyperlink = copy(cell.hyperlink)
                                    if cell.comment:
                                        new_cell.comment = copy(cell.comment)
                            
                                di_number_to_next_row[di_number] += 1
            
                # Save all subset workbooks
                result_queue.put(('status', "Saving updated subset files..."))
            
                saved_count = 0
                for di_number, wb_subset in di_number_to_wb.items():
                    subset_file_path = di_number_to_file_path[di_number]
                    try:
                        wb_subset.save(subset_file_path)
                        wb_subset.close()
                        saved_count += 1
                        if saved_count % 10 == 0:
                            result_queue.put(('detail', f"Saved {saved_count} subset files..."))
                    except Exception as e:
                        result_queue.put(('detail', f"Failed to save subset file for DI Number {di_number}: {e}"))
        
            except Exception as e:
                result_queue.put(('error', f"Error processing main file: {str(e)}"))
                return
            
            # Signal completion
            result_queue.put(('all_complete', None))
        
        except Exception as e:
            result_queue.put(('error', f"Error creating updated subsets: {str(e)}"))


    def prompt_for_pmr_number(self, response_queue):
        """Prompt for PMR number in the main thread and put result in the queue."""
        pmr_number = simpledialog.askinteger("PMR Number", "Enter the PMR number:")
        response_queue.put(pmr_number)  # May be None if the user cancels

    def prompt_for_template(self, response_queue):
        """Prompt for template file in the main thread and put result in the queue."""
        template_file_path = filedialog.askopenfilename(
            title="Select Template Excel File",
            filetypes=[("Excel files", "*.xls;*.xlsx")]
        )
        response_queue.put(template_file_path)  # May be empty string if the user cancels

    def select_base_location(self):
            # Allow the user to select a folder where PMR-related files and subsets will be saved
            selected_directory = filedialog.askdirectory(title="Select Base Directory for PMR Files")
            if selected_directory:
                self.selected_base_path = selected_directory
                messagebox.showinfo("Base Location Selected", f"Base location set to: {self.selected_base_path}")
            else:
                messagebox.showwarning("No Selection", "No base location selected. Using current directory if needed.")

    def set_assigned_verification_cells(self):
        # Check if self.df is loaded
        if self.df is None:
            messagebox.showerror("Error", "No main file loaded. Please upload a main file before using this tool.")
            return

        assigned_column = "Assigned Verification Documents"
        if assigned_column in self.df.columns:
            self.assigned_verification_cells = self.df[assigned_column].tolist()
        else:
            messagebox.showerror("Error", f"Column '{assigned_column}' not found in the loaded DataFrame.")
            self.assigned_verification_cells = []


    # Below are the methods from the provided code snippet, adapted to use self.df, self.assigned_verification_cells,
    # self.excel_file_path, and the currently loaded main file. The upload functionality and separate class have been removed.

    def create_summary_report(self):
        # Collect data from all Assigned Verification Documents
        data_list = []
        for cell_content in self.assigned_verification_cells:
            # Convert cell_content to string to avoid float error
            if pd.isna(cell_content):
                continue
            cell_str = str(cell_content)
            if not cell_str or cell_str.strip().lower() == "nan":
                continue
            # Split entries separated by lines of underscores
            entries = cell_str.split('______________________')
            for entry in entries:
                entry = entry.strip()
                if not entry:
                    continue
                # Initialize variables
                data = {
                    'Object Identifier': "",
                    'DI Number': "",
                    'Government Assessed Status': "",
                    'Contractor Assessed Status': ""
                }
                # Split entry into lines
                lines = entry.split('\n')
                for line in lines:
                    line = line.strip()
                    if ':' in line:
                        key, value = line.split(':', 1)
                        key = key.strip()
                        value = value.strip()
                        if key in data:
                            data[key] = value
                data_list.append(data)

        # Now, data_list contains all the entries
        if not data_list:
            messagebox.showinfo("No Data", "No Assigned Verification Documents data found.")
            return

        # Convert to a DataFrame
        df_data = pd.DataFrame(data_list)

        # Now, group DI Numbers into SWBS groups as provided
        swbs_groups = {
            'SWBS 000': [
                '040-001', '040-002', '040-003', '040-004', '040-005',
                '041-001', '041-002', '041-003', '041-004', '041-005', '041-006', 
                '042-001', '042-003', '042-004', '042-005', '042-006', '042-007', '042-008', '042-009', '042-010',
                '045-001', '045-002', '045-003', 
                '068-001', '068-002', '068-003',
                '070-001', '070-002', '070-003', '070-004', '070-005', '070-006',
                '073-001', '073-002', '073-003', '073-004', '073-005', '073-006', '073-007', '073-008', '073-009', '073-010',
                '074-001', 
                '076-001', '076-002', '076-003', '076-004',
                '077-001', '077-002', '077-003',
                '079-001', '079-002', '079-003',
                '080-001', '080-002', '080-003', 
                '081-001',
                '083-001', '083-002', '083-003', '083-004', '083-005', '083-006',
                '084-001', '084-002', '084-003',
                '085-001', '085-002', '085-003', '085-004', '085-005', '085-006',
                '086-001', '086-002', '086-003', '086-004', '086-005', '086-006', '086-007',
                '088-001', '088-002', '088-003', '088-004', '088-005', '088-006', '088-007', '088-008',
                '092-001', '092-002', '092-003', '092-004',
                '094-001', '094-002',
                '096-001', '096-002', '096-003', '096-004',
                '097-001', '097-002',
                '098-001'
            ],
            'SWBS 100': [
                '100-001', '100-002', '100-003', '100-004', '100-005','100-006','100-007','100-008','100-009','100-010','100-011','100-012', '100-013',
                '167-001'
            ],
            'SWBS 200': [
                '200-001', '200-002', '200-003',
                '233-001', '245-001', '245-002', '245-003',
                '249-001', '249-002', '249-003', '249-004',
                '252-001',
                '259-001', '259-002'
            ],
            'SWBS 202': [
                '202-001','202-002','202-003','202-004','202-005','202-006','202-007','202-008','202-009','202-010','202-011','202-012','202-013','202-014','202-015','202-016'
            ],
            'SWBS 300': [
                '300-001', '300-002', '300-003', '300-006', '300-007','300-008', '300-009', '300-010', '300-011', 
                '302-001',
                '303-001',
                '310-001',
                '313-001',
                '314-001', '314-002',
                '320-001', '320-002', '320-003', '320-004',
                '324-001', '324-002',
                '330-001', '330-002'
            ],
            'SWBS 400': [
                '400-001', '400-002', '400-003', '400-004', '400-005', '400-006', '400-007', '400-008', '400-009', '400-010', '400-011','400-012',
                '402-001', '402-002',
                '405-001',
                '407-001',
                '420-001',
                '428-001',
                '432-001', '432-002'
                '435-001', 
                '440-001' 
            ],
            'SWBS 500': [
                '504-001',
                '506-001', '506-002',
                '507-001',
                '508-001',
                '512-001', '512-002', '512-003',
                '521-001', '521-002', '521-002', '521-003', '521-004',
                '528-001', '528-002',
                '529-001', '529-002', '529-003', '529-004',
                '532-001', '532-002', '532-003', '532-004',
                '541-001', '541-002',
                '549-001', 
                '551-001', '551-002',
                '555-001', '555-002', '555-003', '555-004',
                '580-001', '580-002', '580-003', '580-004', 
                '581-001', '581-002', '581-003', '581-004', 
                '582-001', '582-001',
                '583-001', '583-002', '583-003', '583-004', 
                '589-001', '589-002',
                '593-001', '593-002', '593-003','593-004', '593-005'
            ],
            'SWBS 600': [
                '602-001', 
                '604-001',
                '621-001',
                '625-001',
                '630-001', 
                '631-001',
                '633-001',
                '634-001',
                '635-001',
                '640-001', '640-002'
            ]
        }

        # Add DI Numbers that start with specific numbers
        for swbs in swbs_groups:
            if swbs == 'SWBS 000':
                starts_with = '0'
            elif swbs == 'SWBS 100':
                starts_with = '1'
            elif swbs == 'SWBS 200':
                starts_with = '2'
            elif swbs == 'SWBS 300':
                starts_with = '3'
            elif swbs == 'SWBS 400':
                starts_with = '4'
            elif swbs == 'SWBS 500':
                starts_with = '5'
            elif swbs == 'SWBS 600':
                starts_with = '6'
            else:
                starts_with = None

            if starts_with:
                # Get DI Numbers that start with the specified number
                di_numbers = df_data['DI Number'].unique()
                additional_di_numbers = [di for di in di_numbers if di.startswith(starts_with)]
                # Add them to the group if not already present
                for di in additional_di_numbers:
                    if di not in swbs_groups[swbs]:
                        swbs_groups[swbs].append(di)

        # Now, create a new window to display the pie charts
        self.report_window = tk.Toplevel(self.root)
        self.report_window.title("Summary Report")

        # Create a notebook to organize the charts
        notebook = ttk.Notebook(self.report_window)
        notebook.pack(fill='both', expand=True)

        # Create tabs for overall data
        overall_tabs = ttk.Notebook(self.report_window)
        notebook.add(overall_tabs, text='Overall Summary')
        self.figures_data = []
        # Overall Government Assessed Status
        gov_status_counts = df_data['Government Assessed Status'].value_counts()

        # Define colors
        status_colors = {
            'Disagree': 'red',
            'Agree': 'green',
            'Pending Review': 'orange',
            'Awaiting Input': 'blue'
        }

        gov_colors = [status_colors.get(status, 'grey') for status in gov_status_counts.index]

        gov_frame = ttk.Frame(overall_tabs)
        overall_tabs.add(gov_frame, text='Government Assessed Status')
        fig1, ax1 = plt.subplots(figsize=(6, 6))
        wedges1, texts1, autotexts1 = ax1.pie(
            gov_status_counts, labels=gov_status_counts.index,
            autopct='%1.1f%%', startangle=90, colors=gov_colors
        )
        ax1.axis('equal')
        ax1.set_title('Overall Government Assessed Status')

        # Store data for exporting
        self.figures_data = []
        self.figures_data.append({
            'swbs': 'Overall Status',
            'chart_type': 'Government Assessed Status',
            'counts': gov_status_counts.values,
            'labels': gov_status_counts.index.tolist(),
            'colors': gov_colors,
            'table_data': gov_status_counts.reset_index().values.tolist(),
            'table_columns': ['Status', 'Count']
        })

        canvas1 = FigureCanvasTkAgg(fig1, master=gov_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill='both', expand=True)

        # Overall Contractor Assessed Status
        contractor_status_counts = df_data['Contractor Assessed Status'].value_counts()

        # Define colors for Contractor Assessed Status
        contractor_status_colors = {
            'Satisfactory': 'green',
            'Unsatisfactory': 'red',
            'TBD': 'grey'
        }

        contractor_colors = [contractor_status_colors.get(status, 'grey') for status in contractor_status_counts.index]

        contractor_frame = ttk.Frame(overall_tabs)
        overall_tabs.add(contractor_frame, text='Contractor Assessed Status')
        fig2, ax2 = plt.subplots(figsize=(6, 6))
        wedges2, texts2, autotexts2 = ax2.pie(
            contractor_status_counts, labels=contractor_status_counts.index,
            autopct='%1.1f%%', startangle=90, colors=contractor_colors
        )
        ax2.axis('equal')
        ax2.set_title('Overall Contractor Assessed Status')

        self.figures_data.append({
            'swbs': 'Overall Status',
            'chart_type': 'Contractor Assessed Status',
            'counts': contractor_status_counts.values,
            'labels': contractor_status_counts.index.tolist(),
            'colors': contractor_colors,
            'table_data': contractor_status_counts.reset_index().values.tolist(),
            'table_columns': ['Status', 'Count']
        })

        canvas2 = FigureCanvasTkAgg(fig2, master=contractor_frame)
        canvas2.draw()
        canvas2.get_tk_widget().pack(fill='both', expand=True)

        # Overall DI Number Distribution
        di_number_counts = df_data['DI Number'].value_counts()
        di_frame = ttk.Frame(overall_tabs)
        overall_tabs.add(di_frame, text='DI Number Distribution')
        fig3, ax3 = plt.subplots(figsize=(6, 6))
        ax3.pie(
            di_number_counts, labels=di_number_counts.index,
            autopct='%1.1f%%', startangle=90
        )
        ax3.axis('equal')
        ax3.set_title('Overall DI Number Distribution')

        self.figures_data.append({
            'swbs': 'Overall Status',
            'chart_type': 'DI Number Distribution',
            'counts': di_number_counts.values,
            'labels': di_number_counts.index.tolist(),
            'colors': None,  # No specific colors
            'table_data': di_number_counts.reset_index().values.tolist(),
            'table_columns': ['DI Number', 'Count']
        })

        canvas3 = FigureCanvasTkAgg(fig3, master=di_frame)
        canvas3.draw()
        canvas3.get_tk_widget().pack(fill='both', expand=True)

        # For each SWBS group, create a tab
        for swbs, di_numbers in swbs_groups.items():
             # Strip spaces from swbs to avoid mismatch
            swbs = swbs.strip()

            # Filter the data for this SWBS group
            df_swbs = df_data[df_data['DI Number'].isin(di_numbers)]
            if df_swbs.empty:
                continue

            swbs_notebook = ttk.Notebook(self.report_window)
            notebook.add(swbs_notebook, text=swbs)

            # Government Assessed Status for SWBS
            gov_status_counts_swbs = df_swbs['Government Assessed Status'].value_counts()
            gov_colors_swbs = [status_colors.get(status, 'grey') for status in gov_status_counts_swbs.index]

            gov_frame_swbs = ttk.Frame(swbs_notebook)
            swbs_notebook.add(gov_frame_swbs, text='Government Assessed Status')
            fig_swbs_gov, ax_swbs_gov = plt.subplots(figsize=(6, 6))
            wedges_swbs_gov, texts_swbs_gov, autotexts_swbs_gov = ax_swbs_gov.pie(
                gov_status_counts_swbs, labels=gov_status_counts_swbs.index,
                autopct='%1.1f%%', startangle=90, colors=gov_colors_swbs
            )
            ax_swbs_gov.axis('equal')
            ax_swbs_gov.set_title(f'{swbs} - Government Assessed Status')

            self.figures_data.append({
                'swbs': swbs.replace('SWBS ', ''),
                'chart_type': 'Government Assessed Status',
                'counts': gov_status_counts_swbs.values,
                'labels': gov_status_counts_swbs.index.tolist(),
                'colors': gov_colors_swbs,
                'table_data': gov_status_counts_swbs.reset_index().values.tolist(),
                'table_columns': ['Status', 'Count']
            })

            canvas_swbs_gov = FigureCanvasTkAgg(fig_swbs_gov, master=gov_frame_swbs)
            canvas_swbs_gov.draw()
            canvas_swbs_gov.get_tk_widget().pack(fill='both', expand=True)

            # Contractor Assessed Status for SWBS
            contractor_status_counts_swbs = df_swbs['Contractor Assessed Status'].value_counts()
            contractor_colors_swbs = [contractor_status_colors.get(status, 'grey') for status in contractor_status_counts_swbs.index]

            contractor_frame_swbs = ttk.Frame(swbs_notebook)
            swbs_notebook.add(contractor_frame_swbs, text='Contractor Assessed Status')
            fig_swbs_contractor, ax_swbs_contractor = plt.subplots(figsize=(6, 6))
            wedges_swbs_contractor, texts_swbs_contractor, autotexts_swbs_contractor = ax_swbs_contractor.pie(
                contractor_status_counts_swbs, labels=contractor_status_counts_swbs.index,
                autopct='%1.1f%%', startangle=90, colors=contractor_colors_swbs
            )
            ax_swbs_contractor.axis('equal')
            ax_swbs_contractor.set_title(f'{swbs} - Contractor Assessed Status')

            self.figures_data.append({
                'swbs': swbs.replace('SWBS ', ''),
                'chart_type': 'Contractor Assessed Status',
                'counts': contractor_status_counts_swbs.values,
                'labels': contractor_status_counts_swbs.index.tolist(),
                'colors': contractor_colors_swbs,
                'table_data': contractor_status_counts_swbs.reset_index().values.tolist(),
                'table_columns': ['Status', 'Count']
            })

            canvas_swbs_contractor = FigureCanvasTkAgg(fig_swbs_contractor, master=contractor_frame_swbs)
            canvas_swbs_contractor.draw()
            canvas_swbs_contractor.get_tk_widget().pack(fill='both', expand=True)

            # DI Number Distribution for SWBS
            di_number_counts_swbs = df_swbs['DI Number'].value_counts()
            di_frame_swbs = ttk.Frame(swbs_notebook)
            swbs_notebook.add(di_frame_swbs, text='DI Number Distribution')
            fig_swbs_di, ax_swbs_di = plt.subplots(figsize=(6, 6))
            ax_swbs_di.pie(
                di_number_counts_swbs, labels=di_number_counts_swbs.index,
                autopct='%1.1f%%', startangle=90
            )
            ax_swbs_di.axis('equal')
            ax_swbs_di.set_title(f'{swbs} - DI Number Distribution')

            self.figures_data.append({
                'swbs': swbs.replace('SWBS ', ''),
                'chart_type': 'DI Number Distribution',
                'counts': di_number_counts_swbs.values,
                'labels': di_number_counts_swbs.index.tolist(),
                'colors': None,  # No specific colors
                'table_data': di_number_counts_swbs.reset_index().values.tolist(),
                'table_columns': ['DI Number', 'Count']
            })

            canvas_swbs_di = FigureCanvasTkAgg(fig_swbs_di, master=di_frame_swbs)
            canvas_swbs_di.draw()
            canvas_swbs_di.get_tk_widget().pack(fill='both', expand=True)
            
    def export_photos_of_report(self):
        if not hasattr(self, 'figures_data') or not self.figures_data:
            messagebox.showwarning("No Report Generated", "Please create the summary report first.")
            return

        if not self.selected_base_path:
            messagebox.showerror("No Base Location", "Please select a base location before exporting photos.")
            return

        # Prompt for PMR number
        pmr_number = simpledialog.askinteger("PMR Number", "Enter the PMR number:")
        if pmr_number is None:
            return  # User canceled

        pmr_folder = os.path.join(self.selected_base_path, f"PMR {pmr_number}")

        # Create the PMR folder if it doesn't exist
        if not os.path.exists(pmr_folder):
            os.makedirs(pmr_folder)

        for data in self.figures_data:
            swbs = data['swbs']
            chart_type = data['chart_type']
            counts = data['counts']
            labels = data['labels']
            colors = data['colors']
            table_data = data['table_data']
            table_columns = data['table_columns']

            # Strip spaces from swbs to ensure exact match
            swbs = swbs.strip()

            if swbs == 'Overall Status':
                swbs_folder = os.path.join(pmr_folder, "Overall Status")
            else:
                swbs_folder = os.path.join(pmr_folder, f"{swbs} SWBS", "Status Photos")

            if not os.path.exists(swbs_folder):
                os.makedirs(swbs_folder)

            # Recreate the figure
            fig, ax = plt.subplots(figsize=(6, 6))
            if colors:
                ax.pie(
                    counts, labels=labels,
                    autopct='%1.1f%%', startangle=90, colors=colors
                )
            else:
                ax.pie(
                    counts, labels=labels,
                    autopct='%1.1f%%', startangle=90
                )
            ax.axis('equal')
            ax.set_title(f"{swbs} - {chart_type}")

            # Save the figure
            filename = f"PMR {pmr_number} - SWBS {swbs} - {chart_type}.png"
            save_path = os.path.join(swbs_folder, filename)
        
            # Ensure directories exist before saving
            if not os.path.exists(swbs_folder):
                os.makedirs(swbs_folder)

            fig.savefig(save_path, bbox_inches='tight')
            plt.close(fig)

            # Prepare data for Excel export
            if swbs == 'Overall Status':
                excel_filename = f"PMR {pmr_number} - Overall Status - Raw Data Export.xlsx"
                excel_save_path = os.path.join(swbs_folder, excel_filename)
            else:
                excel_filename = f"PMR {pmr_number} - SWBS {swbs} - Raw Data Export.xlsx"
                excel_save_path = os.path.join(swbs_folder, excel_filename)

            # Write the data to Excel
            table_df = pd.DataFrame(table_data, columns=table_columns)
            sheet_name = chart_type
            if not os.path.exists(excel_save_path):
                with pd.ExcelWriter(excel_save_path, engine='openpyxl') as writer:
                    table_df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                with pd.ExcelWriter(excel_save_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    table_df.to_excel(writer, sheet_name=sheet_name, index=False)

        messagebox.showinfo("Export Completed", f"All charts have been exported to {pmr_folder}")

    def create_subsets(self):
        import shutil
        from openpyxl import load_workbook
        from copy import copy
        from tkinter import ttk, messagebox, simpledialog, filedialog
        import threading
        import queue
        import os

        if self.selected_base_path is None:
            messagebox.showerror("No Base Location", "Please select a base location before creating subsets.")
            return

        if not self.excel_file_path:
            messagebox.showwarning("No File", "Please upload an Excel file first.")
            return

        # Prompt the user for the PMR number
        pmr_number = simpledialog.askinteger("PMR Number", "Enter the PMR number:")
        if pmr_number is None:
            return  # User canceled

        pmr_folder = os.path.join(self.selected_base_path, f"PMR {pmr_number}")

        # Create the PMR folder if it doesn't exist
        if not os.path.exists(pmr_folder):
            os.makedirs(pmr_folder)

        # Check if template_file_path is defined
        if not hasattr(self, 'template_file_path') or not self.template_file_path:
            # Ask the user if they want to upload a template file
            if messagebox.askyesno("No Template File", "No template file is currently selected.\nWould you like to select one now?"):
                template_file_path = filedialog.askopenfilename(
                    title="Select Template Excel File",
                    filetypes=[("Excel files", "*.xls;*.xlsx")]
                )
                if not template_file_path:
                    messagebox.showwarning("No Template File", "No template Excel file was selected. Cannot proceed.")
                    return
                self.template_file_path = template_file_path
            else:
                # User chose not to select a template file
                return

        template_file_path = self.template_file_path

        # Define the SWBS groups and DI Numbers (unchanged)
        swbs_groups = {
                    'SWBS 000': [
                        '040-001', '040-002', '040-003', '040-004', '040-005',
                        '041-001', '041-002', '041-003', '041-004', '041-005', '041-006', 
                        '042-001', '042-003', '042-004', '042-005', '042-006', '042-007', '042-008', '042-009', '042-010',
                        '045-001', '045-002', '045-003', 
                        '068-001', '068-002', '068-003',
                        '070-001', '070-002', '070-003', '070-004', '070-005', '070-006',
                        '073-001', '073-002', '073-003', '073-004', '073-005', '073-006', '073-007', '073-008', '073-009', '073-010',
                        '074-001', 
                        '076-001', '076-002', '076-003', '076-004',
                        '077-001', '077-002', '077-003',
                        '079-001', '079-002', '079-003',
                        '080-001', '080-002', '080-003', 
                        '081-001',
                        '083-001', '083-002', '083-003', '083-004', '083-005', '083-006',
                        '084-001', '084-002', '084-003',
                        '085-001', '085-002', '085-003', '085-004', '085-005', '085-006',
                        '086-001', '086-002', '086-003', '086-004', '086-005', '086-006', '086-007',
                        '088-001', '088-002', '088-003', '088-004', '088-005', '088-006', '088-007', '088-008',
                        '092-001', '092-002', '092-003', '092-004',
                        '094-001', '094-002',
                        '096-001', '096-002', '096-003', '096-004',
                        '097-001', '097-002',
                        '098-001'
                    ],
                    'SWBS 100': [
                        '100-001', '100-002', '100-003', '100-004', '100-005','100-006','100-007','100-008','100-009','100-010','100-011','100-012', '100-013',
                        '167-001'
                    ],
                    'SWBS 200': [
                        '200-001', '200-002', '200-003',
                        '233-001', '245-001', '245-002', '245-003',
                        '249-001', '249-002', '249-003', '249-004',
                        '252-001',
                        '259-001', '259-002'
                    ],
                    'SWBS 202': [
                        '202-001','202-002','202-003','202-004','202-005','202-006','202-007','202-008','202-009','202-010','202-011','202-012','202-013','202-014','202-015','202-016'
                    ],
                    'SWBS 300': [
                        '300-001', '300-002', '300-003', '300-006', '300-007','300-008', '300-009', '300-010', '300-011', 
                        '302-001',
                        '303-001',
                        '310-001',
                        '313-001',
                        '314-001', '314-002',
                        '320-001', '320-002', '320-003', '320-004',
                        '324-001', '324-002',
                        '330-001', '330-002'
                    ],
                    'SWBS 400': [
                        '400-001', '400-002', '400-003', '400-004', '400-005', '400-006', '400-007', '400-008', '400-009', '400-010', '400-011','400-012',
                        '402-001', '402-002',
                        '405-001',
                        '407-001',
                        '420-001',
                        '428-001',
                        '432-001', '432-002'
                        '435-001', 
                        '440-001' 
                    ],
                    'SWBS 500': [
                        '504-001',
                        '506-001', '506-002',
                        '507-001',
                        '508-001',
                        '512-001', '512-002', '512-003',
                        '521-001', '521-002', '521-002', '521-003', '521-004',
                        '528-001', '528-002',
                        '529-001', '529-002', '529-003', '529-004',
                        '532-001', '532-002', '532-003', '532-004',
                        '541-001', '541-002',
                        '549-001', 
                        '551-001', '551-002',
                        '555-001', '555-002', '555-003', '555-004',
                        '580-001', '580-002', '580-003', '580-004', 
                        '581-001', '581-002', '581-003', '581-004', 
                        '582-001', '582-001',
                        '583-001', '583-002', '583-003', '583-004', 
                        '589-001', '589-002',
                        '593-001', '593-002', '593-003','593-004', '593-005'
                    ],
                    'SWBS 600': [
                        '602-001', 
                        '604-001',
                        '621-001',
                        '625-001',
                        '630-001', 
                        '631-001',
                        '633-001',
                        '634-001',
                        '635-001',
                        '640-001', '640-002'
                    ]
                }

        def process_data(progress_queue):
            print("Starting recombination process...")
            # Create a set of all DI Numbers for quick lookup
            all_di_numbers = set()
            di_number_to_swbs = {}
            for swbs, di_numbers in swbs_groups.items():
                for di_number in di_numbers:
                    all_di_numbers.add(di_number)
                    di_number_to_swbs[di_number] = swbs

            # Copy the template file to create subset files
            di_number_to_file_path = {}
            for swbs, di_numbers in swbs_groups.items():
                swbs_number = swbs.replace('SWBS ', '')
                swbs_folder = os.path.join(pmr_folder, f"{swbs_number} SWBS")
                if not os.path.exists(swbs_folder):
                    os.makedirs(swbs_folder)

                for di_number in di_numbers:
                    filename = f"{di_number} - Revision X - RTVM Subset.xlsx"
                    dest_file_path = os.path.join(swbs_folder, filename)

                    try:
                        shutil.copyfile(template_file_path, dest_file_path)
                        di_number_to_file_path[di_number] = dest_file_path
                    except Exception as e:
                        progress_queue.put(('error', f"An error occurred while copying the template file:\n{e}"))
                        return

            # Now, process the main Excel file
            try:
                print("Starting recombination process..2.")
                wb_main = load_workbook(filename=self.excel_file_path)
            except Exception as e:
                progress_queue.put(('error', f"Failed to open the main Excel file:\n{e}"))
                return

            if 'RTVM' not in wb_main.sheetnames:
                progress_queue.put(('error', "The sheet 'RTVM' was not found in the Excel file."))
                return
            ws_main = wb_main['RTVM']

            total_rows = ws_main.max_row - 1
            progress_queue.put(('total', total_rows))

            di_number_to_wb = {}
            di_number_to_ws = {}
            di_number_to_next_row = {}

            for idx, row in enumerate(ws_main.iter_rows(min_row=2, values_only=False), start=1):
                progress_queue.put(('progress', idx))

                cell_f = row[5]
                cell_value = cell_f.value

                if cell_value:
                    entries = cell_value.split('______________________')
                    di_numbers_in_cell = set()
                    for entry in entries:
                        entry = entry.strip()
                        if not entry:
                            continue
                        lines = entry.split('\n')
                        for line in lines:
                            line = line.strip()
                            if line.startswith('DI Number:'):
                                di_number = line[len('DI Number:'):].strip()
                                if di_number:
                                    di_numbers_in_cell.add(di_number)

                    for di_number in di_numbers_in_cell:
                        if di_number in all_di_numbers:
                            if di_number not in di_number_to_wb:
                                subset_file_path = di_number_to_file_path[di_number]
                                wb_subset = load_workbook(filename=subset_file_path)
                                if 'RTVM' in wb_subset.sheetnames:
                                    ws_subset = wb_subset['RTVM']
                                else:
                                    ws_subset = wb_subset.active
                                di_number_to_wb[di_number] = wb_subset
                                di_number_to_ws[di_number] = ws_subset
                                di_number_to_next_row[di_number] = 2
                            else:
                                wb_subset = di_number_to_wb[di_number]
                                ws_subset = di_number_to_ws[di_number]

                            next_row = di_number_to_next_row[di_number]

                            for cell in row:
                                new_cell = ws_subset.cell(row=next_row, column=cell.column, value=cell.value)
                                if cell.has_style:
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.number_format = copy(cell.number_format)
                                    new_cell.protection = copy(cell.protection)
                                    new_cell.alignment = copy(cell.alignment)
                                if cell.hyperlink:
                                    new_cell.hyperlink = copy(cell.hyperlink)
                                if cell.comment:
                                    new_cell.comment = copy(cell.comment)

                            di_number_to_next_row[di_number] += 1

            for di_number, wb_subset in di_number_to_wb.items():
                subset_file_path = di_number_to_file_path[di_number]
                try:
                    wb_subset.save(subset_file_path)
                    wb_subset.close()
                except Exception as e:
                    progress_queue.put(('error', f"Failed to save subset file for DI Number {di_number}:\n{e}"))

            progress_queue.put(('done', None))

        progress_queue = queue.Queue()
        threading.Thread(target=process_data, args=(progress_queue,)).start()

        progress_window = tk.Toplevel()
        progress_window.title("Processing Rows")
        progress_label = tk.Label(progress_window, text="Processing rows...")
        progress_label.pack()
        progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=300, mode='determinate')
        progress_bar.pack()
        progress_info = tk.Label(progress_window, text="0%")
        progress_info.pack()

        total_rows = 0
        current_row = 0

        def update_progress():
            nonlocal total_rows, current_row
            try:
                while True:
                    message_type, data = progress_queue.get_nowait()
                    if message_type == 'total':
                        total_rows = data
                        progress_bar['maximum'] = total_rows
                    elif message_type == 'progress':
                        current_row = data
                        if total_rows > 0:
                            percent = int((current_row / total_rows) * 100)
                            progress_bar['value'] = current_row
                            progress_info.config(text=f"{percent}%")
                    elif message_type == 'error':
                        messagebox.showerror("Error", data)
                        progress_window.destroy()
                        return
                    elif message_type == 'done':
                        progress_bar['value'] = progress_bar['maximum']
                        progress_info.config(text="100%")
                        progress_window.destroy()
                        messagebox.showinfo("Subsets Created", f"The subsets have been created in {pmr_folder}")
                        return
            except queue.Empty:
                pass
            except Exception as e:
                print(f"Error in update_progress: {e}")
            progress_window.after(100, update_progress)

        update_progress()


    def recombine_subsets(self):
        import os
        import time
        from openpyxl import load_workbook
        from copy import copy
        from tkinter import ttk, messagebox, filedialog, simpledialog
        import tkinter as tk
        import threading
        import queue

        # Since you've already uploaded the main Excel file and stored it in self.excel_file_path,
        # we assume self.excel_file_path is valid and no need to check again.
        # If you still want to ensure it's loaded, you can handle it gracefully:
        if not self.excel_file_path:
            messagebox.showerror("Error", "No main Excel file is loaded. Please upload a main file first.")
            return

        # Prompt the user to select the template file if not already selected
        if not hasattr(self, 'template_file_path') or not self.template_file_path:
            template_file_path = filedialog.askopenfilename(
                title="Select the Template Excel File",
                filetypes=[("Excel files", "*.xls;*.xlsx")]
            )
            if not template_file_path:
                messagebox.showwarning("No Template File", "No template Excel file was selected.")
                return
            self.template_file_path = template_file_path
        else:
            template_file_path = self.template_file_path

        # Prompt the user to select the folder containing the subset files
        subset_folder = filedialog.askdirectory(title="Select the PMR Folder Containing Subset Files")
        if not subset_folder:
            messagebox.showwarning("No Folder Selected", "No folder was selected.")
            return

        # Prompt the user to select the output file path for the recombined Excel file
        output_file_path = filedialog.asksaveasfilename(
            title="Save Recombined Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not output_file_path:
            messagebox.showwarning("No Output File", "No output file was specified.")
            return

        def process_data(progress_queue):
            try:
                print("Starting recombination process...")

                # Load the main Excel file to get the structure and headers
                wb_main = load_workbook(filename=self.excel_file_path)
                if 'RTVM' in wb_main.sheetnames:
                    ws_main = wb_main['RTVM']
                else:
                    ws_main = wb_main.active
                print("Main workbook loaded.")

                # Load the template workbook to use as the recombined workbook
                wb_recombined = load_workbook(filename=template_file_path)
                if 'RTVM' in wb_recombined.sheetnames:
                    ws_recombined = wb_recombined['RTVM']
                else:
                    ws_recombined = wb_recombined.active
                print("Template workbook loaded as recombined workbook.")

                # Clear existing data in the recombined worksheet except the header
                ws_recombined.delete_rows(2, ws_recombined.max_row)
                print("Cleared existing data in the recombined worksheet.")

                # Read all subset files and collect data
                data_dict = {}  # Key: DOORS SPEC ID, Value: Row data (as a dict)
                columns_to_combine = [7, 8]  # Columns G and H (1-based indexing)

                print("Collecting subset files...")
                subset_files = []
                for root, dirs, files in os.walk(subset_folder):
                    if 'Status Photos' in dirs:
                        dirs.remove('Status Photos')
                    for file in files:
                        if file.endswith('.xlsx') or file.endswith('.xls'):
                            file_path = os.path.join(root, file)
                            subset_files.append(file_path)
                print(f"Total subset files found: {len(subset_files)}")

                total_files = len(subset_files)
                progress_queue.put(('total', total_files))

                current_file = 0

                for subset_file in subset_files:
                    current_file += 1
                    progress_queue.put(('progress', current_file))
                    print(f"Processing file {current_file}/{total_files}: {subset_file}")

                    wb_subset = load_workbook(filename=subset_file)
                    if 'RTVM' in wb_subset.sheetnames:
                        ws_subset = wb_subset['RTVM']
                    else:
                        ws_subset = wb_subset.active
                    print(f"Opened subset workbook: {subset_file}")

                    for row in ws_subset.iter_rows(min_row=2, values_only=False):
                        doors_spec_id_cell = row[0]
                        doors_spec_id = doors_spec_id_cell.value
                        if not doors_spec_id:
                            continue

                        row_values = {}
                        for cell in row:
                            col_idx = cell.column
                            row_values[col_idx] = cell.value

                        if doors_spec_id not in data_dict:
                            data_dict[doors_spec_id] = row_values
                        else:
                            # Combine columns G and H
                            for col_idx in columns_to_combine:
                                existing_value = data_dict[doors_spec_id].get(col_idx, '')
                                new_value = row_values.get(col_idx, '')
                                if existing_value and new_value and new_value not in existing_value:
                                    combined_value = f"{existing_value}\n{new_value}"
                                    data_dict[doors_spec_id][col_idx] = combined_value
                                elif not existing_value and new_value:
                                    data_dict[doors_spec_id][col_idx] = new_value

                    wb_subset.close()
                    print(f"Finished processing file: {subset_file}")

                print("All subset files processed.")

                print("Preparing to write data to recombined workbook...")
                progress_queue.put(('writing_start', None))

                doors_spec_id_order = []
                for row in ws_main.iter_rows(min_row=2, values_only=False):
                    doors_spec_id = row[0].value
                    if doors_spec_id:
                        doors_spec_id_order.append(doors_spec_id)

                total_rows = len(doors_spec_id_order)
                progress_queue.put(('writing_total', total_rows))
                print(f"Total rows to write: {total_rows}")

                next_row = 2

                for idx, doors_spec_id in enumerate(doors_spec_id_order, start=1):
                    progress_queue.put(('writing_progress', idx))
                    if idx % 100 == 0 or idx == total_rows:
                        print(f"Writing row {idx}/{total_rows}")
                    if doors_spec_id in data_dict:
                        row_values = data_dict[doors_spec_id]
                        for cell in ws_main[next_row]:
                            col_idx = cell.column
                            value = row_values.get(col_idx, cell.value)
                            new_cell = ws_recombined.cell(row=next_row, column=col_idx, value=value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                            if cell.hyperlink:
                                new_cell.hyperlink = copy(cell.hyperlink)
                            if cell.comment:
                                new_cell.comment = copy(cell.comment)
                    else:
                        for cell in ws_main[next_row]:
                            new_cell = ws_recombined.cell(row=next_row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                            if cell.hyperlink:
                                new_cell.hyperlink = copy(cell.hyperlink)
                            if cell.comment:
                                new_cell.comment = copy(cell.comment)
                    next_row += 1

                print("Data written to recombined workbook.")

                wb_recombined.save(output_file_path)
                wb_recombined.close()
                print(f"Recombined workbook saved at: {output_file_path}")

                progress_queue.put(('done', None))
                print("Recombination process completed.")

            except Exception as e:
                progress_queue.put(('error', f"An error occurred during recombination:\n{e}"))
                print(f"An error occurred during recombination: {e}")

        progress_queue = queue.Queue()
        threading.Thread(target=process_data, args=(progress_queue,)).start()

        progress_window = tk.Toplevel()
        progress_window.title("Recombining Subsets")
        progress_label = tk.Label(progress_window, text="Processing subset files...")
        progress_label.pack()
        progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=300, mode='determinate')
        progress_bar.pack()
        progress_info = tk.Label(progress_window, text="0%")
        progress_info.pack()

        total_files = 0
        current_file = 0
        total_rows = 0
        current_row = 0
        write_start_time = None

        def update_progress():
            nonlocal total_files, current_file, total_rows, current_row, write_start_time
            try:
                while True:
                    message_type, data = progress_queue.get_nowait()
                    if message_type == 'total':
                        total_files = data
                        progress_bar['maximum'] = total_files
                        progress_label.config(text="Processing subset files...")
                    elif message_type == 'progress':
                        current_file = data
                        if total_files > 0:
                            percent = int((current_file / total_files) * 100)
                            progress_bar['value'] = current_file
                            progress_info.config(text=f"{percent}%")
                    elif message_type == 'writing_start':
                        current_row = 0
                        total_rows = 0
                        progress_bar['value'] = 0
                        progress_bar['maximum'] = 1
                        progress_label.config(text="Writing data to recombined workbook...")
                        progress_info.config(text="0%")
                        write_start_time = time.time()
                    elif message_type == 'writing_total':
                        total_rows = data
                        progress_bar['maximum'] = total_rows
                    elif message_type == 'writing_progress':
                        current_row = data
                        if total_rows > 0:
                            percent = int((current_row / total_rows) * 100)
                            progress_bar['value'] = current_row
                            elapsed_time = time.time() - write_start_time
                            rows_per_sec = current_row / elapsed_time if elapsed_time > 0 else 0
                            remaining_rows = total_rows - current_row
                            est_remaining_time = remaining_rows / rows_per_sec if rows_per_sec > 0 else 0
                            mins, secs = divmod(int(est_remaining_time), 60)
                            hours, mins = divmod(mins, 60)
                            est_time_str = f"Est. time remaining: {hours:d}h {mins:02d}m {secs:02d}s"
                            progress_info.config(text=f"{percent}% - {est_time_str}")
                    elif message_type == 'error':
                        messagebox.showerror("Error", data)
                        progress_window.destroy()
                        return
                    elif message_type == 'done':
                        progress_bar['value'] = progress_bar['maximum']
                        progress_info.config(text="100%")
                        progress_window.destroy()
                        messagebox.showinfo(
                            "Recombination Complete",
                            f"The subset files have been recombined into {output_file_path}"
                        )
                        return
            except queue.Empty:
                pass
            except Exception as e:
                print(f"Error in update_progress: {e}")
            progress_window.after(100, update_progress)

        update_progress()


    def merge_single_subset(self):
        from openpyxl import load_workbook
        from copy import copy
        import tkinter as tk
        from tkinter import filedialog, messagebox
        import threading
        import queue

        # Prompt the user to select the subset file to merge
        subset_file_path = filedialog.askopenfilename(
            title="Select the Subset Excel File to Merge",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if not subset_file_path:
            messagebox.showwarning("No Subset File", "No subset Excel file was selected.")
            return

        # Prompt the user to select the main file to merge into
        main_file_path = filedialog.askopenfilename(
            title="Select the Main Excel File to Merge Into",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if not main_file_path:
            messagebox.showwarning("No Main File", "No main Excel file was selected.")
            return

        # Prompt the user to select the output file path for the merged Excel file
        output_file_path = filedialog.asksaveasfilename(
            title="Save Merged Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not output_file_path:
            messagebox.showwarning("No Output File", "No output file was specified.")
            return

        # Function to perform the merge in a background thread
        def process_data(progress_queue):
            try:
                # Load the main Excel file
                wb_main = load_workbook(filename=main_file_path)
                if 'RTVM' in wb_main.sheetnames:
                    ws_main = wb_main['RTVM']
                else:
                    ws_main = wb_main.active  # Use the first sheet
                print("Main workbook loaded.")

                # Load the subset Excel file
                wb_subset = load_workbook(filename=subset_file_path)
                if 'RTVM' in wb_subset.sheetnames:
                    ws_subset = wb_subset['RTVM']
                else:
                    ws_subset = wb_subset.active  # Use the first sheet
                print("Subset workbook loaded.")

                # Build a mapping from DOORS SPEC ID to row in main workbook
                doors_spec_id_to_row_main = {}
                for row in ws_main.iter_rows(min_row=2, values_only=False):
                    doors_spec_id = row[0].value  # Column A
                    if doors_spec_id:
                        doors_spec_id_to_row_main[doors_spec_id] = row

                # For progress tracking
                total_rows = ws_subset.max_row - 1  # Exclude header
                progress_queue.put(('total', total_rows))
                current_row_num = 0

                # Iterate over the subset rows
                for row_subset in ws_subset.iter_rows(min_row=2, values_only=False):
                    current_row_num += 1
                    progress_queue.put(('progress', current_row_num))

                    doors_spec_id = row_subset[0].value  # Column A
                    if not doors_spec_id:
                        continue  # Skip rows without DOORS SPEC ID

                    if doors_spec_id in doors_spec_id_to_row_main:
                        # Merge data into the main workbook
                        row_main = doors_spec_id_to_row_main[doors_spec_id]

                        # For columns G and H (columns 7 and 8), combine the data
                        for col_idx in [7, 8]:  # Columns G and H
                            cell_main = row_main[col_idx - 1]  # 0-based index
                            cell_subset = row_subset[col_idx - 1]

                            value_main = cell_main.value or ''
                            value_subset = cell_subset.value or ''
                            if value_subset and value_subset not in value_main:
                                if value_main:
                                    combined_value = f"{value_main}\n{value_subset}"
                                else:
                                    combined_value = value_subset
                                cell_main.value = combined_value
                    else:
                        # DOORS SPEC ID not found in main workbook
                        # You may choose to add the new row or skip it
                        pass

                # Save the merged workbook
                wb_main.save(output_file_path)
                wb_main.close()
                wb_subset.close()
                print("Merged workbook saved.")

                # Signal completion
                progress_queue.put(('done', None))

            except Exception as e:
                progress_queue.put(('error', f"An error occurred during merging:\n{e}"))
                print(f"An error occurred during merging: {e}")

        # Create a queue to communicate with the background thread
        progress_queue = queue.Queue()

        # Start the background thread
        threading.Thread(target=process_data, args=(progress_queue,)).start()

        # Create a progress bar
        progress_window = tk.Toplevel()
        progress_window.title("Merging Subset")
        progress_label = tk.Label(progress_window, text="Merging subset file...")
        progress_label.pack()
        progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=300, mode='determinate')
        progress_bar.pack()
        progress_info = tk.Label(progress_window, text="0%")
        progress_info.pack()

        # Variables to keep track of progress
        total_rows = 0
        current_row = 0

        # Function to update the progress bar
        def update_progress():
            nonlocal total_rows, current_row
            try:
                while True:
                    message_type, data = progress_queue.get_nowait()
                    if message_type == 'total':
                        total_rows = data
                        progress_bar['maximum'] = total_rows
                    elif message_type == 'progress':
                        current_row = data
                        if total_rows > 0:
                            percent = int((current_row / total_rows) * 100)
                            progress_bar['value'] = current_row
                            progress_info.config(text=f"{percent}%")
                    elif message_type == 'error':
                        messagebox.showerror("Error", data)
                        progress_window.destroy()
                        return
                    elif message_type == 'done':
                        progress_bar['value'] = total_rows
                        progress_info.config(text="100%")
                        progress_window.destroy()
                        messagebox.showinfo(
                            "Merge Complete",
                            f"The subset file has been merged into {output_file_path}"
                        )
                        return
            except queue.Empty:
                pass
            except Exception as e:
                print(f"Error in update_progress: {e}")
            # Schedule the next check
            progress_window.after(100, update_progress)

        # Start updating the progress bar
        update_progress()


### End of RTVM Subset tool pack ################################################################################################################################################################################################################################           




    def open_management_window(self):
        # Create a new window
        self.management_window = tk.Toplevel(self.root)
        self.management_window.title("Management")

        # Set window size (optional)
        self.management_window.geometry("800x600")  # Adjust as needed

        # Create a frame for the buttons
        button_frame = tk.Frame(self.management_window)
        button_frame.pack(side=tk.TOP, fill=tk.X)

        # Export to Excel Button
        export_button = tk.Button(button_frame, text="Export to Excel", command=self.export_management_table)
        export_button.pack(side=tk.LEFT, padx=5, pady=5)

        # Create the management table
        columns = ("CDRL NUMBER (DI Number)", "Accepted", "Depreciated", "Proposed Add", "Proposed Delete", "Awaiting Input")
        self.management_table = ttk.Treeview(self.management_window, columns=columns, show="headings")
        self.management_table.pack(fill=tk.BOTH, expand=True)

        # Configure headings with sorting
        for col in columns:
            self.management_table.heading(col, text=col, command=lambda c=col: self.sort_management_table(c))
            self.management_table.column(col, anchor='center', width=130)  # Adjust width as needed

        # Populate the management table
        self.populate_management_table()


    def populate_management_table(self):
        # Clear existing data in the management table
        for item in self.management_table.get_children():
            self.management_table.delete(item)

        # Initialize a dictionary to hold aggregated data
        aggregated_data = {}

        # Loop through the status data
        for item in self.status_data:
            di_number = item.get('di_number', 'Unknown')
            object_status = item.get('object_status', '').lower()
            contractor_status = item.get('contractor_status', '').lower()
            government_status = item.get('government_status', '').lower()

            # Initialize the DI Number entry if not present
            if di_number not in aggregated_data:
                aggregated_data[di_number] = {
                    'Accepted': 0,
                    'Depreciated': 0,
                    'Proposed Add': 0,
                    'Proposed Delete': 0,
                    'Awaiting Input': 0
                }

            # Update counts based on statuses
            if object_status == 'accepted':
                aggregated_data[di_number]['Accepted'] += 1
            elif object_status == 'depreciated':
                aggregated_data[di_number]['Depreciated'] += 1

            # Update counts for 'Proposed Add' and 'Proposed Delete' based on contractor_status
            if object_status == 'proposed add':
                aggregated_data[di_number]['Proposed Add'] += 1
            elif object_status == 'proposed remove' or object_status == 'proposed delete':
                aggregated_data[di_number]['Proposed Delete'] += 1

            # Update count for 'Awaiting Input' based on government_status
            if government_status == 'awaiting input':
                aggregated_data[di_number]['Awaiting Input'] += 1

        # Populate the management table with aggregated data
        for di_number, counts in aggregated_data.items():
            self.management_table.insert(
                "", "end",
                values=(
                    di_number,
                    counts['Accepted'],
                    counts['Depreciated'],
                    counts['Proposed Add'],
                    counts['Proposed Delete'],
                    counts['Awaiting Input']
                )
            )






    def export_management_table(self):
        # Prompt the user to select a file location
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 title="Save Management Table as Excel File")
        if file_path:
            try:
                # Create a DataFrame from the management table
                columns = ["CDRL NUMBER (DI Number)", "Accepted", "Depreciated", "Proposed Add", "Proposed Delete", "Awaiting Input"]
                data = []
                for item in self.management_table.get_children():
                    values = self.management_table.item(item)['values']
                    data.append(values)
                df = pd.DataFrame(data, columns=columns)

                # Save the DataFrame to an Excel file
                df.to_excel(file_path, index=False)

                messagebox.showinfo("Success", f"Management table exported to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export to Excel: {e}")

    def sort_management_table(self, sort_column):
        # Determine the sort order
        if hasattr(self, 'sort_column') and self.sort_column == sort_column:
            # Toggle sort order
            self.sort_descending = not self.sort_descending
        else:
            # New sort column, default to ascending
            self.sort_descending = False
        self.sort_column = sort_column

        # Extract data from the table
        data = []
        for item in self.management_table.get_children():
            values = self.management_table.item(item)['values']
            data.append(values)

        # Get the index of the sort column
        columns = ["CDRL NUMBER (DI Number)", "Accepted", "Depreciated", "Proposed Add", "Proposed Delete", "Awaiting Input"]
        sort_col_index = columns.index(sort_column)

        # Convert the data to appropriate types for sorting
        for row in data:
            # Attempt to convert numerical values (except for the DI Number)
            for i in range(1, len(row)):
                try:
                    row[i] = int(row[i])
                except ValueError:
                    row[i] = 0  # Default to 0 if conversion fails

        # Sort the data
        data.sort(key=lambda x: x[sort_col_index], reverse=self.sort_descending)

        # Clear the table
        for item in self.management_table.get_children():
            self.management_table.delete(item)

        # Insert sorted data back into the table
        for row in data:
            self.management_table.insert("", "end", values=row)

    def on_progress_bar_click(self, event):
        # Identify the row under the cursor
        item_id = self.progress_table.identify_row(event.y)
        if item_id:
            # Get the row number from the item
            row_value = self.progress_table.item(item_id, 'values')[0]
            # Remove '>' if present
            row_value = row_value.lstrip('>')
            try:
                selected_row_number = int(row_value)
                # Adjust for zero-based index and header row
                new_row_index = selected_row_number - 2  # Assuming header is on Excel row 1
                max_row_index = len(self.df) - 1
                if 0 <= new_row_index <= max_row_index:
                    # Save any unsaved data before changing row
                    self.save_comments_to_excel()
                    # Update current row
                    self.current_row = new_row_index
                    # Update the UI
                    self.update_ui_after_navigation()
                    # Update row indicator
                    self.row_indicator_var.set(f"Row: {self.current_row + 2}")
                else:
                    messagebox.showerror(
                        "Invalid Row", "Selected row number is out of range.")
            except ValueError:
                pass  # If conversion to int fails, do nothing

    def toggle_progress_bar(self):
        if self.show_progress_var.get() == 1:
            # Show the progress bar table and label
            self.progress_label.grid()
            self.progress_table.grid()
            # Adjust column weights
            self.root.grid_columnconfigure(0, weight=1)
        else:
            # Hide the progress bar table and label
            self.progress_label.grid_remove()
            self.progress_table.grid_remove()
            # Adjust column weights
            self.root.grid_columnconfigure(0, weight=0)

    def populate_progress_table(self):
        if self.df is None:
            return
        try:
            # Clear existing items
            self.progress_table.delete(*self.progress_table.get_children())
            total_rows = len(self.df)
            for i in range(2, total_rows + 2):  # Rows are numbered starting from 2
                self.progress_table.insert("", "end", values=(str(i),))
        except Exception as e:
            print(f"Error populating progress table: {e}")

            
    def update_progress_bar_highlight(self):
        # First, remove existing highlights
        for item in self.progress_table.get_children():
            # Remove all custom tags
            self.progress_table.item(item, tags=())
            # Get the row number
            row_value = self.progress_table.item(item, 'values')[0]
            # Remove '>' if present
            if row_value.startswith('>'):
                row_value = row_value[1:]
                self.progress_table.item(item, values=(row_value,))
        # Now, highlight the current row and filtered rows
        current_row_number = self.current_row + 2  # Adjust for header
        for item in self.progress_table.get_children():
            row_value = self.progress_table.item(item, 'values')[0]
            df_row_index = int(row_value) - 2  # Adjust for DataFrame index
            tags = []
            # Check if this is the current row
            if int(row_value.lstrip('>')) == current_row_number:
                # Add '>' and 'current_row' tag
                self.progress_table.item(item, values=('>' + row_value.lstrip('>'),))
                tags.append('current_row')
            # Check if filters are applied and this row is in filtered rows
            elif hasattr(self, 'filtered_row_indices') and self.filtered_row_indices and df_row_index in self.filtered_row_indices:
                tags.append('filtered_row')
            # Set the tags
            self.progress_table.item(item, tags=tuple(tags))
        # Configure the highlight styles
        self.progress_table.tag_configure('current_row', background='lightblue')
        self.progress_table.tag_configure('filtered_row', background='yellow')


    def toggle_history_tables(self):
        if self.show_history_var.get() == 1:
            # Show the history tables and labels
            self.comment_history_label.grid()
            self.comment_history_table.grid()
            self.gov_comment_history_label.grid()
            self.gov_comment_history_table.grid()
            # Adjust column weights
            self.root.grid_columnconfigure(3, weight=1)
            self.root.grid_columnconfigure(4, weight=1)
            # Reset weights of columns 0-2
            self.root.grid_columnconfigure(0, weight=1)
            self.root.grid_columnconfigure(1, weight=1)
            self.root.grid_columnconfigure(2, weight=1)
        else:
            # Hide the history tables and labels
            self.comment_history_label.grid_remove()
            self.comment_history_table.grid_remove()
            self.gov_comment_history_label.grid_remove()
            self.gov_comment_history_table.grid_remove()
            # Adjust column weights
            self.root.grid_columnconfigure(3, weight=0)
            self.root.grid_columnconfigure(4, weight=0)
            # Increase weights of columns 0-2 to fill space
            self.root.grid_columnconfigure(0, weight=2)
            self.root.grid_columnconfigure(1, weight=2)
            self.root.grid_columnconfigure(2, weight=2)



    def upload_excel_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            try:
                # Load the main sheet (e.g. the default first sheet)
                self.df = pd.read_excel(file_path)
                # Also load the "VeriDOC Matrix View" sheet into matrix_df
                self.matrix_df = pd.read_excel(file_path, sheet_name="VERI-DOC Matrix")
                self.excel_file_path = file_path  # Store the file path
                self.current_row = 0  # Reset to the first row

                # Debug output: Print column names of both DataFrames
                print(f"Main DataFrame columns: {self.df.columns.tolist()}")
                print(f"Matrix DataFrame columns: {self.matrix_df.columns.tolist()}")

                # Process the main DataFrame to extract statuses
                self.process_statuses()

                # Populate the progress bar table
                self.populate_progress_table()

                # Update the UI
                self.update_ui_after_navigation()
                self.row_indicator_var.set(f"Row: {self.current_row + 2}")
                messagebox.showinfo("Success", "Excel file loaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load Excel file: {e}")


    def process_statuses(self):
        self.status_data = []  # List to store status info for each status entry
        self.unique_object_statuses = set()
        self.unique_contractor_statuses = set()
        self.unique_government_statuses = set()
        try:
            # Loop through the DataFrame
            for index, row in self.df.iterrows():
                data_cell = row[5]  # Adjust if the data is not in column F
                if pd.isna(data_cell):
                    continue
                elif not isinstance(data_cell, str):
                    data_cell = str(data_cell)
                # Split entries separated by lines of underscores
                entries = data_cell.split('______________________')
                for entry in entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                    # Initialize variables
                    data = {
                        'Object Identifier': "",
                        'DI Number': "",
                        'Object Status': "",
                        'Contractor Assessed Status': "",
                        'Government Assessed Status': ""
                    }
                    # Split entry into lines
                    lines = entry.split('\n')
                    for line in lines:
                        line = line.strip()
                        if ':' in line:
                            key, value = line.split(':', 1)
                            key = key.strip()
                            value = value.strip()
                            if key in data:
                                data[key] = value
                    # Add to status_data
                    obj_id = data['Object Identifier']
                    di_number = data['DI Number']
                    object_status = data['Object Status']
                    contractor_status = data['Contractor Assessed Status']
                    government_status = data['Government Assessed Status']
                    self.unique_object_statuses.add(object_status)
                    self.unique_contractor_statuses.add(contractor_status)
                    self.unique_government_statuses.add(government_status)
                    self.status_data.append({
                        'row_index': index,
                        'veridoc_number': obj_id,
                        'di_number': di_number,
                        'object_status': object_status,
                        'contractor_status': contractor_status,
                        'government_status': government_status
                    })
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process statuses: {e}")



    def open_filter_dialog(self):
        if not self.unique_object_statuses and not self.unique_contractor_statuses and not self.unique_government_statuses:
            messagebox.showerror(
                "Error", "No status data available. Please upload an Excel file with valid data.")
            return
        # Create a new window
        self.filter_window = tk.Toplevel(self.root)
        self.filter_window.title("Apply Filter")

        # Object Status
        object_label = tk.Label(self.filter_window, text="Object Status")
        object_label.pack()
        self.object_status_var = tk.StringVar()
        self.object_status_dropdown = ttk.Combobox(
            self.filter_window,
            textvariable=self.object_status_var,
            values=["Any"] + sorted(self.unique_object_statuses),
            state="readonly"
        )
        self.object_status_dropdown.pack()
        self.object_status_var.set("Any")  # Set default value

        # Contractor Assessed Status
        contractor_label = tk.Label(self.filter_window, text="Contractor Assessed Status")
        contractor_label.pack()
        self.contractor_status_var = tk.StringVar()
        self.contractor_status_dropdown = ttk.Combobox(
            self.filter_window,
            textvariable=self.contractor_status_var,
            values=["Any"] + sorted(self.unique_contractor_statuses),
            state="readonly"
        )
        self.contractor_status_dropdown.pack()
        self.contractor_status_var.set("Any")  # Set default value

        # Government Assessed Status
        government_label = tk.Label(self.filter_window, text="Government Assessed Status")
        government_label.pack()
        self.government_status_var = tk.StringVar()
        self.government_status_dropdown = ttk.Combobox(
            self.filter_window,
            textvariable=self.government_status_var,
            values=["Any"] + sorted(self.unique_government_statuses),
            state="readonly"
        )
        self.government_status_dropdown.pack()
        self.government_status_var.set("Any")  # Set default value

        # Apply Button
        apply_btn = tk.Button(self.filter_window, text="Apply", command=self.apply_filters)
        apply_btn.pack(pady=10)
        
    def apply_filters(self):
        # Get selected statuses
        selected_object_status = self.object_status_var.get()
        selected_contractor_status = self.contractor_status_var.get()
        selected_government_status = self.government_status_var.get()
        # Save comments before changing current row
        self.save_comments_to_excel()
        # Treat "Any" as no filter
        if selected_object_status == "Any":
            selected_object_status = ''
        if selected_contractor_status == "Any":
            selected_contractor_status = ''
        if selected_government_status == "Any":
            selected_government_status = ''

        # Filter the status data
        self.filtered_row_indices = []
        filtered_veridoc_numbers = []
        for item in self.status_data:
            # Check if the statuses match
            object_match = True
            contractor_match = True
            government_match = True

            if selected_object_status:
                object_match = (item['object_status'] == selected_object_status)
            if selected_contractor_status:
                contractor_match = (item['contractor_status'] == selected_contractor_status)
            if selected_government_status:
                government_match = (item['government_status'] == selected_government_status)

            if object_match and contractor_match and government_match:
                self.filtered_row_indices.append(item['row_index'])
                # Collect VeriDoc numbers
                filtered_veridoc_numbers.append(item['veridoc_number'])

        # Count unique VeriDoc numbers
        unique_veridoc_numbers = set(filtered_veridoc_numbers)

        # Reset the current row index to 0 in the filtered list
        self.current_filtered_index = 0
        if self.filtered_row_indices:
            self.current_row = self.filtered_row_indices[self.current_filtered_index]
            self.update_ui_after_navigation()
            self.row_indicator_var.set(f"Row: {self.current_row + 2}")
            messagebox.showinfo(
                "Filter Applied",
                f"{len(self.filtered_row_indices)} rows match the selected filters.\n"
                f"{len(unique_veridoc_numbers)} unique VeriDoc(s) found."
            )
        else:
            messagebox.showinfo(
                "No Results", "No rows match the selected filters.")
        # Close the filter window
        self.filter_window.destroy()
        # Update the progress bar highlight
        self.update_progress_bar_highlight()
    def clear_filters(self):
        if hasattr(self, 'filtered_row_indices'):
            del self.filtered_row_indices
        self.current_row = 0
        self.update_ui_after_navigation()
        self.row_indicator_var.set(f"Row: {self.current_row + 2}")
        messagebox.showinfo("Filter Cleared", "Filters have been cleared.")
        # Update the progress bar highlight
        self.update_progress_bar_highlight()

    def navigate_cells(self, direction):
        if self.df is None:
            messagebox.showinfo("No Data", "No file loaded. Please upload an Excel file first.")
            return
        
        if len(self.df) == 0:
            messagebox.showinfo("Empty File", "This file contains no data rows to navigate.")
            return

        # Instead of a blocking save, run the save operation in a background thread.
        self.save_comments_to_excel_background()

        if hasattr(self, 'filtered_row_indices') and self.filtered_row_indices:
            max_index = len(self.filtered_row_indices) - 1
            if direction == 'up':
                self.current_filtered_index = max(self.current_filtered_index - 1, 0)
            elif direction == 'down':
                self.current_filtered_index = min(self.current_filtered_index + 1, max_index)
            self.current_row = self.filtered_row_indices[self.current_filtered_index]
        else:
            max_row = len(self.df) - 1
            if direction == 'up':
                self.current_row = max(self.current_row - 1, 0)
            elif direction == 'down':
                self.current_row = min(self.current_row + 1, max_row)
        self.update_ui_after_navigation()
        self.row_indicator_var.set(f"Row: {self.current_row + 2}")


    def jump_to_cell(self):
        try:
            # Save comments before navigating away
            self.save_comments_to_excel()
            row_number = int(self.jump_to_var.get()) - 2  # Adjust for header
            if 0 <= row_number < len(self.df):
                self.current_row = row_number
                self.update_ui_after_navigation()
                self.row_indicator_var.set(f"Row: {self.current_row + 2}")
            else:
                messagebox.showerror(
                    "Invalid Row", "Row number out of range.")
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Please enter a valid row number.")


    def extract_info(self, data):
        # Clear highlights before clearing tables
        self.clear_highlights()

        # Clear the tables for new extraction
        self.table.delete(*self.table.get_children())
        self.comment_table.delete(*self.comment_table.get_children())

        # Check if data is empty
        if not data.strip():
            messagebox.showwarning("No Data", "No data available to extract.")
            return

        # Get comments from column H (index 7)
        comments_content = self.df.iloc[self.current_row, 7]
        if pd.isna(comments_content):
            comments_content = ""
        elif not isinstance(comments_content, str):
            comments_content = str(comments_content)

        # Split into lines and create a dictionary
        comment_lines = comments_content.strip().split('\n')
        comments_dict = {}
        for line in comment_lines:
            if ' - ' in line:
                obj_id, comment_text = line.split(' - ', 1)
                comments_dict[obj_id.strip()] = comment_text.strip()

        # Store comments for the current row
        self.current_comments = comments_dict

        # Regex pattern to match full entries including Object Identifier, DI Number, etc.
        pattern = (r'Object Identifier:\s*(?P<obj_id>WCC-VERI-DOC-(?P<veridoc_num>\d+)).*?'
                   r'DI Number:\s*(?P<di_num>\d+-\d+).*?'
                   r'CDRL Subtitle:\s*(?P<cdrl_subtitle>.*?)\n.*?'
                   r'Object Status:\s*(?P<object_status>.*?)\n.*?'
                   r'Contractor Assessed Status:\s*(?P<contractor_status>.*?)\n.*?'
                   r'Government Assessed Status:\s*(?P<government_status>.*?)\n')

        matches = re.finditer(pattern, data, re.DOTALL)

        count = 0  # Counter for the number of rows
        for match in matches:
            obj_id = match.group('obj_id')
            veridoc_num = match.group('veridoc_num')
            di_num = match.group('di_num')
            cdrl_subtitle = match.group('cdrl_subtitle').strip()
            object_status = match.group('object_status').strip()
            contractor_status = match.group('contractor_status').strip()
            government_status = match.group('government_status').strip()

            # Only insert into the table if we have a valid DI Number and CDRL Subtitle
            # This ensures that lines that are purely comments (without a proper DI Number or subtitle)
            # will not be inserted.
            if not di_num or not cdrl_subtitle:
                # Skip this match if it doesn't contain the required info
                continue

            # Insert the item into the data display table
            item_id = self.table.insert(
                "", "end",
                values=(obj_id, di_num, cdrl_subtitle, object_status, contractor_status, government_status)
            )

            # Get existing comment if any
            if obj_id in comments_dict:
                comment_text = f"{obj_id} - {comments_dict[obj_id]}"
            else:
                # Prepopulate the comment with the Object Identifier followed by " - "
                comment_text = f"{obj_id} - "

            self.comment_table.insert("", "end", values=(comment_text,))

            # Apply conditional formatting
            self.apply_conditional_formatting(item_id, contractor_status, government_status)

            # Check if Object Status is "DEPRECIATED"
           # if object_status.upper() == "DEPRECIATED" and not getattr(self, 'suppress_depreciated_warning', False):
           #     messagebox.showwarning(
           #         "Warning",
           #         "The Object Status is marked as DEPRECIATED.\nA CDRL that the government concurs should be removed "
           #         "is pending the full removal of the object from DOORs. No DLOC edits will be accepted for a "
           #         "VERI-DOC in this status."
           #     )

            count += 1  # Increment the counter

        # Adjust the table height based on the number of rows
        self.adjust_table_height(count)

    def adjust_table_height(self, row_count):
        # Set a maximum height to prevent the table from becoming too large
        max_height = 20  # You can adjust this value as needed
        height = min(row_count, max_height)
        self.table.config(height=height)
        self.comment_table.config(height=height)

    def apply_conditional_formatting(self, item_id, contractor_status, government_status):
        # Apply conditional formatting based on Contractor Assessed Status
        if contractor_status == "SAT":
            self.table.item(item_id, tags=("sat_row",))
        elif contractor_status == "UNSAT":
            self.table.item(item_id, tags=("unsat_row",))
        elif contractor_status == "TBD":
            self.table.item(item_id, tags=("tbd_row",))

        # Apply conditional formatting based on Government Assessed Status
        if government_status == "Agree":
            self.table.set(item_id, "#6", government_status)
            self.table.tag_configure("GovAgree.Cell", background="green", foreground="white")
            self.table.set(item_id, "Government Assessed Status", government_status)
            current_tags = self.table.item(item_id, "tags")
            if isinstance(current_tags, str):
                current_tags = (current_tags,)
            elif current_tags is None:
                current_tags = ()
            self.table.item(item_id, tags=current_tags + ("GovAgree.Cell",))
        elif government_status == "Disagree":
            self.table.set(item_id, "#6", government_status)
            self.table.tag_configure("GovDisagree.Cell", background="red", foreground="white")
            self.table.set(item_id, "Government Assessed Status", government_status)
            current_tags = self.table.item(item_id, "tags")
            if isinstance(current_tags, str):
                current_tags = (current_tags,)
            elif current_tags is None:
                current_tags = ()
            self.table.item(item_id, tags=current_tags + ("GovDisagree.Cell",))

        # Configure row tags
        self.table.tag_configure("sat_row", background="green", foreground="white")
        self.table.tag_configure("unsat_row", background="red", foreground="white")
        self.table.tag_configure("tbd_row", background="yellow", foreground="black")

    def update_proposed_changes_table(self):
        # Clear the table
        self.proposed_changes_table.delete(
            *self.proposed_changes_table.get_children())

        # Get the content from cell G (column index 6)
        content = self.df.iloc[self.current_row, 6]
        if pd.isna(content):
            content = ""
        elif not isinstance(content, str):
            content = str(content)

        # Split the content into lines and populate the table
        lines = content.strip().split('\n')
        for line in lines:
            if line.strip():
                self.proposed_changes_table.insert(
                    "", "end", values=(line.strip(),))

    def show_table_context_menu(self, event):
        # Identify the row under the cursor
        row_id = self.table.identify_row(event.y)
        if row_id:
            # Get values from the selected row
            item = self.table.item(row_id)
            values = item['values']
            if values:
                object_status = values[3]  # Object Status is at index 3
                if object_status.lower() == "depreciated":
                    messagebox.showinfo("Information", "This item is locked as it is depreciated. You do not need to do anything with this")
                else:
                    # Create the context menu
                    menu = tk.Menu(self.root, tearoff=0)
                    menu.add_command(label="Set DI Number to match CDRL", command=lambda: self.set_di_number(row_id))
                    menu.add_command(label="Delete DI Number", command=lambda: self.delete_di_number(row_id))
                
                    # If the Government Assessed Status is "disagree", add the new option.
                    gov_status = values[5].strip().lower() if len(values) > 5 else ""
                    if gov_status == "disagree":
                        menu.add_command(label="Find Disagreement Report", command=lambda: self.open_disagreement_report(row_id))
                
                    menu.post(event.x_root, event.y_root)
            else:
                messagebox.showerror("Error", "No data available for the selected row.")



    def open_disagreement_report(self, row_id):
        """
        Finds and opens the corresponding disagreement report PDF for the requirement in the
        DI Number Breakdown table, provided its Government Assessed Status is "disagree."
        The method extracts a tracking number from the DOORS SPEC ID (assumed in column 0),
        constructs the expected filename (e.g., "Disagreement Report - WCC-SPEC-1234.pdf"),
        and searches a designated disagreement folder for that file.
        """
        # Get the row values from the DI Number Breakdown table.
        values = self.table.item(row_id, 'values')
        if not values:
            messagebox.showerror("Error", "No data available for the selected row.")
            return

        # Ensure the Government Assessed Status is "disagree".
        gov_status = values[5].strip().lower() if len(values) > 5 else ""
        if gov_status != "disagree":
            messagebox.showinfo("Information", "This requirement does not have a 'Disagree' status.")
            return

        # Get the DOORS SPEC ID from column 0.
        doors_spec_id = values[0]
    
        # Extract the tracking number using a regex (assuming format "WCC-VERI-DOC-XXXX").
        import re
        m = re.search(r'WCC-VERI-DOC-(\d+)', doors_spec_id)
        if m:
            tracking_number = m.group(1)
        else:
            tracking_number = doors_spec_id  # Fallback if the pattern is not found

        # Construct the expected filename.
        expected_filename = f"Disagreement Report - WCC-SPEC-{tracking_number}.pdf"

        # Ensure that the disagreement database folder is set.
        if not hasattr(self, 'disagreement_folder') or not self.disagreement_folder:
            folder = filedialog.askdirectory(title="Select Disagreement Database Folder")
            if not folder:
                messagebox.showerror("Error", "No folder selected.")
                return
            self.disagreement_folder = folder

        # Search for the file in the disagreement folder (including subfolders).
        found_file = None
        for root_dir, dirs, files in os.walk(self.disagreement_folder):
            for file in files:
                if file == expected_filename:
                    found_file = os.path.join(root_dir, file)
                    break
            if found_file:
                break

        if found_file:
            try:
                os.startfile(found_file)  # Windows-specific; on other platforms, use an alternative method.
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file: {e}")
        else:
            messagebox.showinfo("Not Found", f"No disagreement report found with filename:\n{expected_filename}")





    def set_di_number(self, row_id):
        # Get values from the selected row
        item = self.table.item(row_id)
        values = item['values']
        if values:
            obj_id = values[0]  # VeriDoc Number (Object Identifier)
            di_num = values[1]
            # Use obj_id directly since it's already extracted

            # Open the PatternDialog pop-up window
            pattern_dialog = PatternDialog(
                self.root, self, obj_id, di_num, self.current_row)
            self.pattern_dialog = pattern_dialog  # Keep reference to the dialog
            # Pass deletions if needed
            pattern_dialog.deletions = self.deletions

            # Optionally, check for Government Assessed Status
            government_status = values[5]
            if government_status == "Pending Review":
                messagebox.showinfo("Information",
                                    "USCG is currently reviewing this line item, no action is required.")
            # Additional actions can be added here

        else:
            messagebox.showerror(
                "Error", "No data available for the selected row.")

    def delete_di_number(self, row_id):
        # Get values from the selected row
        item = self.table.item(row_id)
        values = item['values']
        if values:
            obj_id = values[0]  # VeriDoc Number (Object Identifier)
            # Create deletion pattern
            del_pattern = f"DEL;{obj_id}"

            # Save the deletion pattern to Excel
            self.save_deletion_to_excel(del_pattern)

            # Update the Proposed Changes Table
            self.update_proposed_changes_table()

            messagebox.showinfo("Success", "DI Number deleted and changes saved to Excel file successfully.")
        else:
            messagebox.showerror("Error", "No data available for the selected row.")


    def save_deletion_to_excel(self, del_pattern):
        # Get existing content from cell G (column index 6) of the current row
        existing_content = self.df.iloc[self.current_row, 6]
        if pd.isna(existing_content):
            existing_content = ""
        elif not isinstance(existing_content, str):
            existing_content = str(existing_content)

        # Append the new deletion pattern to the existing content
        if existing_content.strip():
            new_content = existing_content.strip() + "\n" + del_pattern
        else:
            new_content = del_pattern

        # Update the Excel file directly using openpyxl
        from openpyxl import load_workbook

        try:
            # Load the workbook
            wb = load_workbook(self.excel_file_path)
            ws = wb.active  # You may need to select the correct sheet if there are multiple

            # Calculate the Excel row number (considering headers)
            excel_row = self.current_row + 2  # Assuming header is on the first row

            # For debugging: print the file path and row number before saving
            print(f"Saving to file: {self.excel_file_path} at Excel row: {excel_row}")

            # Update the cell in column G (which is column index 7 in openpyxl)
            ws.cell(row=excel_row, column=7, value=new_content)

            # Save the workbook
            wb.save(self.excel_file_path)

            # Update the DataFrame in memory
            self.df.iloc[self.current_row, 6] = new_content

        except Exception as e:
            print(f"Failed to Save File: {self.excel_file_path} at Excel row: {excel_row}")
            messagebox.showerror("Error", f"Failed to save deletion to Excel file: {e}")

    def on_table_row_select(self, event):
        # Clear previous highlights
        self.clear_highlights()

        # Get selected item
        selected_item = self.table.selection()
        if selected_item:
            selected_item = selected_item[0]
            # Get VeriDoc Number from the selected row
            values = self.table.item(selected_item, 'values')
            veridoc_number = values[0]

            # Highlight rows in DI Number Breakdown table with the same VeriDoc Number
            for item in self.table.get_children():
                item_values = self.table.item(item, 'values')
                if item_values[0] == veridoc_number:
                    # Get current tags and ensure they are a tuple
                    current_tags = self.table.item(item, "tags")
                    if isinstance(current_tags, str):
                        current_tags = (current_tags,)
                    elif current_tags is None:
                        current_tags = ()
                    # Add the "highlight" tag
                    self.table.item(item, tags=current_tags + ("highlight",))
                    self.highlighted_items.append(('table', item))

            # Highlight corresponding rows in Comments table
            for item in self.comment_table.get_children():
                comment = self.comment_table.item(item, 'values')[0]
                # Extract the VeriDoc Number from the comment
                comment_veridoc_number = comment.split(' - ')[0]
                if comment_veridoc_number == veridoc_number:
                    # Get current tags and ensure they are a tuple
                    current_tags = self.comment_table.item(item, "tags")
                    if isinstance(current_tags, str):
                        current_tags = (current_tags,)
                    elif current_tags is None:
                        current_tags = ()
                    # Add the "highlight" tag
                    self.comment_table.item(item, tags=current_tags + ("highlight",))
                    self.highlighted_items.append(('comment_table', item))

            # Highlight corresponding rows in Contractor Proposed Change Request Input table
            for item in self.proposed_changes_table.get_children():
                pattern = self.proposed_changes_table.item(item, 'values')[0]
                # Extract the VeriDoc Number from the pattern
                pattern_veridoc_number = self.extract_veridoc_number_from_pattern(
                    pattern)
                if pattern_veridoc_number == veridoc_number:
                    # Get current tags and ensure they are a tuple
                    current_tags = self.proposed_changes_table.item(item, "tags")
                    if isinstance(current_tags, str):
                        current_tags = (current_tags,)
                    elif current_tags is None:
                        current_tags = ()
                    # Add the "highlight" tag
                    self.proposed_changes_table.item(item, tags=current_tags + ("highlight",))
                    self.highlighted_items.append(('proposed_changes_table', item))

            # Highlight corresponding rows in Contractor Proposed Change Comment History table
            for item in self.comment_history_table.get_children():
                history_entry = self.comment_history_table.item(item, 'values')[0]
                # Extract the VeriDoc Number from the history entry if possible
                history_veridoc_number = self.extract_veridoc_number_from_pattern(history_entry)
                if history_veridoc_number == veridoc_number:
                    # Get current tags and ensure they are a tuple
                    current_tags = self.comment_history_table.item(item, "tags")
                    if isinstance(current_tags, str):
                        current_tags = (current_tags,)
                    elif current_tags is None:
                        current_tags = ()
                    # Add the "highlight" tag
                    self.comment_history_table.item(item, tags=current_tags + ("highlight",))
                    self.highlighted_items.append(('comment_history_table', item))

            # Configure the highlight style
            self.table.tag_configure("highlight", background="lightblue")
            self.comment_table.tag_configure("highlight", background="lightblue")
            self.proposed_changes_table.tag_configure("highlight", background="lightblue")
            self.comment_history_table.tag_configure("highlight", background="lightblue")


    def extract_veridoc_number_from_pattern(self, pattern):
        # Try to extract the VeriDoc Number using regex
        match = re.search(r'(WCC-VERI-DOC-\d+)', pattern)
        if match:
            return match.group(1)
        else:
            return None

    def clear_highlights(self):
        # Remove highlight from previously highlighted items
        for table_name, item in self.highlighted_items:
            try:
                if table_name == 'table':
                    # Remove 'highlight' tag while preserving other tags
                    tags = list(self.table.item(item, 'tags'))
                    if 'highlight' in tags:
                        tags.remove('highlight')
                    self.table.item(item, tags=tuple(tags))
                elif table_name == 'comment_table':
                    self.comment_table.item(item, tags=())
                elif table_name == 'proposed_changes_table':
                    self.proposed_changes_table.item(item, tags=())
            except tk.TclError:
                # Item no longer exists, ignore the error
                pass
        self.highlighted_items = []

    def on_comment_double_click(self, event):
        # Identify the row and column under the cursor
        region = self.comment_table.identify("region", event.x, event.y)
        if region == "cell":
            row_id = self.comment_table.identify_row(event.y)
            column = self.comment_table.identify_column(event.x)
            x, y, width, height = self.comment_table.bbox(row_id, column)

            # Get the current value
            item = self.comment_table.item(row_id)
            current_text = item['values'][0] if item['values'] else ''

            # Create an entry widget over the cell
            self.entry_popup = tk.Entry(self.comment_table)
            self.entry_popup.place(x=x, y=y, width=width, height=height)
            self.entry_popup.insert(0, current_text)
            self.entry_popup.focus_set()
            self.entry_popup.bind(
                "<Return>", lambda event: self.on_entry_confirm(row_id))
            self.entry_popup.bind(
                "<FocusOut>", lambda event: self.on_entry_confirm(row_id))

    def on_entry_confirm(self, row_id):
        new_text = self.entry_popup.get()
        self.comment_table.item(row_id, values=(new_text,))
        self.entry_popup.destroy()

        # Update self.current_comments
        # Get the index of the item in comment_table
        item_index = self.comment_table.get_children().index(row_id)
        # Get the corresponding item in self.table
        table_items = self.table.get_children()
        if item_index < len(table_items):
            table_row_id = table_items[item_index]
            table_values = self.table.item(table_row_id, 'values')
            obj_id = table_values[0]  # VeriDoc Number
            # Extract the comment text from new_text
            if ' - ' in new_text:
                _, comment_text = new_text.split(' - ', 1)
                self.current_comments[obj_id] = comment_text.strip()
            else:
                # No ' - ', treat the entire text as comment
                self.current_comments[obj_id] = new_text.strip()


    def show_proposed_changes_context_menu(self, event):
        # Identify the row under the cursor
        row_id = self.proposed_changes_table.identify_row(event.y)
        if row_id:
            # Create a context menu
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(
                label="Edit Pattern", command=lambda: self.edit_proposed_change_row(row_id))
            menu.add_command(
                label="Delete Row and Save to Excel", command=lambda: self.delete_proposed_change_row(row_id))
            menu.post(event.x_root, event.y_root)
            

    def edit_proposed_change_row(self, row_id):
        # Get the pattern from the selected row
        item = self.proposed_changes_table.item(row_id)
        pattern_to_edit = item['values'][0]
    
        if not pattern_to_edit:
            messagebox.showerror("Error", "No pattern found in the selected row.")
            return
    
        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Pattern")
        edit_window.geometry("800x600")
        edit_window.transient(self.root)
        edit_window.grab_set()  # Make the window modal
    
        # Add guidelines frame from RTVM Desk Guide
        guidelines_frame = tk.LabelFrame(edit_window, text="RTVM Formatting Guidelines")
        guidelines_frame.pack(fill="x", padx=10, pady=5)
    
        guidelines_text = (
            "• NEVER USE SEMICOLONS (;) except as delimiters between elements\n"
            "• Use commas (,) between location elements\n"
            "• Capitalize The First Letter Of Each Word in location\n" 
            "• Avoid abbreviations (use 'Page' not 'PG', 'Sheet' not 'SHT')\n"
            "• For unknown Sheet/Page/Plan elements, use [] as placeholders\n"
            "• Format: ADD;DI-Number;CDRL Name, Page X, Plan View Y;Status\n"
            "• Valid statuses: SAT, UNSAT (don't use TBD)\n"
            "• Use DEL;WCC-VERI-DOC-XXXX to request deletion"
        )
    
        guidelines_label = tk.Label(guidelines_frame, text=guidelines_text, justify="left", anchor="w")
        guidelines_label.pack(fill="x", padx=5, pady=5)
    
        # Toggle button to show/hide guidelines
        guidelines_visible = tk.BooleanVar(value=True)
    
        def toggle_guidelines():
            if guidelines_visible.get():
                guidelines_label.pack_forget()
                guidelines_visible.set(False)
                toggle_button.config(text="Show Guidelines")
            else:
                guidelines_label.pack(fill="x", padx=5, pady=5)
                guidelines_visible.set(True)
                toggle_button.config(text="Hide Guidelines")
    
        toggle_button = tk.Button(guidelines_frame, text="Hide Guidelines", command=toggle_guidelines)
        toggle_button.pack(side="right", padx=5, pady=2)
    
        # Add instructions
        instruction_label = tk.Label(edit_window, text="Edit the pattern below:", anchor="w")
        instruction_label.pack(fill="x", padx=10, pady=(10, 5))
    
        # Create text widget for editing with syntax highlighting (visual cues)
        edit_frame = tk.Frame(edit_window)
        edit_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
        edit_text = tk.Text(edit_frame, wrap="word", height=10, width=80, font=("Courier", 10))
        edit_text.pack(fill="both", expand=True)
        edit_text.insert("1.0", pattern_to_edit)
    
        # Add syntax highlighting for pattern parts
        def highlight_syntax():
            # Clear all tags
            for tag in edit_text.tag_names():
                if tag != "sel":  # Don't remove selection tag
                    edit_text.tag_remove(tag, "1.0", "end")
        
            # Get current text
            text = edit_text.get("1.0", "end-1c")
        
            # Highlight based on pattern type
            if text.startswith("ADD;"):
                # ADD pattern
                edit_text.tag_add("pattern_type", "1.0", "1.4")
            
                parts = text.split(";")
                if len(parts) > 1:
                    # DI Number
                    di_start = text.find(";") + 1
                    di_end = text.find(";", di_start)
                    if di_end > di_start:
                        edit_text.tag_add("di_number", f"1.{di_start}", f"1.{di_end}")
                    
                        # Location
                        if len(parts) > 2:
                            loc_start = di_end + 1
                            status_start = text.rfind(";")
                            if status_start > loc_start:
                                edit_text.tag_add("location", f"1.{loc_start}", f"1.{status_start}")
                            
                                # Status
                                if len(parts) > 3:
                                    edit_text.tag_add("status", f"1.{status_start+1}", "end-1c")
                                
                                    # Check if status is valid
                                    status = parts[3].strip().upper()
                                    if status not in ["SAT", "UNSAT"]:
                                        edit_text.tag_add("error", f"1.{status_start+1}", "end-1c")
        
            elif text.startswith("DEL;"):
                # DEL pattern
                edit_text.tag_add("pattern_type", "1.0", "1.4")
            
                # VeriDoc
                if ";" in text:
                    veridoc_start = text.find(";") + 1
                    edit_text.tag_add("veridoc", f"1.{veridoc_start}", "end-1c")
        
            elif ";" in text:
                # Update pattern (VeriDoc;Location;Status)
                parts = text.split(";")
            
                # VeriDoc
                veridoc_end = text.find(";")
                edit_text.tag_add("veridoc", "1.0", f"1.{veridoc_end}")
            
                if len(parts) > 1:
                    # Location
                    loc_start = veridoc_end + 1
                    if len(parts) > 2:
                        status_start = text.rfind(";")
                        edit_text.tag_add("location", f"1.{loc_start}", f"1.{status_start}")
                    
                        # Status
                        edit_text.tag_add("status", f"1.{status_start+1}", "end-1c")
                    
                        # Check if status is valid
                        status = parts[2].strip().upper()
                        if status not in ["SAT", "UNSAT"]:
                            edit_text.tag_add("error", f"1.{status_start+1}", "end-1c")
                    else:
                        edit_text.tag_add("location", f"1.{loc_start}", "end-1c")
        
            # Check for common errors
            if ";" in text:
                # Check for extra spaces around semicolons (per RTVM Desk Guide)
                for match in re.finditer(r"\s+;|;\s+", text):
                    start, end = match.span()
                    edit_text.tag_add("error", f"1.{start}", f"1.{end}")
            
                # Check for abbreviations in location
                for abbr in ["PG", "SHT", "SEC", "PARA"]:
                    for match in re.finditer(r"\b" + abbr + r"\b", text):
                        start, end = match.span()
                        edit_text.tag_add("warning", f"1.{start}", f"1.{end}")
    
        # Apply tag configurations
        edit_text.tag_configure("pattern_type", foreground="blue", font=("Courier", 10, "bold"))
        edit_text.tag_configure("di_number", foreground="green", font=("Courier", 10, "bold"))
        edit_text.tag_configure("veridoc", foreground="purple", font=("Courier", 10, "bold"))
        edit_text.tag_configure("location", foreground="black")
        edit_text.tag_configure("status", foreground="orange", font=("Courier", 10, "bold"))
        edit_text.tag_configure("error", background="pink", underline=True)
        edit_text.tag_configure("warning", background="yellow")
    
        # Run initial highlighting
        highlight_syntax()
    
        # Bind key release to update highlighting
        edit_text.bind("<KeyRelease>", lambda e: highlight_syntax())
    
        # Parse the pattern to determine its type
        pattern_type = "unknown"
        if pattern_to_edit.startswith("ADD;"):
            pattern_type = "add"
        elif pattern_to_edit.startswith("DEL;"):
            pattern_type = "delete"
        else:
            pattern_type = "update"
    
        # Add pattern structure assistance frame
        structure_frame = tk.LabelFrame(edit_window, text="Pattern Structure Reference")
        structure_frame.pack(fill="x", padx=10, pady=5)
    
        if pattern_type == "add":
            structure_text = "ADD;DI-Number;CDRL Name, Page/Sheet X, Plan View/Section Y;SAT/UNSAT"
        elif pattern_type == "delete":
            structure_text = "DEL;WCC-VERI-DOC-XXXX"
        else:  # update
            structure_text = "WCC-VERI-DOC-XXXX;CDRL Name, Page/Sheet X, Plan View/Section Y;SAT/UNSAT"
    
        structure_label = tk.Label(structure_frame, text=structure_text, font=("Courier", 10))
        structure_label.pack(fill="x", padx=5, pady=5)
    
        # Add example frame from RTVM Desk Guide
        example_frame = tk.LabelFrame(edit_window, text="Examples from RTVM Desk Guide")
        example_frame.pack(fill="x", padx=10, pady=5)
    
        examples_text = (
            "Update Existing: WCC-VERI-DOC-169;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;SAT\n"
            "Add New CDRL: ADD;070-001;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT\n"
            "Delete VeriDoc: DEL;WCC-VERI-DOC-169"
        )
    
        examples_label = tk.Label(example_frame, text=examples_text, justify="left", font=("Courier", 9))
        examples_label.pack(fill="x", padx=5, pady=5)
    
        # Create buttons frame
        button_frame = tk.Frame(edit_window)
        button_frame.pack(fill="x", padx=10, pady=10)
    
        # Helper button frame for quick edits
        helper_frame = tk.Frame(edit_window)
        helper_frame.pack(fill="x", padx=10, pady=5)
    
        # Status quick buttons
        tk.Label(helper_frame, text="Set Status:").pack(side="left", padx=(0, 5))
    
        # Function to set status
        def set_status(status):
            current_text = edit_text.get("1.0", "end-1c")
            parts = current_text.split(";")
        
            # Only applicable for patterns with status (ADD or Update)
            if pattern_type == "add" or pattern_type == "update":
                if len(parts) >= 3:
                    # Replace the last part with the new status
                    parts[-1] = status
                    new_text = ";".join(parts)
                
                    # Update the text widget
                    edit_text.delete("1.0", "end")
                    edit_text.insert("1.0", new_text)
                    highlight_syntax()
    
        tk.Button(helper_frame, text="SAT", command=lambda: set_status("SAT")).pack(side="left", padx=5)
        tk.Button(helper_frame, text="UNSAT", command=lambda: set_status("UNSAT")).pack(side="left", padx=5)
    
        # Function to capitalize properly
        def capitalize_location():
            current_text = edit_text.get("1.0", "end-1c")
            parts = current_text.split(";")
        
            # Only applicable for patterns with location (ADD or Update)
            if (pattern_type == "add" and len(parts) >= 3) or (pattern_type == "update" and len(parts) >= 2):
                location_index = 2 if pattern_type == "add" else 1
            
                # Split location by commas
                location_parts = parts[location_index].split(",")
            
                # Capitalize each part
                for i, part in enumerate(location_parts):
                    # Preserve special formatting like 160_WLIC_WCC
                    if "_" in part:
                        location_parts[i] = part.strip()
                    else:
                        # Capitalize first letter of each word
                        location_parts[i] = " ".join(word.capitalize() for word in part.strip().split())
            
                # Rebuild location
                parts[location_index] = ", ".join(location_parts)
            
                # Rebuild full pattern
                new_text = ";".join(parts)
            
                # Update the text widget
                edit_text.delete("1.0", "end")
                edit_text.insert("1.0", new_text)
                highlight_syntax()
    
        # Fix abbreviations button
        def fix_abbreviations():
            current_text = edit_text.get("1.0", "end-1c")
        
            # Common abbreviations to fix
            replacements = {
                "PG ": "Page ",
                "SHT ": "Sheet ",
                "SEC ": "Section ",
                "PARA ": "Paragraph "
            }
        
            new_text = current_text
            for abbr, full in replacements.items():
                new_text = new_text.replace(abbr, full)
                new_text = new_text.replace(abbr.lower(), full)
        
            # Update the text widget
            edit_text.delete("1.0", "end")
            edit_text.insert("1.0", new_text)
            highlight_syntax()
    
        # Add formatting help buttons
        tk.Button(helper_frame, text="Capitalize Location", command=capitalize_location).pack(side="left", padx=5)
        tk.Button(helper_frame, text="Fix Abbreviations", command=fix_abbreviations).pack(side="left", padx=5)
    
        # Function to save the changes
        def save_changes():
            new_pattern = edit_text.get("1.0", "end-1c").strip()  # Get text without trailing newline
            if not new_pattern:
                messagebox.showerror("Error", "Pattern cannot be empty.")
                return
            
            # Validate the pattern based on type
            parts = new_pattern.split(";")
        
            validation_error = None
        
            # Check for spaces around semicolons
            if re.search(r"\s+;|;\s+", new_pattern):
                validation_error = "Remove spaces before or after semicolons."
        
            # Pattern-specific validation
            if new_pattern.startswith("ADD;"):
                if len(parts) < 4:
                    validation_error = "ADD pattern must have format: ADD;DI-Number;Location;Status"
                elif not parts[3] or parts[3].strip().upper() not in ["SAT", "UNSAT"]:
                    validation_error = "Status must be either SAT or UNSAT, not TBD or empty."
            elif new_pattern.startswith("DEL;"):
                if len(parts) != 2 or not parts[1]:
                    validation_error = "DEL pattern must have format: DEL;WCC-VERI-DOC-XXXX"
            elif ";" in new_pattern:  # Update pattern
                if len(parts) < 3:
                    validation_error = "Update pattern must have format: WCC-VERI-DOC-XXXX;Location;Status"
                elif not parts[2] or parts[2].strip().upper() not in ["SAT", "UNSAT"]:
                    validation_error = "Status must be either SAT or UNSAT, not TBD or empty."
            
            if validation_error:
                if not messagebox.askyesno("Validation Warning", 
                                          f"The pattern may have issues: {validation_error}\n\nSave anyway?"):
                    return
        
            # Get current content from Excel
            content = self.df.iloc[self.current_row, 6]
            if pd.isna(content):
                content = ""
            elif not isinstance(content, str):
                content = str(content)
            
            # Split the content into lines
            lines = content.strip().split('\n')
        
            # Find and replace the pattern
            found = False
            for i, line in enumerate(lines):
                if line == pattern_to_edit:
                    lines[i] = new_pattern
                    found = True
                    break
                
            if not found:
                messagebox.showerror("Error", "Original pattern not found in the Excel data.")
                return
            
            # Join the lines back
            new_content = '\n'.join(lines)
        
            # Update the Excel file
            try:
                # Load the workbook
                wb = load_workbook(self.excel_file_path)
                ws = wb.active  # You may need to select the correct sheet if there are multiple
            
                # Calculate the Excel row number (considering headers)
                excel_row = self.current_row + 2  # Assuming header is on the first row
            
                # Update the cell in column G (which is column index 7 in openpyxl)
                ws.cell(row=excel_row, column=7, value=new_content)
            
                # Save the workbook
                wb.save(self.excel_file_path)
            
                # Update the DataFrame in memory
                self.df.iloc[self.current_row, 6] = new_content
            
                messagebox.showinfo("Success", "Pattern updated and saved to Excel file successfully.")
            
                # Update the table
                self.update_proposed_changes_table()
            
                # Close the edit window
                edit_window.destroy()
            
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save to Excel file: {e}")
    
        # Add save and cancel buttons
        save_button = tk.Button(button_frame, text="Save Changes", command=save_changes)
        save_button.pack(side="left", padx=5)
    
        cancel_button = tk.Button(button_frame, text="Cancel", command=edit_window.destroy)
        cancel_button.pack(side="left", padx=5)

    def add_new_pattern(self):
        # Create a new window for adding a pattern
        add_window = tk.Toplevel(self.root)
        add_window.title("Add New Pattern")
        add_window.geometry("800x650")
        add_window.transient(self.root)
        add_window.grab_set()  # Make the window modal
    
        # Add guidelines frame from RTVM Desk Guide
        guidelines_frame = tk.LabelFrame(add_window, text="RTVM Formatting Guidelines")
        guidelines_frame.pack(fill="x", padx=10, pady=5)
    
        guidelines_text = (
            "• NEVER USE SEMICOLONS (;) except as delimiters between elements\n"
            "• Use commas (,) between location elements\n"
            "• Capitalize The First Letter Of Each Word in location\n" 
            "• Avoid abbreviations (use 'Page' not 'PG', 'Sheet' not 'SHT')\n"
            "• For unknown Sheet/Page/Plan elements, use [] as placeholders\n"
            "• Format: ADD;DI-Number;CDRL Name, Page X, Plan View Y;Status\n"
            "• Valid statuses: SAT, UNSAT (don't use TBD for submissions)\n"
            "• Use DEL;WCC-VERI-DOC-XXXX to request deletion"
        )
    
        guidelines_label = tk.Label(guidelines_frame, text=guidelines_text, justify="left", anchor="w")
        guidelines_label.pack(fill="x", padx=5, pady=5)
    
        # Toggle button to show/hide guidelines
        guidelines_visible = tk.BooleanVar(value=True)
    
        def toggle_guidelines():
            if guidelines_visible.get():
                guidelines_label.pack_forget()
                guidelines_visible.set(False)
                toggle_button.config(text="Show Guidelines")
            else:
                guidelines_label.pack(fill="x", padx=5, pady=5)
                guidelines_visible.set(True)
                toggle_button.config(text="Hide Guidelines")
    
        toggle_button = tk.Button(guidelines_frame, text="Hide Guidelines", command=toggle_guidelines)
        toggle_button.pack(side="right", padx=5, pady=2)
    
        # Pattern Type Selection Frame
        pattern_type_frame = tk.LabelFrame(add_window, text="Pattern Type")
        pattern_type_frame.pack(fill="x", padx=10, pady=5)
    
        pattern_type_var = tk.StringVar(value="add")
    
        # ADD new CDRL
        add_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Add New CDRL (ADD)", 
            variable=pattern_type_var, 
            value="add",
            command=lambda: toggle_input_fields())
        add_radio.grid(row=0, column=0, sticky="w", padx=5, pady=2)
    
        # Update existing VeriDoc
        update_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Update Existing VeriDoc", 
            variable=pattern_type_var, 
            value="update",
            command=lambda: toggle_input_fields())
        update_radio.grid(row=0, column=1, sticky="w", padx=5, pady=2)
    
        # DELete existing VeriDoc
        delete_radio = tk.Radiobutton(
            pattern_type_frame, 
            text="Delete VeriDoc (DEL)", 
            variable=pattern_type_var, 
            value="delete",
            command=lambda: toggle_input_fields())
        delete_radio.grid(row=0, column=2, sticky="w", padx=5, pady=2)
    
        # Create form for input fields
        input_frame = tk.LabelFrame(add_window, text="Pattern Details")
        input_frame.pack(fill="x", padx=10, pady=5)
    
        # VeriDoc Number (for update and delete patterns)
        veridoc_frame = tk.Frame(input_frame)
        veridoc_frame.pack(fill="x", padx=5, pady=5)
    
        veridoc_label = tk.Label(veridoc_frame, text="VeriDoc Number:", width=15, anchor="e")
        veridoc_label.pack(side="left", padx=(5, 2))
    
        veridoc_entry = tk.Entry(veridoc_frame, width=25)
        veridoc_entry.pack(side="left", padx=2)
    
        veridoc_help = tk.Label(veridoc_frame, text="(Format: WCC-VERI-DOC-XXXX)", font=("Helvetica", 9, "italic"))
        veridoc_help.pack(side="left", padx=2)
    
        # Select from table button
        def select_from_table():
            # Get selected item from the DI Number Breakdown table
            selected_items = self.table.selection()
            if not selected_items:
                messagebox.showinfo("No Selection", "Please select a row in the DI Number Breakdown table first.")
                return
            
            # Get the VeriDoc number from the selected row
            item = self.table.item(selected_items[0])
            values = item['values']
            if values and len(values) > 0:
                veridoc = values[0]  # VeriDoc is in the first column
                veridoc_entry.delete(0, tk.END)
                veridoc_entry.insert(0, veridoc)
    
        veridoc_select_button = tk.Button(veridoc_frame, text="Select from Table", command=select_from_table)
        veridoc_select_button.pack(side="left", padx=5)
    
        # DI Number (for ADD patterns)
        di_frame = tk.Frame(input_frame)
        di_frame.pack(fill="x", padx=5, pady=5)
    
        di_label = tk.Label(di_frame, text="DI Number:", width=15, anchor="e")
        di_label.pack(side="left", padx=(5, 2))
    
        di_entry = tk.Entry(di_frame, width=15)
        di_entry.pack(side="left", padx=2)
    
        di_help = tk.Label(di_frame, text="(Format: XXX-XXX)", font=("Helvetica", 9, "italic"))
        di_help.pack(side="left", padx=2)
    
        # Common DI Numbers from desk guide
        common_dis_frame = tk.Frame(di_frame)
        common_dis_frame.pack(side="left", padx=10)
    
        tk.Label(common_dis_frame, text="Common:").pack(side="left")
    
        def insert_di(di):
            di_entry.delete(0, tk.END)
            di_entry.insert(0, di)
    
        common_dis = [
            ("070-001", lambda: insert_di("070-001")),
            ("100-001", lambda: insert_di("100-001")),
            ("073-002", lambda: insert_di("073-002"))
        ]
    
        for di_text, command in common_dis:
            tk.Button(common_dis_frame, text=di_text, command=command, padx=2, pady=0).pack(side="left", padx=2)
    
        # CDRL Name
        cdrl_frame = tk.Frame(input_frame)
        cdrl_frame.pack(fill="x", padx=5, pady=5)
    
        cdrl_label = tk.Label(cdrl_frame, text="CDRL Name:", width=15, anchor="e")
        cdrl_label.pack(side="left", padx=(5, 2))
    
        cdrl_entry = tk.Entry(cdrl_frame, width=30)
        cdrl_entry.pack(side="left", padx=2, fill="x", expand=True)
    
        # Variant template buttons
        variant_frame = tk.Frame(cdrl_frame)
        variant_frame.pack(side="left", padx=10)
    
        tk.Label(variant_frame, text="Insert Template:").pack(side="left")
    
        def insert_variant_template(variant):
            # Get the DI number
            di_num = di_entry.get().strip()
            if not di_num:
                messagebox.showinfo("Info", "Please enter a DI Number first.")
                return
        
            # Create CDRL name with variant and DI number
            variant_format = f"{variant}_WCC_{di_num.replace('-', '_')}"
            cdrl_entry.delete(0, tk.END)
            cdrl_entry.insert(0, variant_format)
    
        tk.Button(variant_frame, text="160-WLIC", 
                  command=lambda: insert_variant_template("160-WLIC")).pack(side="left", padx=2)
        tk.Button(variant_frame, text="180-WLR", 
                  command=lambda: insert_variant_template("180-WLR")).pack(side="left", padx=2)
    
        # Page/Sheet
        page_frame = tk.Frame(input_frame)
        page_frame.pack(fill="x", padx=5, pady=5)
    
        page_label = tk.Label(page_frame, text="Page/Sheet:", width=15, anchor="e")
        page_label.pack(side="left", padx=(5, 2))
    
        page_type_var = tk.StringVar(value="Page")
        page_type = ttk.Combobox(
            page_frame, 
            textvariable=page_type_var, 
            values=["Page", "Sheet"], 
            state="readonly",
            width=8)
        page_type.pack(side="left", padx=2)
    
        page_entry = tk.Entry(page_frame, width=10)
        page_entry.pack(side="left", padx=2)
    
        page_help = tk.Label(page_frame, 
                             text="(Use [] for unknown values per RTVM Desk Guide)", 
                             font=("Helvetica", 9, "italic"))
        page_help.pack(side="left", padx=2)
    
        # Plan View/Section
        plan_frame = tk.Frame(input_frame)
        plan_frame.pack(fill="x", padx=5, pady=5)
    
        plan_label = tk.Label(plan_frame, text="Plan View/Section:", width=15, anchor="e")
        plan_label.pack(side="left", padx=(5, 2))
    
        plan_type_var = tk.StringVar(value="Plan View")
        plan_type = ttk.Combobox(
            plan_frame, 
            textvariable=plan_type_var, 
            values=["Plan View", "Section"], 
            state="readonly",
            width=8)
        plan_type.pack(side="left", padx=2)
    
        plan_entry = tk.Entry(plan_frame, width=10)
        plan_entry.pack(side="left", padx=2)
    
        plan_help = tk.Label(plan_frame, 
                             text="(Use [] for unknown values per RTVM Desk Guide)", 
                             font=("Helvetica", 9, "italic"))
        plan_help.pack(side="left", padx=2)
    
        # Status
        status_frame = tk.Frame(input_frame)
        status_frame.pack(fill="x", padx=5, pady=5)
    
        status_label = tk.Label(status_frame, text="Status:", width=15, anchor="e")
        status_label.pack(side="left", padx=(5, 2))
    
        status_var = tk.StringVar(value="")
        status_dropdown = ttk.Combobox(
            status_frame, 
            textvariable=status_var, 
            values=["SAT", "UNSAT"], 
            state="readonly",
            width=8)
        status_dropdown.pack(side="left", padx=2)
    
        status_help = tk.Label(status_frame, 
                               text="(RTVM Desk Guide states: Do not use TBD for submissions)", 
                               font=("Helvetica", 9, "italic"), fg="red")
        status_help.pack(side="left", padx=2)
    
        # Function to toggle input fields based on pattern type
        def toggle_input_fields():
            pattern_type = pattern_type_var.get()
        
            # Reset all frames
            for frame in [veridoc_frame, di_frame, cdrl_frame, page_frame, plan_frame, status_frame]:
                frame.pack_forget()
        
            # Show relevant frames based on pattern type
            if pattern_type == "add":
                di_frame.pack(fill="x", padx=5, pady=5)
                cdrl_frame.pack(fill="x", padx=5, pady=5)
                page_frame.pack(fill="x", padx=5, pady=5)
                plan_frame.pack(fill="x", padx=5, pady=5)
                status_frame.pack(fill="x", padx=5, pady=5)
            elif pattern_type == "update":
                veridoc_frame.pack(fill="x", padx=5, pady=5)
                cdrl_frame.pack(fill="x", padx=5, pady=5)
                page_frame.pack(fill="x", padx=5, pady=5)
                plan_frame.pack(fill="x", padx=5, pady=5)
                status_frame.pack(fill="x", padx=5, pady=5)
            elif pattern_type == "delete":
                veridoc_frame.pack(fill="x", padx=5, pady=5)
    
        # Initialize visible frames
        toggle_input_fields()
    
        # Pattern preview
        preview_frame = tk.LabelFrame(add_window, text="Pattern Preview")
        preview_frame.pack(fill="x", padx=10, pady=5)
    
        preview_text = tk.Entry(preview_frame, width=80, font=("Courier", 10))
        preview_text.pack(fill="x", padx=5, pady=5)
        preview_text.config(state="readonly")
    
        # Function to generate and update the preview
        def update_preview():
            pattern_type = pattern_type_var.get()
        
            # Clear the preview
            preview_text.config(state="normal")
            preview_text.delete(0, tk.END)
        
            # Get input values
            veridoc = veridoc_entry.get().strip().upper()
            di_num = di_entry.get().strip()
            cdrl_name = cdrl_entry.get().strip().upper()
            page_sheet = page_entry.get().strip()
            page_sheet_type = page_type_var.get()
            plan_view = plan_entry.get().strip()
            plan_view_type = plan_type_var.get()
            status = status_var.get()
        
            # Generate pattern based on type
            pattern = ""
            if pattern_type == "add":
                if not di_num:
                    preview_text.config(state="readonly")
                    return
                
                location = ""
                if cdrl_name or page_sheet or plan_view:
                    location_parts = []
                    if cdrl_name:
                        location_parts.append(cdrl_name)
                    if page_sheet:
                        location_parts.append(f"{page_sheet_type} {page_sheet}")
                    if plan_view:
                        location_parts.append(f"{plan_view_type} {plan_view}")
                    location = ", ".join(location_parts)
                
                if location and status:
                    pattern = f"ADD;{di_num};{location};{status}"
                elif location:
                    pattern = f"ADD;{di_num};{location}"
                else:
                    pattern = f"ADD;{di_num}"
                
            elif pattern_type == "update":
                if not veridoc:
                    preview_text.config(state="readonly")
                    return
                
                location = ""
                if cdrl_name or page_sheet or plan_view:
                    location_parts = []
                    if cdrl_name:
                        location_parts.append(cdrl_name)
                    if page_sheet:
                        location_parts.append(f"{page_sheet_type} {page_sheet}")
                    if plan_view:
                        location_parts.append(f"{plan_view_type} {plan_view}")
                    location = ", ".join(location_parts)
                
                if location and status:
                    pattern = f"{veridoc};{location};{status}"
                elif location:
                    pattern = f"{veridoc};{location}"
                else:
                    pattern = f"{veridoc}"
                
            elif pattern_type == "delete":
                if veridoc:
                    pattern = f"DEL;{veridoc}"
        
            # Update preview
            preview_text.insert(0, pattern)
            preview_text.config(state="readonly")
    
        # Bind field changes to update preview
        def bind_update_preview(widget):
            if isinstance(widget, tk.Entry):
                widget.bind("<KeyRelease>", lambda e: update_preview())
            elif isinstance(widget, ttk.Combobox):
                widget.bind("<<ComboboxSelected>>", lambda e: update_preview())
    
        for widget in [veridoc_entry, di_entry, cdrl_entry, page_entry, plan_entry]:
            bind_update_preview(widget)
    
        for widget in [page_type, plan_type, status_dropdown]:
            bind_update_preview(widget)
    
        # Button frame
        button_frame = tk.Frame(add_window)
        button_frame.pack(fill="x", padx=10, pady=10)
    
        # Function to add the pattern to Excel
        def add_pattern_to_excel():
            # Get the pattern from preview
            preview_text.config(state="normal")
            pattern = preview_text.get().strip()
            preview_text.config(state="readonly")
        
            if not pattern:
                messagebox.showerror("Error", "No pattern to add.")
                return
        
            # Validation
            validation_error = None
            pattern_type = pattern_type_var.get()
        
            if pattern_type == "add":
                di_num = di_entry.get().strip()
                if not di_num:
                    validation_error = "DI Number is required for ADD pattern."
                if status_var.get() and status_var.get() not in ["SAT", "UNSAT"]:
                    validation_error = "Status must be either SAT or UNSAT, not TBD or empty."
            elif pattern_type == "update":
                veridoc = veridoc_entry.get().strip()
                if not veridoc:
                    validation_error = "VeriDoc Number is required for Update pattern."
                if status_var.get() and status_var.get() not in ["SAT", "UNSAT"]:
                    validation_error = "Status must be either SAT or UNSAT, not TBD or empty."
            elif pattern_type == "delete":
                veridoc = veridoc_entry.get().strip()
                if not veridoc:
                    validation_error = "VeriDoc Number is required for DEL pattern."
        
            if validation_error:
                messagebox.showerror("Validation Error", validation_error)
                return
        
            # Get current content from Excel
            content = self.df.iloc[self.current_row, 6]
            if pd.isna(content):
                content = ""
            elif not isinstance(content, str):
                content = str(content)
            
            # Append the new pattern
            if content.strip():
                new_content = content.strip() + "\n" + pattern
            else:
                new_content = pattern
            
            # Update the Excel file
            try:
                # Load the workbook
                wb = load_workbook(self.excel_file_path)
                ws = wb.active
            
                # Calculate the Excel row number
                excel_row = self.current_row + 2
            
                # Update the cell in column G
                ws.cell(row=excel_row, column=7, value=new_content)
            
                # Save the workbook
                wb.save(self.excel_file_path)
            
                # Update the DataFrame in memory
                self.df.iloc[self.current_row, 6] = new_content
            
                messagebox.showinfo("Success", "Pattern added and saved to Excel file successfully.")
            
                # Update the table
                self.update_proposed_changes_table()
            
                # Close the add window
                add_window.destroy()
            
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save to Excel file: {e}")
    
        # Add a checkbox for adding both variants at once
        both_variants_var = tk.BooleanVar(value=False)
    
        def toggle_both_variants():
            # Update visibility of variant warning
            if both_variants_var.get() and pattern_type_var.get() != "delete":
                variant_warning_label.grid()
            else:
                variant_warning_label.grid_forget()
    
        both_variants_check = tk.Checkbutton(
            button_frame, 
            text="Also create 180-WLR variant pattern", 
            variable=both_variants_var,
            command=toggle_both_variants)
        both_variants_check.pack(side="left", padx=5)
    
        # Variant warning label
        variant_warning_label = tk.Label(
            button_frame, 
            text="(Only applicable for ADD and Update patterns)", 
            font=("Helvetica", 9, "italic"))
        # Initially hidden, shown only when checkbox is checked and pattern type is not delete
        if both_variants_var.get() and pattern_type_var.get() != "delete":
            variant_warning_label.grid(row=0, column=1, padx=5)
        else:
            variant_warning_label.grid_forget()
    
        # Add save and cancel buttons
        save_button = tk.Button(button_frame, text="Add Pattern to Excel", command=add_pattern_to_excel)
        save_button.pack(side="right", padx=5)
    
        cancel_button = tk.Button(button_frame, text="Cancel", command=add_window.destroy)
        cancel_button.pack(side="right", padx=5)
    
        # Example frame from RTVM Desk Guide
        example_frame = tk.LabelFrame(add_window, text="Examples from RTVM Desk Guide")
        example_frame.pack(fill="x", padx=10, pady=5)
    
        examples_text = (
            "Update Existing: WCC-VERI-DOC-169;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;SAT\n"
            "Add New CDRL: ADD;070-001;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT\n"
            "Delete VeriDoc: DEL;WCC-VERI-DOC-169"
        )
    
        examples_label = tk.Label(example_frame, text=examples_text, justify="left", font=("Courier", 9))
        examples_label.pack(fill="x", padx=5, pady=5)




    def add_dropdown_to_proposed_changes(self):
        """Add a dropdown button above the proposed changes table"""
        # Create a frame for the label and dropdown button
        self.proposed_changes_header_frame = tk.Frame(self.root)
        self.proposed_changes_header_frame.grid(row=4, column=3, sticky="ew")
    
        # Move the label to this frame
        self.proposed_changes_label.grid_forget()  # Remove from its current position
        self.proposed_changes_label = tk.Label(
            self.proposed_changes_header_frame, text="Contractor Proposed Change Request Input")
        self.proposed_changes_label.pack(side="left", padx=5, pady=5)
    
        # Add a dropdown button
        self.proposed_changes_dropdown_button = tk.Button(
            self.proposed_changes_header_frame, 
            text="Actions ▼", 
            command=self.show_proposed_changes_dropdown)
        self.proposed_changes_dropdown_button.pack(side="right", padx=5, pady=5)
    
        # Add RTVM Desk Guide Button
        self.rtvm_guide_button = tk.Button(
            self.proposed_changes_header_frame,
            text="RTVM Guide",
            command=self.show_rtvm_guide)
        self.rtvm_guide_button.pack(side="right", padx=5, pady=5)

    def show_rtvm_guide(self):
        """Display a window with key information from the RTVM Desk Guide"""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("RTVM Desk Guide - Key Points")
        guide_window.geometry("800x600")
        guide_window.transient(self.root)
    
        # Add a notebook for tabs
        notebook = ttk.Notebook(guide_window)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
        # Tab 1: General Guidelines
        general_frame = tk.Frame(notebook)
        notebook.add(general_frame, text="General Guidelines")
    
        general_text = (
            "RTVM Color Coding:\n"
            "• Grey: Read Only Column as Exported from DOORs.\n"
            "• Blue: Contractor Editable Column.\n"
            "• Orange: USCG Editable Column.\n\n"
        
            "Key Guidelines for Pattern Formatting:\n"
            "• NEVER USE SEMICOLONS (;) except as delimiters between elements\n"
            "• Use commas (,) between location elements\n"
            "• Capitalize The First Letter Of Each Word in location\n" 
            "• Avoid abbreviations (use 'Page' not 'PG', 'Sheet' not 'SHT')\n"
            "• For unknown Sheet/Page/Plan elements, use [] as placeholders\n"
            "• UNSAT status allows you to partially complete with known information\n"
            "• The government will NOT adjudicate any VERI-DOCs in UNSAT status\n"
            "• Do not submit with TBD status\n\n"
        
            "Important Notes:\n"
            "• When a new revision of a CDRL is submitted, a change request referencing the new\n"
            "  revision need not be made unless the revision covers an engineering change affecting\n"
            "  how the spec ID in question is met.\n"
            "• Only use one variant to modify the assignment for any VERI-DOC number.\n"
            "  The other variant is accounted for by using the ADD request on its own line.\n"
            "• If there is only one assignment, you MUST add a replacement assignment when\n"
            "  deleting the existing assignment.\n"
            "• When deleting an assignment, you MUST provide remarks in the Contractor Change\n"
            "  Request Comment Column G justifying the change."
        )
    
        general_label = tk.Label(general_frame, text=general_text, justify="left", anchor="nw")
        general_label.pack(fill="both", expand=True, padx=10, pady=10)
    
        # Tab 2: Pattern Examples
        examples_frame = tk.Frame(notebook)
        notebook.add(examples_frame, text="Pattern Examples")
    
        examples_text = (
            "Examples from RTVM Desk Guide:\n\n"
        
            "1. Format for modifying an existing assignment:\n"
            "   WCC-VERI-DOC-10441;;SAT  (Updating Status Only)\n"
            "   WCC-VERI-DOC-169;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT  (Updating Detailed Location)\n\n"
        
            "2. Format for updating due to a subsequent revision of a CDRL:\n"
            "   Original request:\n"
            "   WCC-VERI-DOC-169;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;SAT\n"
            "   ADD;070-001;180_WLR_WCC_070_2_1, Page 3, Plan View 21-C;SAT\n\n"
        
            "   New entry:\n"
            "   WCC-VERI-DOC-169;160_WLIC_WCC_070_2_2, Page 5, Plan View 21-C;SAT\n"
            "   ADD;070-001;180_WLR_WCC_070_2_2, Page 5, Plan View 21-C;SAT\n\n"
        
            "3. Format for adding a new assignment without Detailed Location Information:\n"
            "   ADD;070-001\n\n"
        
            "4. Format for adding a new assignment when Detailed Location Information is Available:\n"
            "   ADD;070-001;160_WLIC_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT\n"
            "   ADD;070-001;180_WLR_WCC_070_2_1, Page 3, Plan View 21-C;UNSAT\n\n"
        
            "5. Format for deleting an existing assignment:\n"
            "   DEL;WCC-VERI-DOC-169"
        )
    
        examples_text_widget = tk.Text(examples_frame, wrap="word", width=80, height=25)
        examples_text_widget.pack(fill="both", expand=True, padx=10, pady=10)
        examples_text_widget.insert("1.0", examples_text)
        examples_text_widget.config(state="disabled")
    
        # Add scrollbar to examples text widget
        examples_scrollbar = tk.Scrollbar(examples_frame, command=examples_text_widget.yview)
        examples_scrollbar.pack(side="right", fill="y")
        examples_text_widget.config(yscrollcommand=examples_scrollbar.set)
    
        # Tab 3: Object Status Definitions
        status_frame = tk.Frame(notebook)
        notebook.add(status_frame, text="Object Status Definitions")
    
        status_text = (
            "Object Status Definitions from RTVM Desk Guide:\n\n"
        
            "PROPOSED: A CDRL that the contractor has requested be added to the SPEC line item awaiting Government review.\n"
            "  • These typically were changes received as a part of the last submittal and will be changed to\n"
            "    Accepted in the next submittal if the government agrees the recommended DI should be added\n"
            "    to the associated SPEC.\n"
            "  • The Government will note its decision in Column J, Government Adjudication Comment History.\n"
            "  • Updated DLOC information can be submitted as Contractor Proposed Change Request in\n"
            "    Column G if appropriate while a VERI-DOC is in this status.\n\n"
        
            "ACCEPTED: A CDRL that has been assigned to the SPEC line item.\n"
            "  • Updated DLOC information can be submitted as Contractor Proposed Change Request in\n"
            "    Column G if appropriate while a VERI-DOC is in this status.\n\n"
        
            "DELETE: A CDRL the contractor has requested be removed from the SPEC line item currently awaiting Government review.\n"
            "  • These typically were changes received as a part of the last submittal.\n"
            "  • If the request is denied by the government, they will return to Accepted status in the next\n"
            "    RTVM output after the adjudication is completed.\n"
            "  • If the government concurs the associated VERI-DOC will either be completely gone from\n"
            "    the next RTVM or placed in Depreciated status.\n"
            "  • Updated DLOCs are NOT expected for any VERI-DOC in this status.\n\n"
        
            "DEPRECIATED: A CDRL that the government concurs with Birdon's request to remove a CDRL from or disagreed with\n"
            "Birdon's request to add a CDRL trace to the SPEC line item requirement and is pending the full removal of the\n"
            "object from DOORs.\n"
            "  • No DLOC edits will be accepted for a VERI-DOC in this status."
        )
    
        status_text_widget = tk.Text(status_frame, wrap="word", width=80, height=25)
        status_text_widget.pack(fill="both", expand=True, padx=10, pady=10)
        status_text_widget.insert("1.0", status_text)
        status_text_widget.config(state="disabled")
    
        # Add scrollbar to status text widget
        status_scrollbar = tk.Scrollbar(status_frame, command=status_text_widget.yview)
        status_scrollbar.pack(side="right", fill="y")
        status_text_widget.config(yscrollcommand=status_scrollbar.set)
    
        # Tab 4: Depreciated Status Warning
        warning_frame = tk.Frame(notebook)
        notebook.add(warning_frame, text="DEPRECIATED Warning")
    
        warning_text = (
            "Important Note about DEPRECIATED Status:\n\n"
        
            "A CDRL that the government concurs should be removed from or disagreed with adding to the SPEC line\n"
            "item requirement and is pending the full removal of the object from DOORs.\n\n"
        
            "No DLOC edits will be accepted for a VERI-DOC in this status.\n\n"
        
            "When you encounter an Object Status of 'DEPRECIATED' in the DI Number Breakdown table,\n"
            "the system will show a warning to remind you that no changes are allowed for this item."
        )
    
        warning_label = tk.Label(warning_frame, text=warning_text, justify="left", anchor="nw", fg="red")
        warning_label.pack(fill="both", expand=True, padx=10, pady=10)

    def show_proposed_changes_dropdown(self):
        """Display dropdown menu for the proposed changes table"""
        dropdown = tk.Menu(self.root, tearoff=0)
    
        # Add options
        dropdown.add_command(label="Add New Pattern", command=self.add_new_pattern)
        dropdown.add_command(label="Bulk Edit Patterns", command=self.bulk_edit_patterns)
        dropdown.add_command(label="Sort Patterns", command=self.sort_patterns)
        dropdown.add_command(label="Clear All Patterns", command=self.clear_all_patterns)
        dropdown.add_command(label="Validate All Patterns", command=self.validate_all_patterns)
    
        # Display the menu under the dropdown button
        x = self.proposed_changes_dropdown_button.winfo_rootx()
        y = self.proposed_changes_dropdown_button.winfo_rooty() + self.proposed_changes_dropdown_button.winfo_height()
        dropdown.post(x, y)

    def validate_all_patterns(self):
        """Validate all patterns according to RTVM Desk Guide rules"""
        # Get current patterns from column G
        content = self.df.iloc[self.current_row, 6]
        if pd.isna(content):
            content = ""
        elif not isinstance(content, str):
            content = str(content)
        
        if not content.strip():
            messagebox.showinfo("No Patterns", "No patterns to validate.")
            return
        
        # Split into lines
        lines = [line.strip() for line in content.strip().split('\n') if line.strip()]
    
        if not lines:
            messagebox.showinfo("No Patterns", "No patterns to validate.")
            return
    
        # Validate each pattern
        errors = []
        warnings = []
        for i, line in enumerate(lines):
            line_num = i + 1
        
            # Check for spaces around semicolons
            if re.search(r"\s+;|;\s+", line):
                errors.append(f"Line {line_num}: Remove spaces before or after semicolons")
        
            # Validate based on pattern type
            if line.startswith("ADD;"):
                parts = line.split(";")
                if len(parts) < 2:
                    errors.append(f"Line {line_num}: ADD pattern must have at least the DI Number (ADD;DI-Number)")
                elif len(parts) >= 4 and parts[3].strip().upper() not in ["SAT", "UNSAT"] and parts[3].strip():
                    errors.append(f"Line {line_num}: Status must be either SAT or UNSAT, not TBD or other values")
                
                # Check for abbreviations
                abbreviations = self.check_for_abbreviations(line)
                if abbreviations:
                    warnings.append(f"Line {line_num}: Found potential abbreviations: {', '.join(abbreviations)}")
                
            elif line.startswith("DEL;"):
                parts = line.split(";")
                if len(parts) != 2 or not parts[1]:
                    errors.append(f"Line {line_num}: DEL pattern must have format: DEL;WCC-VERI-DOC-XXXX")
                
            elif ";" in line:  # Update pattern
                parts = line.split(";")
                if not parts[0].startswith("WCC-VERI-DOC-"):
                    warnings.append(f"Line {line_num}: VeriDoc number should start with WCC-VERI-DOC-")
                
                if len(parts) >= 3 and parts[2].strip().upper() not in ["SAT", "UNSAT"] and parts[2].strip():
                    errors.append(f"Line {line_num}: Status must be either SAT or UNSAT, not TBD or other values")
                
                # Check for abbreviations
                abbreviations = self.check_for_abbreviations(line)
                if abbreviations:
                    warnings.append(f"Line {line_num}: Found potential abbreviations: {', '.join(abbreviations)}")
            else:
                warnings.append(f"Line {line_num}: Unknown pattern format. Should be ADD;, DEL;, or VeriDoc;")
    
        # Show results
        if not errors and not warnings:
            messagebox.showinfo("Validation Passed", "All patterns are valid according to RTVM Desk Guide rules.")
        else:
            result_window = tk.Toplevel(self.root)
            result_window.title("Validation Results")
            result_window.geometry("600x400")
            result_window.transient(self.root)
        
            if errors:
                error_frame = tk.LabelFrame(result_window, text="Errors (Must Fix)")
                error_frame.pack(fill="x", padx=10, pady=5)
            
                error_text = tk.Text(error_frame, height=len(errors) + 1, width=70, bg="pink")
                error_text.pack(fill="x", padx=5, pady=5)
                error_text.insert("1.0", "\n".join(errors))
                error_text.config(state="disabled")
        
            if warnings:
                warning_frame = tk.LabelFrame(result_window, text="Warnings (Should Fix)")
                warning_frame.pack(fill="x", padx=10, pady=5)
            
                warning_text = tk.Text(warning_frame, height=len(warnings) + 1, width=70, bg="lightyellow")
                warning_text.pack(fill="x", padx=5, pady=5)
                warning_text.insert("1.0", "\n".join(warnings))
                warning_text.config(state="disabled")
        
            # Button to open bulk edit
            if errors or warnings:
                button_frame = tk.Frame(result_window)
                button_frame.pack(fill="x", padx=10, pady=10)
            
                edit_button = tk.Button(button_frame, text="Open Bulk Edit to Fix Issues", 
                                       command=lambda: [result_window.destroy(), self.bulk_edit_patterns()])
                edit_button.pack(side="left", padx=5)
            
                close_button = tk.Button(button_frame, text="Close", command=result_window.destroy)
                close_button.pack(side="right", padx=5)

    def check_for_abbreviations(self, text):
        """Check for common abbreviations in a pattern"""
        common_abbrs = {
            "PG": "Page",
            "SHT": "Sheet",
            "SEC": "Section",
            "PARA": "Paragraph"
        }
    
        found_abbrs = []
        for abbr in common_abbrs:
            if re.search(r"\b" + abbr + r"\b", text, re.IGNORECASE):
                found_abbrs.append(f"{abbr} (use {common_abbrs[abbr]})")
    
        return found_abbrs

    def bulk_edit_patterns(self):
        """Open a window for bulk editing all patterns"""
        # Get current patterns
        content = self.df.iloc[self.current_row, 6]
        if pd.isna(content):
            content = ""
        elif not isinstance(content, str):
            content = str(content)
        
        # Create a new window for bulk editing
        bulk_window = tk.Toplevel(self.root)
        bulk_window.title("Bulk Edit Patterns")
        bulk_window.geometry("800x600")
        bulk_window.transient(self.root)
        bulk_window.grab_set()  # Make the window modal
    
        # Add guidelines frame
        guidelines_frame = tk.LabelFrame(bulk_window, text="RTVM Formatting Guidelines")
        guidelines_frame.pack(fill="x", padx=10, pady=5)
    
        guidelines_text = (
            "• NEVER USE SEMICOLONS (;) except as delimiters between elements\n"
            "• Use commas (,) between location elements\n"
            "• Capitalize The First Letter Of Each Word in location\n" 
            "• Avoid abbreviations (use 'Page' not 'PG', 'Sheet' not 'SHT')\n"
            "• For unknown Sheet/Page/Plan elements, use [] as placeholders\n"
            "• Format: ADD;DI-Number;CDRL Name, Page X, Plan View Y;Status\n"
            "• Valid statuses: SAT, UNSAT (don't use TBD for submissions)\n"
            "• Use DEL;WCC-VERI-DOC-XXXX to request deletion"
        )
    
        guidelines_label = tk.Label(guidelines_frame, text=guidelines_text, justify="left", anchor="w")
        guidelines_label.pack(fill="x", padx=5, pady=5)
    
        # Toggle button to show/hide guidelines
        guidelines_visible = tk.BooleanVar(value=True)
    
        def toggle_guidelines():
            if guidelines_visible.get():
                guidelines_label.pack_forget()
                guidelines_visible.set(False)
                toggle_button.config(text="Show Guidelines")
            else:
                guidelines_label.pack(fill="x", padx=5, pady=5)
                guidelines_visible.set(True)
                toggle_button.config(text="Hide Guidelines")
    
        toggle_button = tk.Button(guidelines_frame, text="Hide Guidelines", command=toggle_guidelines)
        toggle_button.pack(side="right", padx=5, pady=2)
    
        # Add instructions
        instruction_label = tk.Label(bulk_window, text="Edit all patterns below (one pattern per line):", anchor="w")
        instruction_label.pack(fill="x", padx=10, pady=(10, 5))
    
        # Create text widget with scrollbar and line numbers
        edit_frame = tk.Frame(bulk_window)
        edit_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
        # Line numbers
        line_numbers = tk.Text(edit_frame, width=4, padx=3, pady=5, takefocus=0, 
                              border=0, background='lightgrey', state='disabled')
        line_numbers.pack(side="left", fill="y")
    
        # Main text widget
        edit_text = tk.Text(edit_frame, wrap="none", padx=5, pady=5, font=("Courier", 10))
        edit_text.pack(side="left", fill="both", expand=True)
    
        # Set up scrollbars
        y_scrollbar = tk.Scrollbar(edit_frame, command=edit_text.yview)
        y_scrollbar.pack(side="right", fill="y")
    
        x_scrollbar = tk.Scrollbar(bulk_window, orient="horizontal", command=edit_text.xview)
        x_scrollbar.pack(fill="x", padx=10)
    
        edit_text.config(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
    
        # Configure line numbers text widget
        line_numbers.config(yscrollcommand=y_scrollbar.set)
    
        # Insert content into the text widget
        edit_text.insert("1.0", content)
    
        # Function to update line numbers
        def update_line_numbers(*args):
            line_numbers.config(state='normal')
            line_numbers.delete('1.0', 'end')
            line_count = edit_text.get('1.0', 'end').count('\n')
            for i in range(1, line_count + 1):
                line_numbers.insert('end', f"{i}\n")
            line_numbers.config(state='disabled')
    
        # Bind to text changes to update line numbers
        edit_text.bind('<KeyRelease>', update_line_numbers)
        edit_text.bind('<FocusIn>', update_line_numbers)
    
        # Initial update of line numbers
        update_line_numbers()
    
        # Helper buttons frame
        helper_frame = tk.Frame(bulk_window)
        helper_frame.pack(fill="x", padx=10, pady=5)
    
        # Fix common issues button
        def fix_common_issues():
            text = edit_text.get("1.0", "end-1c")
        
            # 1. Remove extra spaces around semicolons
            text = re.sub(r"\s+;", ";", text)
            text = re.sub(r";\s+", ";", text)
        
            # 2. Fix common abbreviations
            replacements = {
                "PG ": "Page ",
                "SHT ": "Sheet ",
                "SEC ": "Section ",
                "PARA ": "Paragraph "
            }
        
            for abbr, full in replacements.items():
                text = text.replace(abbr, full)
                text = text.replace(abbr.lower(), full)
        
            # 3. Capitalize first letter of each word in location parts
            lines = text.split('\n')
            for i, line in enumerate(lines):
                parts = line.split(';')
                if len(parts) >= 3:
                    # For ADD pattern, location is index 2
                    if parts[0] == "ADD" and len(parts) >= 3:
                        location_idx = 2
                    # For Update pattern, location is index 1
                    elif not parts[0].startswith("DEL") and len(parts) >= 2:
                        location_idx = 1
                    else:
                        continue
                
                    # Split location by commas
                    location_parts = parts[location_idx].split(',')
                
                    # Process each part
                    for j, part in enumerate(location_parts):
                        # Skip parts with underscores (like variant identifiers)
                        if "_" in part:
                            continue
                    
                        # Capitalize first letter of each word
                        words = part.strip().split()
                        capitalized = " ".join(word.capitalize() for word in words)
                        location_parts[j] = capitalized
                
                    # Rejoin the location
                    parts[location_idx] = ", ".join(location_parts)
                
                    # Rejoin the pattern
                    lines[i] = ";".join(parts)
        
            # Update the text widget
            edit_text.delete("1.0", "end")
            edit_text.insert("1.0", '\n'.join(lines))
            update_line_numbers()
    
        fix_button = tk.Button(helper_frame, text="Fix Common Issues", command=fix_common_issues)
        fix_button.pack(side="left", padx=5)
    
        # Run validation button
        def run_validation():
            text = edit_text.get("1.0", "end-1c")
            lines = [line.strip() for line in text.split('\n') if line.strip()]
        
            # Clear all tags
            for tag in edit_text.tag_names():
                if tag not in ["sel", "current"]:  # Don't remove selection tags
                    edit_text.tag_remove(tag, "1.0", "end")
        
            # Validate each line
            errors_found = False
            for i, line in enumerate(lines):
                line_num = i + 1
                line_start = f"{line_num}.0"
                line_end = f"{line_num}.end"
            
                # Apply syntax highlighting
                if line.startswith("ADD;"):
                    edit_text.tag_add("add", line_start, f"{line_num}.4")
                elif line.startswith("DEL;"):
                    edit_text.tag_add("del", line_start, f"{line_num}.4")
            
                # Check for common errors
                if re.search(r"\s+;|;\s+", line):
                    edit_text.tag_add("error", line_start, line_end)
                    errors_found = True
            
                # Check for abbreviations
                for abbr in ["PG", "SHT", "SEC", "PARA"]:
                    for match in re.finditer(r"\b" + abbr + r"\b", line, re.IGNORECASE):
                        start, end = match.span()
                        edit_text.tag_add("warning", f"{line_num}.{start}", f"{line_num}.{end}")
                        errors_found = True
            
                # Check status values
                parts = line.split(";")
                if line.startswith("ADD;") and len(parts) >= 4:
                    if parts[3].strip().upper() not in ["SAT", "UNSAT"] and parts[3].strip():
                        status_idx = line.rfind(";") + 1
                        edit_text.tag_add("error", f"{line_num}.{status_idx}", line_end)
                        errors_found = True
                elif not line.startswith("DEL;") and ";" in line and len(parts) >= 3:
                    if parts[2].strip().upper() not in ["SAT", "UNSAT"] and parts[2].strip():
                        status_idx = line.rfind(";") + 1
                        edit_text.tag_add("error", f"{line_num}.{status_idx}", line_end)
                        errors_found = True
        
            # Configure tags
            edit_text.tag_configure("add", foreground="blue", font=("Courier", 10, "bold"))
            edit_text.tag_configure("del", foreground="red", font=("Courier", 10, "bold"))
            edit_text.tag_configure("error", background="pink", underline=True)
            edit_text.tag_configure("warning", background="lightyellow")
        
            if not errors_found:
                messagebox.showinfo("Validation", "No errors found.")
            else:
                messagebox.showwarning("Validation", 
                                      "Errors or warnings found. Red underlines indicate errors, yellow highlights indicate warnings.")
    
        validate_button = tk.Button(helper_frame, text="Validate", command=run_validation)
        validate_button.pack(side="left", padx=5)
    
        # Button frame
        button_frame = tk.Frame(bulk_window)
        button_frame.pack(fill="x", padx=10, pady=10)
    
        # Function to save all patterns
        def save_all_patterns():
            new_content = edit_text.get("1.0", "end-1c").strip()
        
            # Update the Excel file
            try:
                # Load the workbook
                wb = load_workbook(self.excel_file_path)
                ws = wb.active
            
                # Calculate the Excel row number
                excel_row = self.current_row + 2
            
                # Update the cell in column G
                ws.cell(row=excel_row, column=7, value=new_content)
            
                # Save the workbook
                wb.save(self.excel_file_path)
            
                # Update the DataFrame in memory
                self.df.iloc[self.current_row, 6] = new_content
            
                messagebox.showinfo("Success", "Patterns updated and saved to Excel file successfully.")
            
                # Update the table
                self.update_proposed_changes_table()
            
                # Close the window
                bulk_window.destroy()
            
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save to Excel file: {e}")
    
        # Add save and cancel buttons
        save_button = tk.Button(button_frame, text="Save All Changes", command=save_all_patterns)
        save_button.pack(side="left", padx=5)
    
        cancel_button = tk.Button(button_frame, text="Cancel", command=bulk_window.destroy)
        cancel_button.pack(side="left", padx=5)

    def sort_patterns(self):
        # Implementation shown in the previous code block...
        # This will sort patterns by type (ADD, DEL, VeriDoc) or by DI Number, etc.
        pass

    def clear_all_patterns(self):
        # Implementation shown in the previous code block...
        # This will clear all patterns after confirmation
        pass


    def delete_proposed_change_row(self, row_id):
        # Get the pattern from the selected row
        item = self.proposed_changes_table.item(row_id)
        pattern_to_delete = item['values'][0]
        if pattern_to_delete:
            # Remove the pattern from the DataFrame
            content = self.df.iloc[self.current_row, 6]
            if pd.isna(content):
                content = ""
            elif not isinstance(content, str):
                content = str(content)

            # Split the content into lines
            lines = content.strip().split('\n')
            if pattern_to_delete in lines:
                lines.remove(pattern_to_delete)
                # Join the remaining lines
                new_content = '\n'.join(lines)

                # Update the Excel file directly using openpyxl
                from openpyxl import load_workbook

                try:
                    # Load the workbook
                    wb = load_workbook(self.excel_file_path)
                    ws = wb.active  # You may need to select the correct sheet if there are multiple

                    # Calculate the Excel row number (considering headers)
                    excel_row = self.current_row + 2  # Assuming header is on the first row

                    # Update the cell in column G (which is column index 7 in openpyxl)
                    ws.cell(row=excel_row, column=7, value=new_content)

                    # Save the workbook
                    wb.save(self.excel_file_path)

                    # Update the DataFrame in memory
                    self.df.iloc[self.current_row, 6] = new_content

                    messagebox.showinfo(
                        "Success", "Row deleted and changes saved to Excel file successfully.")
                    # Update the Proposed Changes Table
                    self.update_proposed_changes_table()
                except Exception as e:
                    messagebox.showerror(
                        "Error", f"Failed to save to Excel file: {e}")
            else:
                messagebox.showerror(
                    "Error", "Pattern not found in the Excel data.")
        else:
            messagebox.showerror(
                "Error", "No pattern found in the selected row.")


    def update_ui_after_navigation(self):
        # First check if DataFrame exists and has rows
        if self.df is None or len(self.df) == 0:
            # Clear UI elements for empty file
            self.spec_text_box.config(state=tk.NORMAL)
            self.spec_text_box.delete("1.0", tk.END)
            self.spec_text_box.insert(tk.END, "No data available in this file.")
            self.spec_text_box.config(state=tk.DISABLED)
        
            # Clear all tables
            self.table.delete(*self.table.get_children())
            self.comment_table.delete(*self.comment_table.get_children())
            self.proposed_changes_table.delete(*self.proposed_changes_table.get_children())
            self.comment_history_table.delete(*self.comment_history_table.get_children())
            self.gov_comment_history_table.delete(*self.gov_comment_history_table.get_children())
        
            messagebox.showinfo("Empty File", "This file contains no data rows to navigate.")
            return

        # Check if current_row is valid (within bounds)
        if self.current_row < 0:
            self.current_row = 0
        elif self.current_row >= len(self.df):
            self.current_row = max(0, len(self.df) - 1)

        # Now proceed with the original function if we have valid data
        try:
            # Get data from column F (index 5)
            data = self.df.iloc[self.current_row, 5]
            # Update Specification Text Box with content from column B (index 1)
            spec_text = self.df.iloc[self.current_row, 1]
            if pd.isna(spec_text):
                spec_text = ""
            elif not isinstance(spec_text, str):
                spec_text = str(spec_text)
            self.spec_text_box.config(state=tk.NORMAL)
            self.spec_text_box.delete("1.0", tk.END)
            self.spec_text_box.insert(tk.END, spec_text)
            self.spec_text_box.config(state=tk.DISABLED)
            # Check if data is NaN or not a string
            if pd.isna(data):
                data = ""
            elif not isinstance(data, str):
                data = str(data)
            # Run extract_info automatically
            self.extract_info(data)
            # Update the Proposed Changes Table
            self.update_proposed_changes_table()
            # **Update the Comment History Table**
            self.update_comment_history_table()
            # Update the progress bar highlight
            self.update_progress_bar_highlight()
        except (IndexError, KeyError) as e:
            # Handle case where column structure doesn't match expectations
            self.spec_text_box.config(state=tk.NORMAL)
            self.spec_text_box.delete("1.0", tk.END)
            self.spec_text_box.insert(tk.END, f"Error accessing data: {str(e)}\nThe file structure may not be compatible.")
            self.spec_text_box.config(state=tk.DISABLED)
        
            # Clear all tables
            self.table.delete(*self.table.get_children())
            self.comment_table.delete(*self.comment_table.get_children())
            self.proposed_changes_table.delete(*self.proposed_changes_table.get_children())
            self.comment_history_table.delete(*self.comment_history_table.get_children())
            self.gov_comment_history_table.delete(*self.gov_comment_history_table.get_children())

    def update_comment_history_table(self):
        # Clear the table
        self.comment_history_table.delete(
            *self.comment_history_table.get_children())

        # Get the content from column I (index 8)
        content = self.df.iloc[self.current_row, 8]
        if pd.isna(content):
            content = ""
        elif not isinstance(content, str):
            content = str(content)

        # Split the content into lines and populate the table
        lines = content.strip().split('\n')
        for line in lines:
            if line.strip():
                self.comment_history_table.insert(
                    "", "end", values=(line.strip(),))


    def open_pie_chart_window(self):
        if self.df is None:
            messagebox.showerror("Error", "Please upload an Excel file first.")
            return

        self.pie_chart_window = tk.Toplevel(self.root)
        self.pie_chart_window.title("Pie Charts")

        # Create filter frame (this stays the same)
        filter_frame = tk.Frame(self.pie_chart_window)
        filter_frame.pack(side=tk.TOP, fill=tk.X)

        # Object Status Filter
        tk.Label(filter_frame, text="Object Status").pack(side=tk.LEFT, padx=5, pady=5)
        self.pie_object_status_var = tk.StringVar()
        self.pie_object_status_dropdown = ttk.Combobox(
            filter_frame,
            textvariable=self.pie_object_status_var,
            values=["Any"] + sorted(self.unique_object_statuses),
            state="readonly",
            width=15
        )
        self.pie_object_status_dropdown.pack(side=tk.LEFT)
        self.pie_object_status_var.set("Any")

        # Contractor Assessed Status Filter
        tk.Label(filter_frame, text="Contractor Assessed Status").pack(side=tk.LEFT, padx=5, pady=5)
        self.pie_contractor_status_var = tk.StringVar()
        self.pie_contractor_status_dropdown = ttk.Combobox(
            filter_frame,
            textvariable=self.pie_contractor_status_var,
            values=["Any"] + sorted(self.unique_contractor_statuses),
            state="readonly",
            width=15
        )
        self.pie_contractor_status_dropdown.pack(side=tk.LEFT)
        self.pie_contractor_status_var.set("Any")

        # Government Assessed Status Filter
        tk.Label(filter_frame, text="Government Assessed Status").pack(side=tk.LEFT, padx=5, pady=5)
        self.pie_government_status_var = tk.StringVar()
        self.pie_government_status_dropdown = ttk.Combobox(
            filter_frame,
            textvariable=self.pie_government_status_var,
            values=["Any"] + sorted(self.unique_government_statuses),
            state="readonly",
            width=15
        )
        self.pie_government_status_dropdown.pack(side=tk.LEFT)
        self.pie_government_status_var.set("Any")

        # Update Charts Button
        tk.Button(filter_frame, text="Update Charts", command=self.update_pie_charts).pack(side=tk.LEFT, padx=10)

        # Now create a frame to hold four charts in a 2×2 grid
        charts_frame = tk.Frame(self.pie_chart_window)
        charts_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Create the four figures and embed them using grid.
        self.object_status_fig = plt.Figure(figsize=(4, 4), dpi=100)
        self.object_status_canvas = FigureCanvasTkAgg(self.object_status_fig, charts_frame)
        self.object_status_canvas.get_tk_widget().grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        self.contractor_status_fig = plt.Figure(figsize=(4, 4), dpi=100)
        self.contractor_status_canvas = FigureCanvasTkAgg(self.contractor_status_fig, charts_frame)
        self.contractor_status_canvas.get_tk_widget().grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

        self.government_status_fig = plt.Figure(figsize=(4, 4), dpi=100)
        self.government_status_canvas = FigureCanvasTkAgg(self.government_status_fig, charts_frame)
        self.government_status_canvas.get_tk_widget().grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.gov_req_status_fig = plt.Figure(figsize=(4, 4), dpi=100)
        self.gov_req_status_canvas = FigureCanvasTkAgg(self.gov_req_status_fig, charts_frame)
        self.gov_req_status_canvas.get_tk_widget().grid(row=1, column=1, padx=5, pady=5, sticky="nsew")

        # Optionally configure the grid so cells expand equally.
        charts_frame.columnconfigure(0, weight=1)
        charts_frame.columnconfigure(1, weight=1)
        charts_frame.rowconfigure(0, weight=1)
        charts_frame.rowconfigure(1, weight=1)

        # Create a frame for the counts table (same as before)
        counts_frame = tk.Frame(self.pie_chart_window)
        counts_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.counts_table = ttk.Treeview(counts_frame, columns=("Status", "Count", "Total Count"), show="headings")
        self.counts_table.heading("Status", text="Status")
        self.counts_table.heading("Count", text="Count")
        self.counts_table.heading("Total Count", text="Total Count")
        self.counts_table.pack(fill=tk.BOTH, expand=True)

        self.compute_total_counts()
        self.update_pie_charts()

    def update_pie_charts(self):
        # Get the selected filters
        selected_object_status = self.pie_object_status_var.get()
        selected_contractor_status = self.pie_contractor_status_var.get()
        selected_government_status = self.pie_government_status_var.get()

        if selected_object_status == "Any":
            selected_object_status = ''
        if selected_contractor_status == "Any":
            selected_contractor_status = ''
        if selected_government_status == "Any":
            selected_government_status = ''

        # Filter the status data based on the selected filters (for the first three charts)
        filtered_items = []
        for item in self.status_data:
            object_match = True
            contractor_match = True
            government_match = True
            if selected_object_status:
                object_match = (item['object_status'] == selected_object_status)
            if selected_contractor_status:
                contractor_match = (item['contractor_status'] == selected_contractor_status)
            if selected_government_status:
                government_match = (item['government_status'] == selected_government_status)
            if object_match and contractor_match and government_match:
                filtered_items.append(item)

        filtered_statuses = {
            'object_status': [item['object_status'] for item in filtered_items],
            'contractor_status': [item['contractor_status'] for item in filtered_items],
            'government_status': [item['government_status'] for item in filtered_items]
        }

        # Plot the first three charts using the filtered data
        self.plot_pie_chart(self.object_status_fig, filtered_statuses['object_status'], 'Object Status')
        self.plot_pie_chart(self.contractor_status_fig, filtered_statuses['contractor_status'], 'Contractor Assessed Status')
        self.plot_pie_chart(self.government_status_fig, filtered_statuses['government_status'], 'Government Assessed Status')

        # --- Now compute the per-requirement overall government status ---
        # We will group by row index (each row is one requirement).
        from collections import defaultdict, Counter
        req_status = defaultdict(list)
        for item in self.status_data:
            row_idx = item['row_index']
            status = item['government_status']
            if isinstance(status, str):
                req_status[row_idx].append(status.strip().lower())

        # Define the order of precedence (lower number = higher precedence)
        precedence = {"agree": 1, "disagree": 2, "pending review": 3, "awaiting input": 4}

        overall_status_list = []
        for row_idx, statuses in req_status.items():
            overall = None
            min_rank = float('inf')
            for s in statuses:
                rank = precedence.get(s, 100)  # default to a high rank if unknown
                if rank < min_rank:
                    min_rank = rank
                    overall = s
            if overall:
                # Capitalize the first letter (so "agree" becomes "Agree", etc.)
                overall_status_list.append(overall.capitalize())

        # Plot the new pie chart using the overall status per requirement.
        self.plot_pie_chart(self.gov_req_status_fig, overall_status_list, 'Government Assessed Status Per Requirement')

        # Redraw all canvases
        self.object_status_canvas.draw()
        self.contractor_status_canvas.draw()
        self.government_status_canvas.draw()
        self.gov_req_status_canvas.draw()

        # Update counts table (using filtered_statuses as before)
        self.update_counts_table(filtered_statuses)


    def plot_pie_chart(self, fig, data_list, title):
        # Clear the figure
        fig.clear()
        from collections import Counter
        counter = Counter(data_list)
        labels = list(counter.keys())
        sizes = list(counter.values())

        # Define your desired color mapping:
        status_color_map = {
            'Agree': 'green',
            'Disagree': 'red',
            'Awaiting input': 'yellow',
            'Pending eview': 'blue'
        }

        # We standardize each label by capitalizing it before checking the mapping.
        colors = []
        for label in labels:
            label_cap = label.capitalize()
            colors.append(status_color_map.get(label_cap, 'grey'))



        # Handle the case of no data.
        if not sizes:
            labels = ['No Data']
            sizes = [1]
            colors = ['lightgrey']

        # Create the pie chart with the specified colors.
        ax = fig.add_subplot(111)
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors)
        ax.axis('equal')
        ax.set_title(title)

    def update_counts_table(self, filtered_statuses):
        # Clear the table
        self.counts_table.delete(*self.counts_table.get_children())

        # Prepare data for the table
        status_types = {
            'Object Status': 'object_status',
            'Contractor Assessed Status': 'contractor_status',
            'Government Assessed Status': 'government_status'
        }

        for status_type, key in status_types.items():
            # Insert a heading row for the status type
            self.counts_table.insert("", "end", values=(status_type, "", ""), tags=('heading',))
            # Count the occurrences in filtered data
            filtered_counter = Counter(filtered_statuses[key])
            # Total counts from the entire dataset
            total_counter = self.total_counts[key]

            # Get all unique statuses to ensure alignment
            all_statuses = set(filtered_counter.keys()).union(set(total_counter.keys()))

            for status in all_statuses:
                filtered_count = filtered_counter.get(status, 0)
                total_count = total_counter.get(status, 0)
                self.counts_table.insert("", "end", values=(status, filtered_count, total_count))
            # Insert an empty row for spacing
            self.counts_table.insert("", "end", values=("", "", ""))
        # Configure the heading row style
        self.counts_table.tag_configure('heading', font=('Arial', 10, 'bold'))


    def compute_total_counts(self):
        # Initialize dictionaries to store total counts
        self.total_counts = {
            'object_status': {},
            'contractor_status': {},
            'government_status': {}
        }

        # Collect all statuses from the entire dataset
        total_statuses = {
            'object_status': [],
            'contractor_status': [],
            'government_status': []
        }
        for item in self.status_data:
            total_statuses['object_status'].append(item['object_status'])
            total_statuses['contractor_status'].append(item['contractor_status'])
            total_statuses['government_status'].append(item['government_status'])

        # Count occurrences for each status type
        for key in self.total_counts.keys():
            counter = Counter(total_statuses[key])
            self.total_counts[key] = counter

    def save_comments_to_excel(self):
        # Prepare the content for the 'Contractor Proposed Change Comment Input' column
        comment_lines = []
        for obj_id, comment_text in self.current_comments.items():
            comment_line = f"{obj_id} - {comment_text}"
            comment_lines.append(comment_line)
        comments_content = '\n'.join(comment_lines)

        # Use openpyxl to write to column H (index 8)
        from openpyxl import load_workbook

        try:
            # Load the workbook
            wb = load_workbook(self.excel_file_path)
            ws = wb.active  # Or specify the sheet if needed

            # Calculate the Excel row number (considering headers)
            excel_row = self.current_row + 2  # Adjust if your header is on a different row

            # Find the column index for 'Contractor Proposed Change Comment Input'
            column_letter = None
            for col in ws.iter_cols(1, ws.max_column):
                if col[0].value == 'Contractor Proposed Change Comment Input':
                    column_letter = col[0].column_letter
                    break

            if column_letter is None:
                messagebox.showerror("Error", "Column 'Contractor Proposed Change Comment Input' not found in Excel file.")
                return

            # Update the cell in the correct column
            ws[f"{column_letter}{excel_row}"] = comments_content

            # Save the workbook
            wb.save(self.excel_file_path)

            # Update the DataFrame in memory
            self.df.at[self.current_row, 'Contractor Proposed Change Comment Input'] = comments_content

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save comments to Excel file: {e}")


    def update_gov_comment_history_table(self):
        print("Updating Government Adjudication Comment History Table")
        # Clear the table
        self.gov_comment_history_table.delete(*self.gov_comment_history_table.get_children())

        # Get the content from column J (index 9)
        try:
            content = self.df.iloc[self.current_row, 9]
        except IndexError:
            messagebox.showerror("Error", "Column J not found in Excel file.")
            return

        if pd.isna(content):
            content = ""
        elif not isinstance(content, str):
            content = str(content)

        # Debugging output
        print(f"Content from column J at row {self.current_row}: {content}")

        # Split the content into lines and populate the table
        lines = content.strip().split('\n')
        for line in lines:
            if line.strip():
                self.gov_comment_history_table.insert("", "end", values=(line.strip(),))



    def on_comment_double_click(self, event):
        # Disable double-click editing
        pass  # Do nothing on double-click

    def show_comment_table_context_menu(self, event):
        # Identify the row under the cursor
        row_id = self.comment_table.identify_row(event.y)
        if row_id:
            # Create a context menu
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(
                label="Edit Comment", command=lambda: self.open_comment_editor(row_id))
            menu.post(event.x_root, event.y_root)

    def open_comment_editor(self, row_id):
        # Get the current comment text
        item = self.comment_table.item(row_id)
        current_text = item['values'][0] if item['values'] else ''

        # Create a new window
        comment_window = tk.Toplevel(self.root)
        comment_window.title("Edit Comment")

        # Initialize the spell checker
        spell = SpellChecker()

        # Large text box
        text_box = tk.Text(comment_window, width=60, height=20)
        text_box.pack(padx=10, pady=10)
        text_box.insert(tk.END, current_text)

        # Bind the key release event for automatic spell checking
        text_box.bind("<KeyRelease>", lambda event: self.check_spelling(text_box, spell))

        # Configure the tag for misspelled words
        text_box.tag_config("misspelled", underline=True, foreground="red")

        # Save Comment button
        save_button = tk.Button(comment_window, text="Save Comment",
                                command=lambda: self.save_comment(row_id, text_box.get("1.0", tk.END), comment_window))
        save_button.pack(pady=10)

    def check_spelling(self, text_widget, spell):
        # Get the content of the text widget
        content = text_widget.get("1.0", tk.END)

        # Split the content into words with positions
        words = content.split()
        index = "1.0"

        # Clear previous tags
        text_widget.tag_remove("misspelled", "1.0", tk.END)

        for word in words:
            # Clean the word by removing punctuation
            clean_word = ''.join(char for char in word if char.isalpha())

            # Get the start and end index of the word
            start_index = text_widget.search(word, index, stopindex=tk.END)
            if not start_index:
                continue
            end_index = f"{start_index}+{len(word)}c"

            # Update the index for the next search
            index = end_index

            if clean_word.lower() not in spell:
                # Highlight the misspelled word
                text_widget.tag_add("misspelled", start_index, end_index)

    def on_right_click(self, event, text_widget, spell):
        # Get the index of the mouse click
        index = text_widget.index(f"@{event.x},{event.y}")

        # Get the word at that index
        word_start = text_widget.search(r'\m', index, backwards=True, regexp=True)
        word_end = text_widget.search(r'\M', index, forwards=True, regexp=True)
        word = text_widget.get(word_start, word_end)

        # Clean the word
        clean_word = ''.join(char for char in word if char.isalpha())

        # Check if the word is misspelled
        if clean_word.lower() not in spell:
            # Get suggestions
            suggestions = spell.candidates(clean_word.lower())

            # Create a context menu
            menu = tk.Menu(self.root, tearoff=0)
            for suggestion in suggestions:
                menu.add_command(label=suggestion, command=lambda s=suggestion: self.replace_word(text_widget, word_start, word_end, s))
            menu.post(event.x_root, event.y_root)

    def save_comment(self, row_id, new_text, window):
        # Update the comment in the comment_table
        self.comment_table.item(row_id, values=(new_text.strip(),))

        # Close the comment window
        window.destroy()

        # Update self.current_comments
        # Get the index of the item in comment_table
        item_index = self.comment_table.get_children().index(row_id)
        # Get the corresponding item in self.table
        table_items = self.table.get_children()
        if item_index < len(table_items):
            table_row_id = table_items[item_index]
            table_values = self.table.item(table_row_id, 'values')
            obj_id = table_values[0]  # VeriDoc Number
            # Extract the comment text from new_text
            if ' - ' in new_text:
                _, comment_text = new_text.split(' - ', 1)
                self.current_comments[obj_id] = comment_text.strip()
            else:
                # No ' - ', treat the entire text as comment
                self.current_comments[obj_id] = new_text.strip()

        # Save comments to Excel file immediately
        self.save_comments_to_excel()
        

import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from textwrap import wrap
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from datetime import datetime
from textwrap import wrap
import queue
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

class DisagreementManager:
    def __init__(self, master, app, disagreement_items):
        """
        Initialize the Disagreement Manager with updated column headers.
        """
        self.master = master
        self.app = app
        self.disagreement_items = disagreement_items  # Already filtered
        self.output_folder = None   # To be set via "Select Database Location"
        self.report_list = []       # To store info about each generated report

        self.master.title("Disagreement Manager - Batch Reports")
        self.master.geometry("800x600")
        
        # Top frame with buttons
        top_frame = tk.Frame(self.master)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        self.db_button = tk.Button(top_frame, text="Select Database Location", command=self.select_output_folder)
        self.db_button.pack(side=tk.LEFT, padx=5)
        
        self.create_reports_button = tk.Button(top_frame, text="B.1 Create Disagreement Reports", command=self.create_all_reports)
        self.create_reports_button.pack(side=tk.LEFT, padx=5)
        
        self.refresh_button = tk.Button(top_frame, text="Refresh Reports", command=self.refresh_report_table)
        self.refresh_button.pack(side=tk.LEFT, padx=5)
        
        # Table frame
        table_frame = tk.Frame(self.master)
        table_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Updated treeview with more meaningful column headers
        self.report_tree = ttk.Treeview(table_frame, 
                                        columns=("SpecID", "Button1", "Button2", "Button3", "Report_File"),
                                        show="headings")
        self.report_tree.heading("SpecID", text="Spec ID")
        self.report_tree.heading("Button1", text="Not Clear")
        self.report_tree.heading("Button2", text="Can Resolve")
        self.report_tree.heading("Button3", text="Cannot Resolve")
        self.report_tree.heading("Report_File", text="Report File")
        
        self.report_tree.column("SpecID", width=100)
        self.report_tree.column("Button1", width=100)
        self.report_tree.column("Button2", width=100)
        self.report_tree.column("Button3", width=100)
        self.report_tree.column("Report_File", width=300)
        
        self.report_tree.bind("<Double-1>", self.on_report_tree_double_click)
        self.report_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.suppress_depreciated_warning = True


    def read_pdf_fields(self, report):
        """
        Read fields from a single PDF report and handle potential errors
        """
        try:
            pdf_path = report["Report_File"]
            print(f"Reading fields from: {os.path.basename(pdf_path)}")
        
            fields = get_pdf_form_fields(pdf_path)
        
            # Get checkbox states
            not_clear = "✓" if fields.get("disagreementNotClear", False) else ""
            can_resolve = "✓" if fields.get("disagreementResolvedLocations", False) else ""
            cannot_resolve = "✓" if fields.get("disagreementNotResolved", False) else ""
        
            # Log the fields for debugging
            print(f"Fields for {os.path.basename(pdf_path)}: not_clear={not_clear}, can_resolve={can_resolve}, cannot_resolve={cannot_resolve}")
        
            return (report["SpecID"], not_clear, can_resolve, cannot_resolve, pdf_path)
        except Exception as e:
            print(f"Error reading PDF fields from {report['Report_File']}: {str(e)}")
            # Return a default result rather than failing
            return (report["SpecID"], "", "", "", report["Report_File"])


    def select_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder for PDFs")
        if folder:
            self.output_folder = folder
            messagebox.showinfo("Folder Selected", f"PDFs will be saved to: {folder}")

    def prev_item(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.show_item()

    def next_item(self):
        if self.current_index < len(self.disagreement_items) - 1:
            self.current_index += 1
            self.show_item()

    def show_item(self):
        # Clear the preview
        self.preview_text.delete("1.0", tk.END)

        # Get current row_index
        item = self.disagreement_items[self.current_index]
        row_index = item['row_index']

        # Set current_row in the app and update UI so that extract_info() is called
        self.app.current_row = row_index
        self.app.update_ui_after_navigation()

        # Now self.app.table contains all lines (agreements and disagreements) for this row
        # Also self.app.spec_text_box, self.app.comment_table, etc. are updated

        # For preview, we can just show the spec text and the lines we have in self.app.table.
        spec_text = self.app.spec_text_box.get("1.0", tk.END).strip()
        self.preview_text.insert(tk.END, "Specification Text:\n" + spec_text + "\n\n---\n\n")

        # Insert DI Number Breakdown lines
        self.preview_text.insert(tk.END, "DI Number Breakdown:\n")
        for line_id in self.app.table.get_children():
            values = self.app.table.item(line_id, 'values')
            # Each line is a tuple like (VeriDoc, DI Number, CDRL Subtitle, Object Status, Contractor Status, Government Status)
            self.preview_text.insert(tk.END, ", ".join(str(v) for v in values) + "\n")

        self.preview_text.insert(tk.END, "\n---\n\nComments:\n")
        # Insert comments
        for c_id in self.app.comment_table.get_children():
            c_values = self.app.comment_table.item(c_id, 'values')
            if c_values:
                self.preview_text.insert(tk.END, c_values[0] + "\n")



### This creates the diagreement reports as an excel file. 

    def generate_excel_for_current(self, show_popup=True):
        """
        Creates an Excel file (.xlsx) containing the same data that would normally go into the PDF.
        """
        # 1. Gather the same data as in generate_pdf_for_current
        item = self.disagreement_items[self.current_index]
        row_index = item['row_index']

        # Update UI and load current row data
        self.app.current_row = row_index
        self.app.update_ui_after_navigation()

        # DOORS SPEC ID
        doors_spec_id = self.app.df.iloc[row_index, 0]
        if pd.isna(doors_spec_id):
            doors_spec_id = ""
        elif not isinstance(doors_spec_id, str):
            doors_spec_id = str(doors_spec_id)

        # Spec text
        spec_text = self.app.spec_text_box.get("1.0", "end").strip()

        # Contractor Proposed Change Comment History (col 8)
        contractor_history_content = ""
        if len(self.app.df.columns) > 8:
            val = self.app.df.iloc[row_index, 8]
            if pd.isna(val):
                val = ""
            contractor_history_content = str(val)

        # Government Adjudication Comment History (col 9)
        gov_history_content = ""
        if len(self.app.df.columns) > 9:
            val = self.app.df.iloc[row_index, 9]
            if pd.isna(val):
                val = ""
            gov_history_content = str(val)

        # Table data from self.app.table
        items = self.app.table.get_children()
        breakdown_data = [
            ["VeriDoc Number", "DI Number", "CDRL Subtitle", "Gov. Assessed Status"]
        ]
        for line_id in items:
            values = self.app.table.item(line_id, 'values')
            # (VeriDoc, DI Number, CDRL Subtitle, Object Status, Contractor Status, Government Status)
            # We'll store relevant columns
            breakdown_data.append([values[0], values[1], values[2], values[5]])

        # Count how many are "agree" vs "disagree"
        agree_count = 0
        disagree_count = 0
        for i in range(1, len(breakdown_data)):
            gov_status = str(breakdown_data[i][3]).lower()
            if gov_status == "agree":
                agree_count += 1
            elif gov_status == "disagree":
                disagree_count += 1

        # Extract specifically which lines are "disagree"
        disagreement_rows = []
        for i in range(1, len(breakdown_data)):
            gov_status = str(breakdown_data[i][3]).lower()
            if gov_status == "disagree":
                disagreement_rows.append(breakdown_data[i])

        # 2. Prepare an Excel workbook with openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Disagreement Report"

        # Basic styling references
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # 3. Fill some header info
        current_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A1"] = "Contract: 70Z02323D93270001"
        ws["A2"] = f"Date/Time: {current_dt}"
        ws["A3"] = ("Distribution Statement D: Distribution authorized to DHS/CG/DOD "
                    "and their contractors only due to administrative or operational use.\n"
                    "Other requests shall be referred to COMMANDANT (CG-9327).")
        ws["A4"] = ("DESTRUCTION NOTICE: Destroy this document by any method that "
                    "will prevent disclosure of contents or reconstruction of the document.")
        for row in range(1, 5):
            ws.row_dimensions[row].height = 30  # Make them a bit taller

        # 4. Summary info table: DOORS SPEC, row, total agreements, disagreements
        ws["A6"] = "DOORS SPEC ID"
        ws["B6"] = "Excel Row"
        ws["C6"] = "Total Agreements"
        ws["D6"] = "Total Disagreements"

        ws["A7"] = doors_spec_id
        ws["B7"] = str(row_index + 2)  # +2 for Excel-based row counting
        ws["C7"] = str(agree_count)
        ws["D7"] = str(disagree_count)

        for col in range(1, 5):
            cell = ws.cell(row=6, column=col)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

            cell2 = ws.cell(row=7, column=col)
            cell2.border = thin_border

        # 5. Write out the Specification Text
        spec_start = 9
        ws.cell(spec_start, 1, "Specification Text:")
        ws.cell(spec_start, 1).font = bold_font
        # Put the spec text in the next row
        ws.cell(spec_start+1, 1, spec_text)
        # Optionally wrap text
        ws.cell(spec_start+1, 1).alignment = Alignment(wrap_text=True)

        # 6. Comments
        comment_start = spec_start + 3
        ws.cell(comment_start, 1, "Contractor Proposed Change Comment History:")
        ws.cell(comment_start, 1).font = bold_font

        # We'll split by new lines and write each line in a new row
        c_lines = contractor_history_content.strip().split('\n')
        row_c = comment_start + 1
        for cline in c_lines:
            if not cline.strip() or '_____' in cline:
                continue
            ws.cell(row_c, 1, cline)
            row_c += 1

        # Then government
        row_c += 2
        ws.cell(row_c, 1, "Government Adjudication Comment History:")
        ws.cell(row_c, 1).font = bold_font
        row_c += 1
        g_lines = gov_history_content.strip().split('\n')
        for gline in g_lines:
            if not gline.strip() or '_____' in gline:
                continue
            ws.cell(row_c, 1, gline)
            row_c += 1

        # 7. Breakdown table
        # Let's put it under everything else
        breakdown_start = row_c + 2
        ws.cell(breakdown_start, 1, "Breakdown Data:")
        ws.cell(breakdown_start, 1).font = bold_font
        breakdown_start += 1

        # Insert headers
        for col_idx, header in enumerate(breakdown_data[0], start=1):
            cell = ws.cell(breakdown_start, col_idx, header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # Insert data rows
        for i in range(1, len(breakdown_data)):
            row_data = breakdown_data[i]
            row_num = breakdown_start + i
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row_num, col_idx, val)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border
            
                # Simple color-coding if Government Status is "disagree" or "agree"
                if col_idx == 4:  # Gov. Status
                    if str(val).lower() == "disagree":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF")  # White text
                    elif str(val).lower() == "agree":
                        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                        cell.font = Font(color="FFFFFF")  # White text
                    elif str(val).lower() == "pending review":
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                        cell.font = Font(color="000000")  # Black text

        # 8. Disagreement Comments
        if disagreement_rows:
            c.drawString(left_margin, y, "Disagreement Comments:")
            y -= 20
            for d_row in disagreement_rows:
                veridoc = str(d_row[0]).strip()  # VeriDoc Number from breakdown data
                di_num = str(d_row[1]).strip()
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                c.line(left_margin, y, width - right_margin, y)
                y -= 10
                c.drawString(left_margin, y, f"VeriDoc: {veridoc}")
                y -= 14
                c.drawString(left_margin, y, f"DI Number: {di_num}")
                y -= 28
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                c.drawString(left_margin, y, "Government Comments:")
                y -= 14
                # First try to get the comment from the matrix view (cross-reference)
                matrix_comment = self.get_government_status_comment(veridoc)
                if matrix_comment:
                    y = wrap_text_to_pdf(c, matrix_comment, left_margin, y, usable_width)
                else:
                    # Fall back to filtering gov_lines from the main sheet (using DI Number)
                    related_gov_lines = [gl for gl in gov_lines if di_num in gl]
                    if not related_gov_lines:
                        c.setFillColor(colors.red)
                        c.drawString(left_margin, y, "No specific government comments related to this item.")
                        c.setFillColor(colors.black)
                        y -= 14
                    else:
                        for gl in related_gov_lines:
                            y = wrap_text_to_pdf(c, gl, left_margin, y, usable_width)
                # Add a Birdon Response Comments text field
                y -= 10
                c.drawString(left_margin, y, "Birdon Response Comments:")
                y -= 14
                # Create a text field of height 50 pixels; the name is made unique by appending the VeriDoc ID
                form.textfield(
                    name=f"birdonResponse_{veridoc}",
                    tooltip="Birdon Response Comments",
                    x=left_margin,
                    y=y-50,
                    width=usable_width,
                    height=50,
                    borderStyle="inset",
                    borderWidth=1,
                    fillColor=colors.white,
                )
                y -= 60  # Adjust y for spacing after the text field

        # 9. Optional "form fields" replaced with placeholders
        # Excel doesn't have fillable forms in the same sense, but you can label cells
        form_start = disagree_section + 2
        ws.cell(form_start, 1, "PWG assigned to resolve disagreement:")
        ws.cell(form_start, 2, "")  # user can fill in
        form_start += 1

        ws.cell(form_start, 1, "USCG POC assigned to resolve disagreement:")
        ws.cell(form_start, 2, "")
        form_start += 1

        ws.cell(form_start, 1, "BIRDON POC assigned to resolve disagreement:")
        ws.cell(form_start, 2, "")
        form_start += 2

        ws.cell(form_start, 1, "General Comments:")
        ws.cell(form_start+1, 1, "Enter general comments in the cell below:")
        form_start += 3

        # 10. Finally, save the workbook
        # Decide filename
        filename = f"Disagreement Report - WCC-SPEC-{doors_spec_id}.xlsx"
        # If user selected a folder
        if self.output_folder:
            xlsx_path = os.path.join(self.output_folder, filename)
        else:
            xlsx_path = filename

        try:
            wb.save(xlsx_path)
            if show_popup:
                from tkinter import messagebox
                messagebox.showinfo("Excel Generated", f"Excel saved as {xlsx_path}")
        except Exception as e:
            if show_popup:
                messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")




###This Creates the Diagreement report PDF's 
import os
import re
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import pandas as pd
from pdfrw import PdfReader  # Make sure to install pdfrw: pip install pdfrw
from concurrent.futures import ThreadPoolExecutor

# ReportLab imports for PDF generation
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import acroform
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from textwrap import wrap

def get_pdf_form_fields(pdf_path):
    """
    Reads a PDF file and returns a dictionary of form fields.
    This enhanced version correctly handles checkbox states and provides better error handling.
    """
    fields = {}
    try:
        # Open the PDF with error handling
        pdf = None
        try:
            pdf = PdfReader(pdf_path)
        except Exception as e:
            print(f"Failed to open PDF {os.path.basename(pdf_path)}: {e}")
            return fields  # Return empty fields dictionary if PDF can't be opened
        
        # Check if PDF has form fields
        if not hasattr(pdf, 'Root') or not hasattr(pdf.Root, 'AcroForm') or not pdf.Root.AcroForm or not pdf.Root.AcroForm.Fields:
            # No form fields in PDF
            return fields
        
        # Process each field
        for field in pdf.Root.AcroForm.Fields:
            try:
                # Get field name
                name = ""
                if hasattr(field, 'T'):
                    name = field.T.to_unicode().strip() if hasattr(field.T, 'to_unicode') else str(field.T).strip()
                    if name.startswith('(') and name.endswith(')'):
                        name = name[1:-1]
                if not name:
                    continue  # Skip fields with no name
                
                # Get field value based on field type
                value = None
                
                # The appearance state often indicates checkbox status
                checked = False
                
                # Check for checkbox by examining various properties
                if hasattr(field, 'AS'):
                    # AS property often indicates the appearance state for checkboxes
                    appearance_state = field.AS
                    if appearance_state and (str(appearance_state).strip() not in ['/Off', 'Off']):
                        checked = True
                
                if hasattr(field, 'V'):
                    # V property contains the value
                    if field.V:
                        # Convert to string and handle parentheses
                        raw_value = field.V.to_unicode().strip() if hasattr(field.V, 'to_unicode') else str(field.V).strip()
                        if raw_value.startswith('(') and raw_value.endswith(')'):
                            raw_value = raw_value[1:-1]
                        
                        # Common patterns for checked checkboxes
                        if raw_value in ['/Yes', 'Yes', 'On', '/On', 'True', '/True'] or raw_value == True:
                            checked = True
                            value = True
                        elif raw_value in ['Off', '/Off', 'No', '/No', 'False', '/False'] or not raw_value:
                            value = False
                        else:
                            # For non-checkbox fields, store the actual value
                            value = raw_value
                    else:
                        value = False  # No value typically means unchecked for checkboxes
                
                # For checkboxes, use the checked flag
                if name in ['disagreementNotClear', 'disagreementResolvedLocations', 'disagreementNotResolved']:
                    fields[name] = checked
                else:
                    # For other fields, use the processed value
                    fields[name] = value if value is not None else ""
                    
            except Exception as field_error:
                print(f"Error processing field in {os.path.basename(pdf_path)}: {field_error}")
                continue
                
    except Exception as e:
        print(f"Error reading form fields from {os.path.basename(pdf_path)}: {e}")
    
    return fields

class DisagreementManager:
    def __init__(self, master, app, disagreement_items):
        """
        master: Toplevel window.
        app: instance of your main application.
        disagreement_items: list of dictionaries (filtered so that government_status == "Disagree").
        """
        self.master = master
        self.app = app
        self.disagreement_items = disagreement_items  # Already filtered
        self.output_folder = None   # To be set via "Select Database Location"
        self.report_list = []       # To store info about each generated report

        self.master.title("Disagreement Manager - Batch Reports")
        self.master.geometry("800x600")
        
        # Top frame with buttons
        top_frame = tk.Frame(self.master)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        self.db_button = tk.Button(top_frame, text="Select Database Location", command=self.select_output_folder)
        self.db_button.pack(side=tk.LEFT, padx=5)
        
        self.create_reports_button = tk.Button(top_frame, text="B.1 Create Disagreement Reports", command=self.create_all_reports)
        self.create_reports_button.pack(side=tk.LEFT, padx=5)
        
        self.refresh_button = tk.Button(top_frame, text="Refresh Reports", command=self.refresh_report_table)
        self.refresh_button.pack(side=tk.LEFT, padx=5)
        
        # Table frame
        table_frame = tk.Frame(self.master)
        table_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Updated treeview columns: SpecID, Button1, Button2, Button3, Report_File
        self.report_tree = ttk.Treeview(table_frame, 
                                        columns=("SpecID", "Button1", "Button2", "Button3", "Report_File"),
                                        show="headings")
        self.report_tree.heading("SpecID", text="Spec ID")
        self.report_tree.heading("Button1", text="Button 1")
        self.report_tree.heading("Button2", text="Button 2")
        self.report_tree.heading("Button3", text="Button 3")
        self.report_tree.heading("Report_File", text="Report File")
        
        self.report_tree.column("SpecID", width=100)
        self.report_tree.column("Button1", width=100)
        self.report_tree.column("Button2", width=100)
        self.report_tree.column("Button3", width=100)
        self.report_tree.column("Report_File", width=300)
        
        self.report_tree.bind("<Double-1>", self.on_report_tree_double_click)
        self.report_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.suppress_depreciated_warning = True

    def on_report_tree_double_click(self, event):
        item_id = self.report_tree.focus()
        if item_id:
            values = self.report_tree.item(item_id, "values")
            if values and len(values) >= 5:
                report_file = values[4]
                if os.path.exists(report_file):
                    try:
                        os.startfile(report_file)
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to open file: {e}")


    def refresh_report_table(self):
        """
        Scans the output folder for PDF disagreement reports and then updates the treeview.
        Shows a progress dialog and provides detailed status updates.
        """
        print("Refresh button clicked – refreshing report table...")
        if not self.output_folder:
            print("No output folder set.")
            messagebox.showerror("Error", "Please select a database folder first.")
            return

        # Create progress dialog
        progress_window = tk.Toplevel(self.master)
        progress_window.title("Refreshing Reports")
        progress_window.geometry("400x150")
        progress_window.transient(self.master)
        progress_window.grab_set()  # Make the window modal
    
        # Center the window
        progress_window.update_idletasks()
        width = progress_window.winfo_width()
        height = progress_window.winfo_height()
        x = (progress_window.winfo_screenwidth() - width) // 2
        y = (progress_window.winfo_screenheight() - height) // 2
        progress_window.geometry(f"{width}x{height}+{x}+{y}")
    
        # Progress status label
        status_var = tk.StringVar(value="Scanning for PDF files...")
        status_label = tk.Label(progress_window, textvariable=status_var)
        status_label.pack(pady=(20, 10))
    
        # Progress bar
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100, length=350)
        progress_bar.pack(pady=10, padx=20)
    
        # Detail label
        detail_var = tk.StringVar(value="Finding PDF files...")
        detail_label = tk.Label(progress_window, textvariable=detail_var, font=("Helvetica", 9))
        detail_label.pack(pady=10)
    
        # First, clear the treeview
        for child in self.report_tree.get_children():
            self.report_tree.delete(child)
    
        # Reset report list
        self.report_list = []

        # Use a queue for thread-safe communication
        result_queue = queue.Queue()
    
        def scan_directory():
            """Background thread function to scan for PDF files"""
            try:
                report_files = []
                total_files_scanned = 0
                pdf_files_found = 0
            
                # Get all files in directory tree
                all_files = []
                for root_dir, dirs, files in os.walk(self.output_folder):
                    for file in files:
                        all_files.append((root_dir, file))
            
                total_files = len(all_files)
                print(f"Total files found in directory: {total_files}")
            
                # Process the files
                for root_dir, file in all_files:
                    total_files_scanned += 1
                    # Update progress approximately every 20 files or on last file
                    if total_files_scanned % 20 == 0 or total_files_scanned == total_files:
                        result_queue.put(('progress', (total_files_scanned, total_files, f"Scanning file {total_files_scanned} of {total_files}")))
                
                    # Check if it's a disagreement report PDF
                    if file.endswith(".pdf") and "Disagreement Report - WCC-SPEC-" in file:
                        pdf_files_found += 1
                        print(f"Found report: {file}")
                        result_queue.put(('detail', f"Found report: {file}"))
                    
                        # Extract the SPEC ID
                        m = re.search(r"Disagreement Report - (WCC-SPEC-[^.]+)\.pdf", file)
                        if m:
                            spec_id = m.group(1)
                        else:
                            spec_id = file
                    
                        file_path = os.path.join(root_dir, file)
                        report_files.append({
                            "SpecID": spec_id,
                            "Report_File": file_path
                        })
            
                # Signal completion of scan
                print(f"Scan complete. Found {pdf_files_found} PDF report files.")
                result_queue.put(('scan_complete', (pdf_files_found, report_files)))
            except Exception as e:
                print(f"Error scanning directory: {str(e)}")
                result_queue.put(('error', f"Error scanning directory: {str(e)}"))
    
        def process_fields(report_files):
            """Process the PDF fields in batches for better progress feedback"""
            try:
                total_files = len(report_files)
                print(f"Processing {total_files} PDF files")
                self.report_list = report_files
                result_queue.put(('status', f"Reading form fields from {total_files} PDFs..."))
            
                processed_results = []
                files_processed = 0
                errors = 0
            
                # Process in smaller batches for more frequent updates
                batch_size = 5  # Smaller batch size for better progress updates
                batches = [report_files[i:i + batch_size] for i in range(0, len(report_files), batch_size)]
            
                for batch_idx, batch in enumerate(batches):
                    batch_results = []
                    print(f"Processing batch {batch_idx+1}/{len(batches)}")
                
                    with ThreadPoolExecutor(max_workers=batch_size) as executor:
                        futures = [executor.submit(self.read_pdf_fields, report) for report in batch]
                    
                        for future in as_completed(futures):
                            files_processed += 1
                            progress_percent = files_processed / total_files * 100
                            result_queue.put(('progress', (files_processed, total_files, f"Processing file {files_processed} of {total_files}")))
                        
                            try:
                                result = future.result(timeout=5)  # Timeout after 5 seconds
                                print(f"Successfully processed PDF: {files_processed}")
                                batch_results.append(result)
                            except TimeoutError:
                                errors += 1
                                print(f"Timeout reading PDF #{files_processed}")
                                result_queue.put(('detail', f"Timeout reading PDF #{files_processed}"))
                            except Exception as e:
                                errors += 1
                                print(f"Error reading PDF #{files_processed}: {str(e)}")
                                result_queue.put(('detail', f"Error reading PDF #{files_processed}: {str(e)[:50]}..."))
                
                    processed_results.extend(batch_results)
                
                    # Update status periodically
                    if batch_idx % 2 == 0 or batch_idx == len(batches) - 1:
                        result_queue.put(('status', f"Processed {files_processed}/{total_files} PDFs ({errors} errors)"))
            
                # Signal completion with final results
                print(f"PDF processing complete. {len(processed_results)} processed with {errors} errors.")
                result_queue.put(('complete', (processed_results, errors)))
        
            except Exception as e:
                print(f"Error in process_fields: {str(e)}")
                result_queue.put(('error', f"Error processing PDF fields: {str(e)}"))
    
        def update_progress():
            """Update the progress dialog with queue messages"""
            try:
                # Check if there's a message in the queue
                try:
                    message_type, data = result_queue.get_nowait()
                
                    if message_type == 'progress':
                        current, total, detail_text = data
                        progress_var.set((current / total) * 100 if total > 0 else 0)
                        detail_var.set(detail_text)
                        print(f"Progress update: {detail_text}")
                
                    elif message_type == 'status':
                        status_var.set(data)
                        print(f"Status update: {data}")
                
                    elif message_type == 'detail':
                        detail_var.set(data)
                        print(f"Detail update: {data}")
                
                    elif message_type == 'scan_complete':
                        pdf_count, report_files = data
                        print(f"Scan phase complete. Found {pdf_count} PDFs. Starting processing phase.")
                        status_var.set(f"Found {pdf_count} reports. Processing form fields...")
                        progress_var.set(0)  # Reset progress for next phase
                    
                        # If no PDFs found, we're done
                        if pdf_count == 0:
                            messagebox.showinfo("No Reports", "No PDF reports were found in the selected folder.")
                            progress_window.destroy()
                            return
                    
                        # Start processing PDF fields in a new thread
                        processing_thread = threading.Thread(target=process_fields, args=(report_files,), daemon=True)
                        processing_thread.start()
                
                    elif message_type == 'complete':
                        results, error_count = data
                        progress_var.set(100)  # Ensure progress bar shows 100%
                        status_var.set(f"Completed with {error_count} errors")
                        detail_var.set(f"Refreshing display with {len(results)} reports...")
                        print(f"Processing complete. {len(results)} reports loaded, {error_count} errors.")
                    
                        # Update the treeview with results
                        for item in results:
                            self.report_tree.insert("", "end", values=item)
                    
                        # Close the progress window after a short delay
                        progress_window.after(1000, progress_window.destroy)
                        return  # Exit the update loop
                
                    elif message_type == 'error':
                        error_msg = data
                        status_var.set("Error occurred")
                        detail_var.set(error_msg)
                        print(f"ERROR: {error_msg}")
                        # Allow window to stay open a bit longer to show the error
                        progress_window.after(3000, progress_window.destroy)
                        return  # Exit the update loop
                
                except queue.Empty:
                    # Queue is empty, no messages to process
                    pass
                
                # Schedule another check after a short delay
                progress_window.after(100, update_progress)
            
            except Exception as e:
                print(f"Error in update_progress: {str(e)}")
                messagebox.showerror("Error", f"An error occurred updating progress: {str(e)}")
                progress_window.destroy()
    
        # Start the background scanning thread
        scanning_thread = threading.Thread(target=scan_directory, daemon=True)
        scanning_thread.start()
    
        # Start the progress update loop after a short delay
        progress_window.after(100, update_progress)


    def get_swbs_group(self, detailed_location):
            # Example logic: look for a SWBS keyword in the detailed_location.
            if "SWBS" in detailed_location:
                # Return the first word (adjust logic as needed)
                return detailed_location.split()[0]
            return "Default SWBS"

    def generate_tracking_number(self, item):
        """
        Generate a tracking number from the item.
        If the veridoc_number is something like "WCC-VERI-DOC-2", return just "2".
        """
        veridoc = item.get("veridoc_number", "0000")
        prefix = "WCC-VERI-DOC-"
        if veridoc.startswith(prefix):
            veridoc = veridoc[len(prefix):]  # Strip the prefix
        return veridoc


    def get_government_status_comment(self, veridoc):
        # Example: return an empty string (or implement logic to look up a comment)
        return ""


    def select_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder for Disagreement PDFs")
        if folder:
            self.output_folder = folder
            messagebox.showinfo("Database Location Set", f"Reports will be stored in:\n{folder}")

    # (Other helper methods such as get_swbs_group, extract_detailed_location, etc., remain unchanged.)

    def generate_pdf_for_disagreement_item(self, item, output_folder):
        """
        Generates a PDF disagreement report for a single disagreement item.
        This version retains the full formatting from your previous reports:
         - Distribution and Destruction notices,
         - DOORS SPEC ID summary table,
         - Specification Text,
         - Two-column Comments Table,
         - Breakdown Table,
         - Disagreement Comments Section with a horizontal line and Birdon Response Comments text field under each entry,
         - General Comments field with a horizontal line above,
         - Options for Birdon,
         - USCG Response field, and USCG Signature/Date fields.
        The tracking number is based on the DOORS SPEC ID.
        The PDF is saved into a subfolder (named by the SWBS group) within the selected output folder.
        Returns a tuple: (pdf_path, swbs_group).
        """
        # Set the current row and update UI so that spec_text_box and table data are up-to-date.
        row_index = item['row_index']
        self.app.current_row = row_index
        self.app.update_ui_after_navigation()

        # Get DOORS SPEC ID from the first column.
        doors_spec_id = self.app.df.iloc[row_index, 0]
        if pd.isna(doors_spec_id):
            doors_spec_id = ""
        elif not isinstance(doors_spec_id, str):
            doors_spec_id = str(doors_spec_id)

        # Retrieve the specification text.
        spec_text = self.app.spec_text_box.get("1.0", "end").strip()

        # Get Contractor Proposed Change Comment History (assumed column index 8).
        contractor_history_content = ""
        if len(self.app.df.columns) > 8:
            val = self.app.df.iloc[row_index, 8]
            if pd.isna(val):
                val = ""
            contractor_history_content = str(val)

        # Get Government Adjudication Comment History (assumed column index 9).
        gov_history_content = ""
        if len(self.app.df.columns) > 9:
            val = self.app.df.iloc[row_index, 9]
            if pd.isna(val):
                val = ""
            gov_history_content = str(val)

        # Use the "Assigned Verification Documents" cell to extract the Detailed Location and determine SWBS.
        try:
            row = self.app.df.iloc[row_index]
            assigned_docs = row["Assigned Verification Documents"]
        except Exception:
            assigned_docs = ""
        detailed_location = self.extract_detailed_location(assigned_docs)
        swbs_group = self.get_swbs_group(detailed_location)

        # Determine tracking number and PDF filename.
        tracking_number = self.generate_tracking_number(item)
        filename = f"Disagreement Report - WCC-SPEC-{tracking_number}.pdf"
        # Create a subfolder based on SWBS group.
        target_folder = os.path.join(output_folder, swbs_group)
        os.makedirs(target_folder, exist_ok=True)
        pdf_path = os.path.join(target_folder, filename)

        # Page setup.
        width, height = letter
        left_margin = 72
        right_margin = 72
        top_margin = 50
        bottom_margin = 72
        usable_width = width - (left_margin + right_margin)

        # Build breakdown table data from self.app.table.
        items_ids = self.app.table.get_children()
        breakdown_data = [["VeriDoc Number", "DI Number", "CDRL Subtitle", "Government Assessed Status"]]
        for line_id in items_ids:
            values = self.app.table.item(line_id, 'values')
            breakdown_data.append([values[0], values[1], values[2], values[5]])

        # Count agreements and disagreements.
        agree_count = 0
        disagree_count = 0
        for i in range(1, len(breakdown_data)):
            gov_status = str(breakdown_data[i][3]).strip().lower()
            if gov_status == "agree":
                agree_count += 1
            elif gov_status == "disagree":
                disagree_count += 1

        # Extract rows with disagreement.
        disagreement_rows = []
        for i in range(1, len(breakdown_data)):
            if str(breakdown_data[i][3]).strip().lower() == "disagree":
                disagreement_rows.append(breakdown_data[i])

        # Set up the canvas and acroform.
        c = canvas.Canvas(pdf_path, pagesize=letter)
        form = acroform.AcroForm(c)
        styles = getSampleStyleSheet()
        styleN = styles["Normal"]

        def wrap_text_to_pdf(c, text, x, y, max_width):
            chars_per_line = int(max_width / 6)  # Approximation for 12pt font
            wrapped_lines = wrap(text, width=chars_per_line)
            for wline in wrapped_lines:
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                c.drawString(x, y, wline)
                y -= 14
            return y

        # Write header information.
        c.setFont("Helvetica", 8)
        y = height - top_margin
        current_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.drawString(left_margin, y, f"Date/Time: {current_dt}")
        y -= 10
        c.drawString(left_margin, y, "Contract: 70Z02323D93270001")
        y -= 10

        distribution_text = (
            "DISTRIBUTION STATEMENT D: DISTRIBUTION AUTHORIZED TO DHS/CG/DOD AND THEIR "
            "CONTRACTORS ONLY DUE TO ADMINISTRATIVE OR OPERATIONAL USE (5 OCT 2022). "
            "OTHER REQUESTS SHALL BE REFERRED TO COMMANDANT (CG-9327)."
        )
        destruction_text = (
            "DESTRUCTION NOTICE: DESTROY THIS DOCUMENT BY ANY METHOD THAT WILL "
            "PREVENT DISCLOSURE OF CONTENTS OR RECONSTRUCTION OF THE DOCUMENT."
        )
        y = wrap_text_to_pdf(c, distribution_text, left_margin, y, usable_width)
        y -= 10
        y = wrap_text_to_pdf(c, destruction_text, left_margin, y, usable_width)

        # DOORS SPEC ID Summary Table.
        id_table_data = [
            ["DOORS SPEC ID", "Excel Row", "Total Agreements", "Total Disagreements"],
            [doors_spec_id, str(row_index + 2), str(agree_count), str(disagree_count)],
        ]
        id_table = Table(id_table_data, colWidths=[130, 60, 100, 120])
        id_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        id_table.setStyle(id_style)
        w_id, h_id = id_table.wrap(usable_width, 50)
        c.setFont("Helvetica", 12)
        if y - h_id < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        id_table.drawOn(c, left_margin, y - h_id)
        y = y - h_id - 30

        # Specification Text Section.
        c.drawString(left_margin, y, "Specification Text:")
        y -= 20
        y = wrap_text_to_pdf(c, spec_text, left_margin, y, usable_width)

        # Two-column Comments Table for Contractor vs. Government.
        y -= 30
        c.drawString(left_margin, y, "Comments:")
        y -= 20
        contractor_lines = [line.strip() for line in contractor_history_content.split("\n") if line.strip() and "_____" not in line]
        gov_lines = [line.strip() for line in gov_history_content.split("\n") if line.strip() and "_____" not in line]
        comments_data = [["Contractor Proposed Change Comment History", "Government Adjudication Comment History"]]
        max_len = max(len(contractor_lines), len(gov_lines))
        for i in range(max_len):
            c_text = contractor_lines[i] if i < len(contractor_lines) else ""
            g_text = gov_lines[i] if i < len(gov_lines) else ""
            comments_data.append([Paragraph(c_text, styleN), Paragraph(g_text, styleN)])
        comments_table = Table(comments_data, colWidths=[usable_width / 2, usable_width / 2])
        comments_table_style = TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ])
        comments_table.setStyle(comments_table_style)
        w_comments, h_comments = comments_table.wrap(usable_width, y)
        if y - h_comments < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        comments_table.drawOn(c, left_margin, y - h_comments)
        y -= h_comments + 20

        # Breakdown Table.
        approx_char_width = 6
        max_lengths = [0, 0, 0, 0]
        for row_val in breakdown_data:
            for j, val in enumerate(row_val):
                length = len(str(val))
                if length > max_lengths[j]:
                    max_lengths[j] = length
        column_widths = [length * approx_char_width for length in max_lengths]
        if column_widths[2] > 200:
            column_widths[2] = 200
        for i in range(1, len(breakdown_data)):
            cdrl_text = breakdown_data[i][2]
            breakdown_data[i][2] = Paragraph(cdrl_text, styleN)
        t = Table(breakdown_data, colWidths=column_widths)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        for i in range(1, len(breakdown_data)):
            gov_status = str(breakdown_data[i][3]).strip().lower()
            if gov_status == "disagree":
                style.add("BACKGROUND", (3, i), (3, i), colors.red)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.white)
            elif gov_status == "agree":
                style.add("BACKGROUND", (3, i), (3, i), colors.green)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.white)
            elif gov_status == "pending review":
                style.add("BACKGROUND", (3, i), (3, i), colors.yellow)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.black)
        t.setStyle(style)
        w, h = t.wrap(usable_width, 400)
        if y - h < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        t.drawOn(c, left_margin, y - h)
        y = y - h - 60

        # Disagreement Comments Section (with Birdon Response Comments text fields).
        if disagreement_rows:
            c.drawString(left_margin, y, "Disagreement Comments:")
            y -= 20
            for d_row in disagreement_rows:
                veridoc = str(d_row[0]).strip()  # VeriDoc Number from breakdown data
                di_num = str(d_row[1]).strip()
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                # Draw a horizontal line for separation
                c.line(left_margin, y, width - right_margin, y)
                y -= 10
                c.drawString(left_margin, y, f"VeriDoc: {veridoc}")
                y -= 14
                c.drawString(left_margin, y, f"DI Number: {di_num}")
                y -= 28
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                c.drawString(left_margin, y, "Government Comments:")
                y -= 14
                matrix_comment = self.get_government_status_comment(veridoc)
                if matrix_comment:
                    y = wrap_text_to_pdf(c, matrix_comment, left_margin, y, usable_width)
                else:
                    related_gov_lines = [gl for gl in gov_lines if di_num in gl]
                    if not related_gov_lines:
                        c.setFillColor(colors.red)
                        c.drawString(left_margin, y, "No specific government comments related to this item.")
                        c.setFillColor(colors.black)
                        y -= 14
                    else:
                        for gl in related_gov_lines:
                            y = wrap_text_to_pdf(c, gl, left_margin, y, usable_width)
                # Add Birdon Response Comments text field for this disagreement.
                y -= 10
                c.drawString(left_margin, y, "Birdon Response Comments:")
                y -= 14
                form.textfield(
                    name=f"birdonResponse_{veridoc}",
                    tooltip="Birdon Response Comments",
                    x=left_margin,
                    y=y - 50,
                    width=usable_width,
                    height=50,
                    borderStyle="inset",
                    borderWidth=1,
                    fillColor=colors.white,
                )
                y -= 60
        else:
            c.drawString(left_margin, y, "No items are marked 'Disagree' in this row.")
            y -= 20

        # General Comments Text Field with horizontal line above.
        single_line_height = 20
        if y < 200:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        c.line(left_margin, y, width - right_margin, y)
        y -= 10
        c.drawString(left_margin, y, "General Comments:")
        general_box_width = usable_width
        general_box_height = 80
        form.textfield(
            name="generalComments",
            tooltip="General Comments",
            x=left_margin,
            y=y - general_box_height,
            width=general_box_width,
            height=general_box_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= (general_box_height + 40)

        # Options for Birdon (checkboxes)
        c.drawString(left_margin, y, "Options for Birdon")
        y -= 20
        c.drawString(left_margin, y, "Disagreement Not Clear - Send to USCG for Clarification")
        form.checkbox(
            name="disagreementNotClear",
            tooltip="Check if the disagreement is not clear and needs USCG clarification",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 20
        c.drawString(left_margin, y, "Disagreement can be resolved with updated locations flag for RTVM")
        form.checkbox(
            name="disagreementResolvedLocations",
            tooltip="Check if the disagreement can be resolved with updated locations in RTVM",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 20
        c.drawString(left_margin, y, "Disagreement can not be resolved at this time")
        form.checkbox(
            name="disagreementNotResolved",
            tooltip="Check if the disagreement cannot be resolved at this time",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 40

        # USCG Response text field
        c.drawString(left_margin, y, "USCG Response:")
        uscg_box_height = 60
        form.textfield(
            name="uscgResponceBox",
            tooltip="Enter USCG Response",
            x=left_margin,
            y=y - uscg_box_height,
            width=usable_width,
            height=uscg_box_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= (uscg_box_height + 40)

        # USCG Signature and Date of Resolution
        c.drawString(left_margin, y, "USCG Signature (approved to disregard disagreement):")
        form.textfield(
            name="uscgSignature",
            tooltip="USCG Signature",
            x=left_margin + 364,
            y=y - 12,
            width=150,
            height=single_line_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= 40
        c.drawString(left_margin, y, "Date of Resolution:")
        form.textfield(
            name="resolutionDate",
            tooltip="Date of Resolution",
            x=left_margin + 228,
            y=y - 12,
            width=200,
            height=single_line_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= 40

        c.showPage()
        c.save()

        return pdf_path, swbs_group

    def extract_detailed_location(self, assigned_docs):
        # Example: simply return the assigned_docs as a string, or a default value if empty.
        return str(assigned_docs) if assigned_docs else "Location Not Provided"

    def create_all_reports(self):
        if not self.disagreement_items:
            messagebox.showinfo("No Disagreements", "No disagreement items available.")
            return
        if not self.output_folder:
            messagebox.showerror("Output Folder Not Set", "Please select a database folder first.")
            return

        # Group disagreement items by DOORS SPEC ID instead of VeriDoc number
        grouped_by_spec_id = {}
        for item in self.disagreement_items:
            row_index = item['row_index']
        
            # Get the DOORS SPEC ID from the first column of the DataFrame
            spec_id = self.app.df.iloc[row_index, 0]
            if pd.isna(spec_id):
                spec_id = "Unknown"
            elif not isinstance(spec_id, str):
                spec_id = str(spec_id)
            
            # Add this item to the appropriate group
            if spec_id not in grouped_by_spec_id:
                grouped_by_spec_id[spec_id] = []
            grouped_by_spec_id[spec_id].append(item)

        self.report_list = []
    
        # Process each DOORS SPEC ID group
        for spec_id, items in grouped_by_spec_id.items():
            # Generate a single report for each DOORS SPEC ID
            pdf_path = self.generate_pdf_for_spec_id(spec_id, items, self.output_folder)
            if pdf_path:
                report_info = {
                    "SpecID": spec_id,
                    "Report_File": pdf_path
                }
                self.report_list.append(report_info)

        self.update_report_table()
        messagebox.showinfo("Reports Created", f"{len(self.report_list)} disagreement reports have been generated.")

    def generate_pdf_for_spec_id(self, spec_id, items, output_folder):
        """
        Generates a single PDF disagreement report for a DOORS SPEC ID group.
        All VeriDoc items belonging to this SPEC ID will be included in one report.
        """
        # Choose a representative item to get the row data, just to initialize
        rep_item = items[0]
        row_index = rep_item['row_index']
    
        # Set the current row in the app and update UI
        self.app.current_row = row_index
        self.app.update_ui_after_navigation()
    
        # Create filename based on DOORS SPEC ID
        # Extract just the numeric part if needed
        match = re.search(r'WCC-SPEC-(\d+)', spec_id)
        tracking_number = match.group(1) if match else spec_id
        filename = f"Disagreement Report - WCC-SPEC-{tracking_number}.pdf"
    
        # Get detailed location to determine SWBS group
        try:
            row = self.app.df.iloc[row_index]
            assigned_docs = row["Assigned Verification Documents"]
        except Exception:
            assigned_docs = ""
    
        detailed_location = self.extract_detailed_location(assigned_docs)
        swbs_group = self.get_swbs_group(detailed_location)
    
        # Create a subfolder based on SWBS group
        target_folder = os.path.join(output_folder, swbs_group)
        os.makedirs(target_folder, exist_ok=True)
        pdf_path = os.path.join(target_folder, filename)
    
        # Get specification text and other content from the DataFrame
        spec_text = self.app.spec_text_box.get("1.0", "end").strip()
    
        # Get Contractor Proposed Change Comment History (column index 8)
        contractor_history_content = ""
        if len(self.app.df.columns) > 8:
            val = self.app.df.iloc[row_index, 8]
            if pd.isna(val):
                val = ""
            contractor_history_content = str(val)

        # Get Government Adjudication Comment History (column index 9)
        gov_history_content = ""
        if len(self.app.df.columns) > 9:
            val = self.app.df.iloc[row_index, 9]
            if pd.isna(val):
                val = ""
            gov_history_content = str(val)
    
        # Page setup
        width, height = letter
        left_margin = 72
        right_margin = 72
        top_margin = 50
        bottom_margin = 72
        usable_width = width - (left_margin + right_margin)
    
        # Create the PDF document
        c = canvas.Canvas(pdf_path, pagesize=letter)
        form = acroform.AcroForm(c)
        styles = getSampleStyleSheet()
        styleN = styles["Normal"]
    
        # Helper function for wrapping text
        def wrap_text_to_pdf(c, text, x, y, max_width):
            chars_per_line = int(max_width / 6)  # Approx for 12pt font
            wrapped_lines = wrap(text, width=chars_per_line)
            for wline in wrapped_lines:
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
                c.drawString(x, y, wline)
                y -= 14
            return y
    
        # Write header information
        c.setFont("Helvetica", 8)
        y = height - top_margin
        current_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.drawString(left_margin, y, f"Date/Time: {current_dt}")
        y -= 10
        c.drawString(left_margin, y, "Contract: 70Z02323D93270001")
        y -= 10
    
        distribution_text = (
            "DISTRIBUTION STATEMENT D: DISTRIBUTION AUTHORIZED TO DHS/CG/DOD AND THEIR "
            "CONTRACTORS ONLY DUE TO ADMINISTRATIVE OR OPERATIONAL USE (5 OCT 2022). "
            "OTHER REQUESTS SHALL BE REFERRED TO COMMANDANT (CG-9327)."
        )
        destruction_text = (
            "DESTRUCTION NOTICE: DESTROY THIS DOCUMENT BY ANY METHOD THAT WILL "
            "PREVENT DISCLOSURE OF CONTENTS OR RECONSTRUCTION OF THE DOCUMENT."
        )
        y = wrap_text_to_pdf(c, distribution_text, left_margin, y, usable_width)
        y -= 10
        y = wrap_text_to_pdf(c, destruction_text, left_margin, y, usable_width)
    
        # Build data for the main table by looking at all items in the group
        all_breakdown_rows = []
    
        # Track the total agrees and disagrees
        agree_count = 0
        disagree_count = 0
    
        # Collect all VeriDoc entries from the table
        # We need to manually rebuild the table since we're switching the row
        # Use a dictionary to track unique entries by VeriDoc number
        unique_entries = {}
    
        for item in items:
            row_index = item['row_index']
            # Temporarily set the current row to get the table data
            self.app.current_row = row_index
            self.app.update_ui_after_navigation()
        
            # Get data from the table
            for line_id in self.app.table.get_children():
                values = self.app.table.item(line_id, 'values')
                veridoc = values[0]
                di_number = values[1]
                cdrl_subtitle = values[2]
                gov_status = values[5]
            
                # Store only unique entries using veridoc as key
                if veridoc not in unique_entries:
                    unique_entries[veridoc] = [veridoc, di_number, cdrl_subtitle, gov_status]
                
                    # Track agreement counts
                    if gov_status.lower() == "agree":
                        agree_count += 1
                    elif gov_status.lower() == "disagree":
                        disagree_count += 1
    
        # Convert the dictionary of unique entries to a list for the table
        all_breakdown_rows = list(unique_entries.values())
    
        # Return to the representative row for consistency
        self.app.current_row = row_index
        self.app.update_ui_after_navigation()
    
        # DOORS SPEC ID Summary Table
        id_table_data = [
            ["DOORS SPEC ID", "Excel Row", "Total Agreements", "Total Disagreements"],
            [spec_id, str(row_index + 2), str(agree_count), str(disagree_count)],
        ]
        id_table = Table(id_table_data, colWidths=[130, 60, 100, 120])
        id_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        id_table.setStyle(id_style)
        w_id, h_id = id_table.wrap(usable_width, 50)
        c.setFont("Helvetica", 12)
        if y - h_id < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        id_table.drawOn(c, left_margin, y - h_id)
        y = y - h_id - 30
    
        # Specification Text Section
        c.drawString(left_margin, y, "Specification Text:")
        y -= 20
        y = wrap_text_to_pdf(c, spec_text, left_margin, y, usable_width)
    
        # Comments Table
        y -= 30
        c.drawString(left_margin, y, "Comments:")
        y -= 20
        contractor_lines = [line.strip() for line in contractor_history_content.split("\n") if line.strip() and "_____" not in line]
        gov_lines = [line.strip() for line in gov_history_content.split("\n") if line.strip() and "_____" not in line]
        comments_data = [["Contractor Proposed Change Comment History", "Government Adjudication Comment History"]]
        max_len = max(len(contractor_lines), len(gov_lines))
        for i in range(max_len):
            c_text = contractor_lines[i] if i < len(contractor_lines) else ""
            g_text = gov_lines[i] if i < len(gov_lines) else ""
            comments_data.append([Paragraph(c_text, styleN), Paragraph(g_text, styleN)])
        comments_table = Table(comments_data, colWidths=[usable_width / 2, usable_width / 2])
        comments_table_style = TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ])
        comments_table.setStyle(comments_table_style)
        w_comments, h_comments = comments_table.wrap(usable_width, y)
        if y - h_comments < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        comments_table.drawOn(c, left_margin, y - h_comments)
        y -= h_comments + 20
    
        # Breakdown Table
        breakdown_data = [["VeriDoc Number", "DI Number", "CDRL Subtitle", "Government Assessed Status"]]
        breakdown_data.extend(all_breakdown_rows)
    
        # Calculate column widths
        approx_char_width = 6
        max_lengths = [0, 0, 0, 0]
        for row_val in breakdown_data:
            for j, val in enumerate(row_val):
                length = len(str(val))
                if length > max_lengths[j]:
                    max_lengths[j] = length
        column_widths = [min(length * approx_char_width, 200) for length in max_lengths]
    
        # Convert long text to paragraphs
        for i in range(1, len(breakdown_data)):
            cdrl_text = breakdown_data[i][2]
            breakdown_data[i][2] = Paragraph(cdrl_text, styleN)
    
        # Create the breakdown table
        t = Table(breakdown_data, colWidths=column_widths)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    
        # Color code the government status cells
        for i in range(1, len(breakdown_data)):
            gov_status = str(breakdown_data[i][3]).strip().lower()
            if gov_status == "disagree":
                style.add("BACKGROUND", (3, i), (3, i), colors.red)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.white)
            elif gov_status == "agree":
                style.add("BACKGROUND", (3, i), (3, i), colors.green)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.white)
            elif gov_status == "pending review":
                style.add("BACKGROUND", (3, i), (3, i), colors.yellow)
                style.add("TEXTCOLOR", (3, i), (3, i), colors.black)
    
        t.setStyle(style)
        w, h = t.wrap(usable_width, 400)
        if y - h < bottom_margin:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        t.drawOn(c, left_margin, y - h)
        y = y - h - 30
    
        # Disagreement Comments Section for each disagreed item
        disagreement_rows = [row for row in all_breakdown_rows if row[3].lower() == "disagree"]
    
        if disagreement_rows:
            c.drawString(left_margin, y, "Disagreement Comments:")
            y -= 20
            for row in disagreement_rows:
                veridoc = row[0]
                di_number = row[1]
            
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
            
                # Draw a horizontal line for separation
                c.line(left_margin, y, width - right_margin, y)
                y -= 10
                c.drawString(left_margin, y, f"VeriDoc: {veridoc}")
                y -= 14
                c.drawString(left_margin, y, f"DI Number: {di_number}")
                y -= 28
            
                if y < bottom_margin:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - top_margin
            
                c.drawString(left_margin, y, "Government Comments:")
                y -= 14
            
                # Try to find government comments related to this VeriDoc or DI Number
                matrix_comment = self.get_government_status_comment(veridoc)
                if matrix_comment:
                    y = wrap_text_to_pdf(c, matrix_comment, left_margin, y, usable_width)
                else:
                    related_gov_lines = [gl for gl in gov_lines if di_number in gl]
                    if not related_gov_lines:
                        c.setFillColor(colors.red)
                        c.drawString(left_margin, y, "No specific government comments related to this item.")
                        c.setFillColor(colors.black)
                        y -= 14
                    else:
                        for gl in related_gov_lines:
                            y = wrap_text_to_pdf(c, gl, left_margin, y, usable_width)
            
                # Add Birdon Response Comments text field
                y -= 10
                c.drawString(left_margin, y, "Birdon Response Comments:")
                y -= 14
            
                form.textfield(
                    name=f"birdonResponse_{veridoc}",
                    tooltip="Birdon Response Comments",
                    x=left_margin,
                    y=y - 50,
                    width=usable_width,
                    height=50,
                    borderStyle="inset",
                    borderWidth=1,
                    fillColor=colors.white,
                )
                y -= 60
        else:
            c.drawString(left_margin, y, "No items are marked 'Disagree' in this row.")
            y -= 20
    
        # General Comments field
        # Make sure there's enough space for the general comments section (title + field + margin)
        required_space = 150  # 10px for line + 10px for title + 80px for field + 50px margin
        if y < bottom_margin + required_space:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
    
        c.line(left_margin, y, width - right_margin, y)
        y -= 10
        c.drawString(left_margin, y, "General Comments:")
        y -= 20
    
        form.textfield(
            name="generalComments",
            tooltip="General Comments",
            x=left_margin,
            y=y - 80,
            width=usable_width,
            height=80,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= (80 + 40)
    
        # Options for Birdon (checkboxes)
        # Check if we need a new page for the options section
        required_space = 120  # Approximately 20px per line × 6 lines
        if y < bottom_margin + required_space:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        
        c.drawString(left_margin, y, "Options for Birdon")
        y -= 20
        c.drawString(left_margin, y, "Disagreement Not Clear - Send to USCG for Clarification")
        form.checkbox(
            name="disagreementNotClear",
            tooltip="Check if the disagreement is not clear and needs USCG clarification",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 20
        c.drawString(left_margin, y, "Disagreement can be resolved with updated locations flag for RTVM")
        form.checkbox(
            name="disagreementResolvedLocations",
            tooltip="Check if the disagreement can be resolved with updated locations in RTVM",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 20
        c.drawString(left_margin, y, "Disagreement can not be resolved at this time")
        form.checkbox(
            name="disagreementNotResolved",
            tooltip="Check if the disagreement cannot be resolved at this time",
            x=left_margin + 420,
            y=y - 8,
            size=12,
            borderWidth=1,
            checked=False,
            buttonStyle="check",
        )
        y -= 40
    
        # USCG Response field
        # Check if we need a page break for the USCG response section
        required_space = 110  # Text label + text field + margin
        if y < bottom_margin + required_space:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        
        c.drawString(left_margin, y, "USCG Response:")
        y -= 20
    
        form.textfield(
            name="uscgResponceBox",
            tooltip="Enter USCG Response",
            x=left_margin,
            y=y - 60,
            width=usable_width,
            height=60,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= (60 + 40)
    
        # USCG Signature and Date
        # Check if we need a page break for the signature and date section
        required_space = 100  # For both signature and date fields + margin
        if y < bottom_margin + required_space:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
    
        single_line_height = 20
        c.drawString(left_margin, y, "USCG Signature (approved to disregard disagreement):")
        form.textfield(
            name="uscgSignature",
            tooltip="USCG Signature",
            x=left_margin + 364,
            y=y - 12,
            width=150,
            height=single_line_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
        y -= 40
    
        # Make sure date of resolution doesn't go off the page
        if y < bottom_margin + 40:  # Need at least 40px for the date field
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - top_margin
        
        c.drawString(left_margin, y, "Date of Resolution:")
        form.textfield(
            name="resolutionDate",
            tooltip="Date of Resolution",
            x=left_margin + 228,
            y=y - 12,
            width=200,
            height=single_line_height,
            borderStyle="inset",
            borderWidth=1,
            fillColor=colors.white,
        )
    
        # Make sure we have adequate margin at the bottom of the last page
        y -= (single_line_height + 30)
        if y < bottom_margin:
            c.showPage()
    
        # Save the PDF
        c.showPage()
        c.save()
    
        return pdf_path




    def update_report_table(self):
        """
        Reads the PDF form fields for each report in self.report_list concurrently
        and updates the treeview with the checkbox states.
        """
        print("Updating report table with generated report list...")
        
        # Clear existing items in the treeview
        for child in self.report_tree.get_children():
            self.report_tree.delete(child)
        
        # Update column headers to meaningful names
        self.report_tree.heading("Button1", text="Not Clear")
        self.report_tree.heading("Button2", text="Can Resolve")
        self.report_tree.heading("Button3", text="Cannot Resolve")

        def read_pdf_fields(self, report):
            """
            Read fields from a single PDF report and handle potential errors
            """
            try:
                pdf_path = report["Report_File"]
                print(f"Reading fields from: {os.path.basename(pdf_path)}")
        
                fields = get_pdf_form_fields(pdf_path)
        
                # Get checkbox states
                not_clear = "✓" if fields.get("disagreementNotClear", False) else ""
                can_resolve = "✓" if fields.get("disagreementResolvedLocations", False) else ""
                cannot_resolve = "✓" if fields.get("disagreementNotResolved", False) else ""
        
                # Log the fields for debugging
                print(f"Fields for {os.path.basename(pdf_path)}: not_clear={not_clear}, can_resolve={can_resolve}, cannot_resolve={cannot_resolve}")
        
                return (report["SpecID"], not_clear, can_resolve, cannot_resolve, pdf_path)
            except Exception as e:
                print(f"Error reading PDF fields from {report['Report_File']}: {str(e)}")
                # Return a default result rather than failing
                return (report["SpecID"], "", "", "", report["Report_File"])

        results = []
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(read_pdf_fields, report) for report in self.report_list]
            for future in futures:
                try:
                    # Use a timeout to prevent a single PDF from hanging
                    result = future.result(timeout=10)
                    results.append(result)
                except TimeoutError:
                    print("Timeout reading a PDF file; skipping it.")
                except Exception as e:
                    print(f"Error reading PDF fields: {e}")

        # Now update the treeview on the main thread
        def update_treeview():
            for item in results:
                self.report_tree.insert("", "end", values=item)
        self.master.after(0, update_treeview)

    #################################################################################################END END END ##########################################################

if __name__ == "__main__":
    root = tk.Tk()
    app = RTVMApp(root)
    root.mainloop()
    
